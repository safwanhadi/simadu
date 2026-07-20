from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.db import transaction
from django.views import View, generic
# from django.views.generic.edit import FormView
from django.db.models.query import QuerySet
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.db.models import Q, Min, Max, Value, CharField, Count, Case, When, F, OuterRef, Subquery, IntegerField, Exists, Prefetch, FilteredRelation
from django.db.models.functions import TruncDate, Coalesce
from django.contrib.staticfiles import finders
from django.db.models.fields import TimeField
# from django.db.models.functions import TruncMonth, ExtractMonth, TruncDate, TruncYear
from django.utils.functional import cached_property
from django.http import HttpResponse, HttpResponseRedirect
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse
from datetime import date, datetime, time, timedelta
from calendar import monthrange, month_name
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.http import HttpResponse
from collections import defaultdict
from decimal import Decimal
from openpyxl.drawing.image import Image as XLImage
import requests
from django.views.generic import ListView, TemplateView
from django.conf import settings
from django.db import IntegrityError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from .models import MappingMesinAbsensi, LogKehadiran

from django.views.generic import FormView
from django.http import JsonResponse
from .services import BridgeSyncService
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
import json
import os

from .services import KehadiranService, ApelPagiService, AttendanceOrchestrator, NewAttendanceOrchestrator

from django.utils import timezone
from .utils import get_mingguan_lengkap, hitung_total_jam, is_user_authorized_to_approve
from openpyxl import load_workbook
import calendar
import qrcode
from io import BytesIO
import base64
from PIL import Image
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


from strukturorg.models import SatuanKerjaInduk, UnitOrganisasi, Bidang, SubBidang, UnitInstalasi

from .models import (
    JenisSDMPerinstalasi, 
    JadwalDinasSDM, 
    ApprovedJadwalDinasSDM,
    DaftarKegiatanPegawai, 
    AlasanTidakHadir, 
    KehadiranKegiatan, 
    JenisKegiatan,
    DetailKategoriJadwalDinas,
    HariLibur,
    AturanToleransiKeterlambatan,
)
from dokumen.models import RiwayatPenempatan, RiwayatJabatan
from strukturorg.models import UnitInstalasi
from myaccount.models import Users, ProfilSDM
from .forms import (
    HariLiburForm,
    jadwal_formset, 
    update_jadwal_formset,
    jumlah_hari_dalam_bulan, 
    JenisSDMPerinstalasiBasicForm,
    JenisSDMPerinstalasiForm, 
    JenisSDMPerinstalasiCustomForm, 
    DaftarKegiatanPegawaiForm,
    kehadiran_formset,
    UploadFingerprintForm,
    FormCopyJadwalSDM,
    SalinJadwalForm,
    SalinJadwalInstalasiForm,
    SearchForm,
    PengajuanJadwalForm,
    PersetujuanForm,
    ProsesKehadiranForm,
    PenilaianKehadiranForm,
    LogAktivitasFormSet,
    )

# Create your views here.
notfoundview = 'riwayat_urls:notfound_view'

def get_nip(user):
    try:
        nip = user.profil_user.nip
        return nip
    except Exception:
        return None
    
def get_date_from_string(tanggal: str):
    # 1. PENJAGA GERBANG (GUARD CLAUSE)
    if not tanggal or not tanggal.strip():
        return date.today()

    # 2. BLOK UTAMA
    try:
        return datetime.strptime(tanggal, "%Y-%m-%d").date()
    except ValueError:
        return date.today()
    
    
def get_day_in_a_month():
    sekarang = date.today().replace(day=1)
    tanggal = [sekarang + relativedelta(day=i) for i in range(jumlah_hari_dalam_bulan())]
    return tanggal


def get_evaluasi_tabel(inst_id, users, bulan=None, tahun=None):
    today = date.today()
    bulan = bulan or today.month
    tahun = tahun or today.year
    jumlah_hari = monthrange(tahun, bulan)[1]

    try:
        instalasi = UnitInstalasi.objects.get(pk=inst_id)
    except UnitInstalasi.DoesNotExist:
        return {
            'error': f'Instalasi dengan ID {inst_id} tidak ditemukan.'
        }

    # semua_pegawai = Users.objects.filter(instalasi=instalasi)

    # Ambil semua JenisSDMPerinstalasi bulan ini
    jenis_qs = JenisSDMPerinstalasi.objects.filter(
        bulan=bulan, tahun=tahun, instalasi=instalasi
    ).select_related('pegawai')

    jenis_map = {item.pegawai_id: item for item in jenis_qs}

    # Hitung jumlah jadwal per pegawai
    jadwal_count_qs = JadwalDinasSDM.objects.filter(
        tanggal__month=bulan,
        tanggal__year=tahun,
        pegawai__instalasi=instalasi
    ).values('pegawai__pegawai_id').annotate(jumlah=Count('id'))
    
    jadwal_map = {item['pegawai__pegawai_id']: item['jumlah'] for item in jadwal_count_qs}

    # Susun data per pegawai
    data_tabel = []
    for idx, peg in enumerate(users, start=1):
        terdaftar = peg.id in jenis_map
        jumlah_jadwal = jadwal_map.get(peg.id, 0)

        if not terdaftar:
            status = "❌ Belum terdaftar di JenisSDM"
        elif jumlah_jadwal == 0:
            status = "⚠️ Belum ada jadwal"
        elif jumlah_jadwal < jumlah_hari:
            status = "⚠️ Jadwal belum lengkap"
        else:
            status = "✅ Jadwal lengkap"

        data_tabel.append({
            'no': idx,
            'nama': peg.full_name,
            'terdaftar': "✅ Ya" if terdaftar else "❌ Tidak",
            'jumlah_jadwal': jumlah_jadwal,
            'status': status
        })

    return {
        'instalasi': instalasi.instalasi,
        'bulan': bulan,
        'tahun': tahun,
        'jumlah_hari': jumlah_hari,
        'data': data_tabel
    }


def get_pimpinan_id():
    pimpinan_ids = Users.objects.filter(
        Q(id__in=SatuanKerjaInduk.objects.values_list('nama_pimpinan_id', flat=True)) |
        Q(id__in=UnitOrganisasi.objects.values_list('nama_pimpinan_id', flat=True)) |
        Q(id__in=Bidang.objects.values_list('nama_pimpinan_id', flat=True)) |
        Q(id__in=SubBidang.objects.values_list('nama_pimpinan_id', flat=True))
    ).exclude(is_active=False).values_list('id', flat=True)
    
    return pimpinan_ids

class EvaluasiJadwal(LoginRequiredMixin, UserPassesTestMixin, generic.TemplateView):
    template_name = 'jadwal_piket/evaluasijadwal_list.html'
    paginate_by = 20

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_disiplin_admin

    def handle_no_permission(self):
        messages.error(self.request, 'Anda tidak memiliki izin untuk melihat menu ini.')
        return redirect(reverse('disiplinsdm_urls:jadwal_list'))

    def get_instalasi_queryset(self):
        user = self.request.user
        qs = UnitInstalasi.objects.all()
        if user.is_disiplin_admin:
            return qs
        profil = getattr(user, 'profil_admin', None)
        if profil:
            if profil.instalasi.exists():
                return qs.filter(pk__in=profil.instalasi.values_list('pk', flat=True))
            if profil.sub_bidang.exists():
                return qs.filter(sub_bidang__in=profil.sub_bidang.values_list('pk', flat=True))
            if profil.bidang.exists():
                return qs.filter(sub_bidang__bidang__in=profil.bidang.values_list('pk', flat=True))
        return qs.none()

    def get_inst_id(self, instalasi_qs):
        get = self.request.GET.get
        inst_id = get('inst')
        if not inst_id and instalasi_qs.exists():
            inst_id = str(instalasi_qs.first().pk)
        return inst_id

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        get = self.request.GET.get
        tanggal = date.today()
        bulan = int(get('bulan') or tanggal.month)
        tahun = int(get('tahun') or tanggal.year)

        # Ambil daftar instalasi sesuai role
        instalasi_qs = self.get_instalasi_queryset()

        # Ambil inst_id dari GET atau fallback ke instalasi pertama
        inst_id = self.get_inst_id(instalasi_qs)

        # Saring user berdasarkan instalasi yang dipilih
        data_user = Users.objects.exclude(is_superuser=True, is_active=False).prefetch_related('riwayat_penempatan').order_by('-id')
        if inst_id:
            data_user = data_user.filter(
                riwayat_penempatan__penempatan_level4__id=inst_id,
                riwayat_penempatan__status=True
            )
        else:
            data_user = data_user.none()

        # Evaluasi berdasarkan user yang tersaring
        full_data_table = get_evaluasi_tabel(inst_id, data_user, bulan, tahun)
        data = full_data_table.get('data') if inst_id else []
        # Paginate hasil evaluasi
        paginator = Paginator(data, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context.update({
            'data': page_obj.object_list,
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': paginator.num_pages > 1,

            'instalasi_list': instalasi_qs,
            'bulan_list': [(i, month_name[i]) for i in range(1, 13)],
            'tahun_list': list(range(datetime.now().year - 5, datetime.now().year + 6)),

            'selected_inst': int(inst_id) if inst_id else None,
            'selected_bulan': bulan,
            'selected_tahun': tahun,
            'preserved_query': self._get_preserved_query(),

            'title': 'Evaluasi Pembuatan Jadwal',
            'url': reverse('disiplinsdm_urls:jadwal_list'),
            'riwayat': 'active',
            'selected': 'disiplin',
        })

        return context


    def _get_preserved_query(self):
        querydict = self.request.GET.copy()
        querydict.pop('page', None)
        return querydict.urlencode()


class HariLiburView(LoginRequiredMixin, generic.ListView):
    model = HariLibur
    template_name = 'jadwal_piket/harilibur_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        querydict = self.request.GET.copy()
        querydict.pop('page', None)
        context['preserved_query'] = querydict.urlencode()
        context['query'] = self.request.GET.get('q', '')
        context['bulan_list'] = [(i, month_name[i]) for i in range(1, 13)]
        current_year = datetime.now().year
        context['tahun_list'] = [year for year in range(current_year - 5, current_year + 6)]
        context['title'] = 'Daftar Hari Libur'
        context['url'] = reverse('disiplinsdm_urls:jadwal_list')
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        return context
    
    def get_paginate_by(self, queryset):
        per_page = self.request.GET.get('per_page', 10)
        return per_page
    
    def get_queryset(self):
        tanggal = date.today()
        get = self.request.GET.get
        bulan = get('bulan')
        tahun = get('tahun')
        try:
            bulan = int(get('bulan')) if get('bulan') else tanggal.month
        except ValueError:
            bulan = tanggal.month

        try:
            tahun = int(get('tahun')) if get('tahun') else tanggal.year
        except ValueError:
            tahun = tanggal.year
        
        queryset = HariLibur.objects.filter(tanggal__month=bulan, tanggal__year=tahun).order_by('id')
        return queryset
    
    
class HariLiburCreateView(LoginRequiredMixin, UserPassesTestMixin, generic.CreateView):
    model = HariLibur
    template_name='kehadirankegiatan/form.html'
    success_url=reverse_lazy('disiplinsdm_urls:harilibur_list')
    form_class=HariLiburForm
    
    def test_func(self):
        if self.request.user.is_disiplin_admin:
            return True
        return False

    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        messages.error(self.request, 'Anda tidak memiliki izin untuk menambah hari libur.')
        return redirect(reverse('disiplinsdm_urls:harilibur_list'))
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['url'] = reverse('disiplinsdm_urls:harilibur_list')
        context['title'] = 'Tambah Hari Libur'
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        return context
    
    
class HariLiburUpdateView(LoginRequiredMixin, UserPassesTestMixin, generic.UpdateView):
    model = HariLibur
    template_name='kehadirankegiatan/form.html'
    success_url=reverse_lazy('disiplinsdm_urls:harilibur_list')
    form_class=HariLiburForm
    
    def test_func(self):
        if self.request.user.is_disiplin_admin:
            return True
        return False

    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        messages.error(self.request, 'Anda tidak memiliki izin untuk mengedit hari libur.')
        return redirect(reverse('disiplinsdm_urls:harilibur_list'))
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['url'] = reverse('disiplinsdm_urls:harilibur_list')
        context['title'] = 'Edit Hari Libur'
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        return context
    

class HariLiburDeleteView(LoginRequiredMixin, UserPassesTestMixin, generic.DeleteView):
    model = HariLibur
    template_name='jadwal_piket/validasi_delete_jadwal.html'
    success_url=reverse_lazy('disiplinsdm_urls:harilibur_list')
    
    def test_func(self):
        if self.request.user.is_disiplin_admin:
            return True
        return False

    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        messages.error(self.request, 'Anda tidak memiliki izin untuk menghapus hari libur.')
        return redirect(reverse('disiplinsdm_urls:harilibur_list'))
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['url'] = reverse('disiplinsdm_urls:harilibur_list')
        context['title'] = 'Hapus Hari Libur'
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        return context
    

def approve_instalasi_bulan(instalasi, bulan, tahun, user):
    with transaction.atomic():
        # Ambil ID pegawai yang statusnya 'diajukan'
        initial_jadwal_ids = list(JenisSDMPerinstalasi.objects.filter(
            instalasi=instalasi,
            bulan=bulan,
            tahun=tahun,
            status='diajukan'
        ).values_list('id', flat=True))
        
        # Ubah status jadi disetujui
        JenisSDMPerinstalasi.objects.filter(id__in=initial_jadwal_ids).update(status='disetujui')

        # Ambil jadwal yang akan di-approve
        draft_jadwal = JadwalDinasSDM.objects.filter(
            pegawai_id__in=initial_jadwal_ids,
            tanggal__year=tahun,
            tanggal__month=bulan
        )
        
        # Mapping baru: {(pegawai_id, tanggal): set(kategori_id)}
        new_shift_map = defaultdict(set)
        for jd in draft_jadwal:
            new_shift_map[(jd.pegawai_id, jd.tanggal)].add(jd.kategori_jadwal_id)
       
        # Ambil data ApprovedJadwalDinasSDM yang sudah ada
        old_approvals = ApprovedJadwalDinasSDM.objects.filter(
            pegawai_id__in=[j.pegawai_id for j in draft_jadwal],
            tanggal__year=tahun,
            tanggal__month=bulan
        ).select_related('kategori_jadwal', 'pegawai')

        # Mapping lama: {(pegawai_id, tanggal): set(kategori_id)}
        old_shift_map = defaultdict(set)
        for ap in old_approvals:
            old_shift_map[(ap.pegawai_id, ap.tanggal)].add(ap.kategori_jadwal_id)
        
        # Hapus shift yang tidak lagi dipakai (diganti)
        for (pegawai_id, tanggal), old_kategori_ids in old_shift_map.items():
            new_kategori_ids = new_shift_map.get((pegawai_id, tanggal), set())
            removed_shifts = old_kategori_ids - new_kategori_ids
            # added_ids = new_kategori_ids - old_kategori_ids
            # existing_ids = old_kategori_ids & new_kategori_ids
            # print(f"[{pegawai_id} - {tanggal}] Old: {old_kategori_ids} | New: {new_kategori_ids} | Removed: {removed_shifts} | Added: {added_ids}")
            
            if removed_shifts:
                ApprovedJadwalDinasSDM.objects.filter(
                    pegawai_id=pegawai_id,
                    tanggal=tanggal,
                    kategori_jadwal_id__in=removed_shifts
                ).delete()

        # Buat ulang existing_lookup setelah penghapusan
        existing_approvals = ApprovedJadwalDinasSDM.objects.filter(
            pegawai__in=[j.pegawai for j in draft_jadwal],
            tanggal__year=tahun,
            tanggal__month=bulan
        )
        existing_lookup = {
            (a.pegawai_id, a.tanggal, a.kategori_jadwal_id): a for a in existing_approvals
        }

        to_create = []
        to_update = []

        for jadwal in draft_jadwal:
            key = (jadwal.pegawai_id, jadwal.tanggal, jadwal.kategori_jadwal_id)

            if key in existing_lookup:
                obj = existing_lookup[key]
                obj.catatan = jadwal.catatan
                obj.is_approved = True
                obj.approved_by = user
                to_update.append(obj)
            else:
                to_create.append(ApprovedJadwalDinasSDM(
                    pegawai=jadwal.pegawai,
                    tanggal=jadwal.tanggal,
                    kategori_jadwal=jadwal.kategori_jadwal,
                    catatan=jadwal.catatan,
                    is_approved=True,
                    approved_by=user
                ))

        if to_create:
            ApprovedJadwalDinasSDM.objects.bulk_create(to_create)
        if to_update:
            ApprovedJadwalDinasSDM.objects.bulk_update(
                to_update,
                fields=['kategori_jadwal_id', 'catatan', 'is_approved', 'approved_by', 'updated_at']
            )
    

def generate_qr_with_logo(data: str, logo_path: str, size=300) -> str:
    # 1. Buat QR Code Dasar
    qr = qrcode.QRCode(
        # Menggunakan ERROR_CORRECT_H agar aman jika ditimpa logo
        error_correction=qrcode.constants.ERROR_CORRECT_H  
    )
    qr.add_data(data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert('RGB')

    # Resize QR ke ukuran target
    qr_img = qr_img.resize((size, size), Image.LANCZOS)

    # 2. PROSES PENEMPELAN LOGO (Hanya berjalan jika logo_path valid & file-nya ada)
    if logo_path and os.path.exists(logo_path):
        try:
            # Buka dan sesuaikan ukuran logo
            logo = Image.open(logo_path)
            logo_size = size // 4  # Logo 1/4 dari ukuran QR
            logo = logo.resize((logo_size, logo_size), Image.LANCZOS)

            # Tempel logo ke tengah QR
            pos = ((qr_img.size[0] - logo.size[0]) // 2, (qr_img.size[1] - logo.size[1]) // 2)
            qr_img.paste(logo, pos, mask=logo if logo.mode == 'RGBA' else None)
        except Exception as e:
            # Jika file gambar logo korup atau gagal dibaca, log error tapi biarkan QR tetap terbuat
            print(f"Gagal menempelkan logo pada QR Code: {e}")
    else:
        # Tampilkan log peringatan di server console jika file logo menghilang
        print(f"WARNING: Logo path '{logo_path}' tidak valid atau file tidak ditemukan. Menghasilkan QR standar.")

    # 3. Simpan ke BytesIO dan ubah ke Base64 Data URI
    buffer = BytesIO()
    qr_img.save(buffer, format='PNG')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    
    return f"data:image/png;base64,{img_base64}"


class JadwalBulananListView(LoginRequiredMixin, generic.ListView):
    model = JenisSDMPerinstalasi
    template_name = 'jadwal_piket/jadwal_pivot.html'
    context_object_name = 'metadata'

    def get_queryset(self):
        return JenisSDMPerinstalasi.objects.filter(
            bulan=self.bulan,
            tahun=self.tahun,
            instalasi=self.instalasi_id
        ).select_related('pegawai')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ambil data pivot terlebih dahulu
        pivot_data = self.get_pivot_data()
        context.update({
            'bulan': self.bulan,
            'tahun': self.tahun,
            'tanggal_range': range(1, self.jumlah_hari + 1),
            'data': self.get_pivot_data().values(),
            'title': f'Jadwal instalasi {self.instalasi.instalasi}',
            'url': reverse('disiplinsdm_urls:jadwal_list'),
            'inst': self.instalasi_id,
            'tab': 'draft',
            'riwayat': 'active',
            'selected': 'disiplin',
        })

        # Tambahkan status dan QR
        context.update(self.get_status_and_qr())
        
        # Tambahkan data summary ke context
        context['summary_data'] = self.get_summary_data(pivot_data)
        
        # BARU: Definisikan konfigurasi summary di sini
        context['shift_summary_config'] = [
            {'acr': 'P', 'label': 'PAGI'},
            {'acr': 'S', 'label': 'SIANG'},
            {'acr': 'M', 'label': 'MALAM'},
            {'acr': 'L', 'label': 'LIBUR'},
            {'acr': 'C', 'label': 'CUTI'},
            {'acr': 'LX', 'label': 'LIBUR EXTRA'},
        ]

        return context

    def get_pivot_data(self):
        pegawai_list = [md.pegawai for md in self.object_list]
        draft = JenisSDMPerinstalasi.objects.filter(bulan=self.bulan, tahun=self.tahun, status__in=['draft', 'ditolak', 'diajukan']).exists()
        
        jadwal_qs = ApprovedJadwalDinasSDM.objects.filter(
            tanggal__month=self.bulan,
            tanggal__year=self.tahun,
            pegawai__pegawai__in=pegawai_list
        ).select_related('kategori_jadwal', 'pegawai__pegawai')
        if draft:
            jadwal_qs = JadwalDinasSDM.objects.filter(
                tanggal__month=self.bulan,
                tanggal__year=self.tahun,
                pegawai__pegawai__in=pegawai_list
            ).select_related('kategori_jadwal', 'pegawai__pegawai')

        data = {
            md.pegawai.id: {
                'nama': md.pegawai.full_name,
                'jadwal': {}
            } for md in self.object_list
        }

        for j in jadwal_qs:
            pegawai_id = j.pegawai.pegawai.id
            hari = j.tanggal.day
            kategori = j.kategori_jadwal.akronim if j.kategori_jadwal else "-"

            # Pastikan key 'jadwal' sudah punya list untuk tanggal itu
            if hari not in data[pegawai_id]['jadwal']:
                data[pegawai_id]['jadwal'][hari] = []

            # Tambahkan kategori (hindari duplikat jika perlu)
            if kategori not in data[pegawai_id]['jadwal'][hari]:
                data[pegawai_id]['jadwal'][hari].append(kategori)

        return data
    
    # METODE BARU: Untuk menghitung total setiap jenis piket per hari
    def get_summary_data(self, pivot_data):
        """
        Menghitung total Pagi, Siang, Malam, L, C, dan LX untuk setiap tanggal.
        Menerima `pivot_data` sebagai argumen untuk efisiensi.
        """
        summary_counts = {}
        # Asumsi akronim: P=Pagi, S=Siang, M=Malam, L=Libur, C=Cuti, LX=Libur Extra
        shift_keys = ['P', 'S', 'M', 'L', 'C', 'LX'] 

        # Inisialisasi struktur data untuk total
        for day in range(1, self.jumlah_hari + 1):
            summary_counts[day] = {key: 0 for key in shift_keys}

        # Iterasi melalui data pegawai untuk mengakumulasi total
        for pegawai_data in pivot_data.values():
            for day, shifts in pegawai_data['jadwal'].items():
                for shift in shifts:
                    # Periksa apakah shift ada di dalam shift_keys yang ingin dihitung
                    if shift in summary_counts[day]:
                        summary_counts[day][shift] += 1
                        
        return summary_counts

    def get_status_and_qr(self):
        data = {}
        object_data = self.object_list.first()
        objects_draft = self.object_list.filter(status__in=['draft', 'ditolak']).exists()
        objects_pengajuan = self.object_list.filter(status='diajukan').exists()

        status = 'Draft'
        tanggal_pengajuan = datetime.now()
        tanggal_persetujuan = datetime.now()
        qr_image_pengajuan = None
        qr_image_persetujuan = None
        approved_jadwal = None

        if object_data and not objects_draft:
            status = 'Diajukan'
            tanggal_pengajuan = object_data.updated_at
            qr_image_pengajuan = self.generate_qr(
                f'diajukan oleh: {self.pimpinan_instalasi}\n'
                f'tanggal: {tanggal_pengajuan}\n'
                f'url: {self.get_absolute_url()}'
            )

            approved_jadwal = object_data.approvedjadwaldinassdm_set.filter(
                pegawai__bulan=self.bulan,
                pegawai__tahun=self.tahun
            ).first()

            if approved_jadwal and not objects_pengajuan:
                status = 'Disetujui'
                tanggal_persetujuan = approved_jadwal.updated_at
                qr_image_persetujuan = self.generate_qr(
                    f'disetujui oleh: {self.pimpinan}\n'
                    f'tanggal: {tanggal_persetujuan}\n'
                    f'url: {self.get_absolute_url()}'
                )

        return {
            'tanggal_pengajuan': tanggal_pengajuan,
            'qr_image_pengajuan': qr_image_pengajuan,
            'tanggal': tanggal_persetujuan,
            'qr_image': qr_image_persetujuan,
            'status': status,
            'pimpinan_instalasi': self.pimpinan_instalasi,
            'pimpinan': self.pimpinan,
            'object_draft': objects_draft,
            'object_pengajuan': objects_pengajuan,
            'approved_jadwal': approved_jadwal
        }
        
    # def generate_qr(self, data_str):
    #     image_path = '/var/www/html/prod/static/dist/img/logo_rsmandalika.png'
    #     if not os.path.exists(image_path):
    #         raise FileNotFoundError("Logo tidak ditemukan.")
    #     return generate_qr_with_logo(data_str, image_path)

    def generate_qr(self, data_str):
        image_path = finders.find('dist/img/logo_rsmandalika.png')
        
        return generate_qr_with_logo(data_str, image_path)

    def get_absolute_url(self):
        return self.request.build_absolute_uri(
            reverse('disiplinsdm_urls:jadwal_pivot', kwargs={'inst': self.instalasi_id})
        )

    @cached_property
    def bulan(self):
        return int(self.request.GET.get('bulan', date.today().month))

    @cached_property
    def tahun(self):
        return int(self.request.GET.get('tahun', date.today().year))

    @cached_property
    def jumlah_hari(self):
        return monthrange(self.tahun, self.bulan)[1]

    @cached_property
    def instalasi_id(self):
        return self.kwargs.get('inst')

    @cached_property
    def instalasi(self):
        return UnitInstalasi.objects.filter(pk=self.instalasi_id).select_related('sub_bidang').first()

    @cached_property
    def pimpinan_instalasi(self):
        return self.instalasi.nama_pimpinan if self.instalasi else '-'

    @cached_property
    def pimpinan(self):
        return self.instalasi.sub_bidang.nama_pimpinan if self.instalasi and self.instalasi.sub_bidang else '-'


class ApprovedJadwalBulananListView(LoginRequiredMixin, generic.ListView):
    model = JenisSDMPerinstalasi
    template_name = 'jadwal_piket/jadwal_pivot_approved.html'
    context_object_name = 'metadata'

    def get_queryset(self):
        return JenisSDMPerinstalasi.objects.filter(
            bulan=self.bulan,
            tahun=self.tahun,
            instalasi=self.instalasi_id
        ).select_related('pegawai')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ambil data pivot terlebih dahulu
        pivot_data = self.get_pivot_data()
        context.update({
            'bulan': self.bulan,
            'tahun': self.tahun,
            'tanggal_range': range(1, self.jumlah_hari + 1),
            'data': self.get_pivot_data().values(),
            'title': f'Jadwal instalasi {self.instalasi.instalasi}',
            'url': reverse('disiplinsdm_urls:jadwal_list'),
            'inst': self.instalasi_id,
            'tab': 'approved',
            'riwayat': 'active',
            'selected': 'disiplin',
        })

        # Tambahkan status dan QR
        context.update(self.get_status_and_qr())
        
        # Tambahkan data summary ke context
        context['summary_data'] = self.get_summary_data(pivot_data)
        
        # BARU: Definisikan konfigurasi summary di sini
        context['shift_summary_config'] = [
            {'acr': 'P', 'label': 'PAGI'},
            {'acr': 'S', 'label': 'SIANG'},
            {'acr': 'M', 'label': 'MALAM'},
            {'acr': 'L', 'label': 'LIBUR'},
            {'acr': 'C', 'label': 'CUTI'},
            {'acr': 'LX', 'label': 'LIBUR EXTRA'},
        ]

        return context

    def get_pivot_data(self):
        pegawai_list = [md.pegawai for md in self.object_list]
        
        jadwal_qs = ApprovedJadwalDinasSDM.objects.filter(
            tanggal__month=self.bulan,
            tanggal__year=self.tahun,
            pegawai__pegawai__in=pegawai_list
        ).select_related('kategori_jadwal', 'pegawai__pegawai')

        data = {
            md.pegawai.id: {
                'nama': md.pegawai.full_name,
                'jadwal': {}
            } for md in self.object_list
        }

        for j in jadwal_qs:
            pegawai_id = j.pegawai.pegawai.id
            hari = j.tanggal.day
            kategori = j.kategori_jadwal.akronim if j.kategori_jadwal else "-"

            # Pastikan key 'jadwal' sudah punya list untuk tanggal itu
            if hari not in data[pegawai_id]['jadwal']:
                data[pegawai_id]['jadwal'][hari] = []

            # Tambahkan kategori (hindari duplikat jika perlu)
            if kategori not in data[pegawai_id]['jadwal'][hari]:
                data[pegawai_id]['jadwal'][hari].append(kategori)

        return data
    
    # METODE BARU: Untuk menghitung total setiap jenis piket per hari
    def get_summary_data(self, pivot_data):
        """
        Menghitung total Pagi, Siang, Malam, L, C, dan LX untuk setiap tanggal.
        Menerima `pivot_data` sebagai argumen untuk efisiensi.
        """
        summary_counts = {}
        # Asumsi akronim: P=Pagi, S=Siang, M=Malam, L=Libur, C=Cuti, LX=Libur Extra
        shift_keys = ['P', 'S', 'M', 'L', 'C', 'LX'] 

        # Inisialisasi struktur data untuk total
        for day in range(1, self.jumlah_hari + 1):
            summary_counts[day] = {key: 0 for key in shift_keys}

        # Iterasi melalui data pegawai untuk mengakumulasi total
        for pegawai_data in pivot_data.values():
            for day, shifts in pegawai_data['jadwal'].items():
                for shift in shifts:
                    # Periksa apakah shift ada di dalam shift_keys yang ingin dihitung
                    if shift in summary_counts[day]:
                        summary_counts[day][shift] += 1
                        
        return summary_counts

    def get_status_and_qr(self):
        data = {}
        object_data = self.object_list.first()
        objects_draft = self.object_list.filter(status__in=['draft', 'ditolak', 'diajukan']).exists()

        status = 'Draft' if objects_draft else 'Disetujui'
        tanggal_pengajuan = datetime.now()
        tanggal_persetujuan = datetime.now()
        qr_image_pengajuan = None
        qr_image_persetujuan = None
        approved_jadwal = None

        if object_data:
            tanggal_pengajuan = object_data.updated_at
            qr_image_pengajuan = self.generate_qr(
                f'diajukan oleh: {self.pimpinan_instalasi}\n'
                f'tanggal: {tanggal_pengajuan}\n'
                f'url: {self.get_absolute_url()}'
            )

            approved_jadwal = object_data.approvedjadwaldinassdm_set.filter(
                pegawai__bulan=self.bulan,
                pegawai__tahun=self.tahun
            ).first()

            if approved_jadwal:
                tanggal_persetujuan = approved_jadwal.updated_at
                qr_image_persetujuan = self.generate_qr(
                    f'disetujui oleh: {self.pimpinan}\n'
                    f'tanggal: {tanggal_persetujuan}\n'
                    f'url: {self.get_absolute_url()}'
                )

        return {
            'tanggal_pengajuan': tanggal_pengajuan,
            'qr_image_pengajuan': qr_image_pengajuan,
            'tanggal': tanggal_persetujuan,
            'qr_image': qr_image_persetujuan,
            'status': status,
            'pimpinan_instalasi': self.pimpinan_instalasi,
            'pimpinan': self.pimpinan,
            'approved_jadwal': approved_jadwal
        }
        
    # def generate_qr(self, data_str):
    #     image_path = '/var/www/html/prod/static/dist/img/logo_rsmandalika.png'
    #     if not os.path.exists(image_path):
    #         raise FileNotFoundError("Logo tidak ditemukan.")
    #     return generate_qr_with_logo(data_str, image_path)

    def generate_qr(self, data_str):
        image_path = finders.find('dist/img/logo_rsmandalika.png')
        return generate_qr_with_logo(data_str, image_path)

    def get_absolute_url(self):
        return self.request.build_absolute_uri(
            reverse('disiplinsdm_urls:jadwal_pivot', kwargs={'inst': self.instalasi_id})
        )

    @cached_property
    def bulan(self):
        return int(self.request.GET.get('bulan', date.today().month))

    @cached_property
    def tahun(self):
        return int(self.request.GET.get('tahun', date.today().year))

    @cached_property
    def jumlah_hari(self):
        return monthrange(self.tahun, self.bulan)[1]

    @cached_property
    def instalasi_id(self):
        return self.kwargs.get('inst')

    @cached_property
    def instalasi(self):
        return UnitInstalasi.objects.filter(pk=self.instalasi_id).select_related('sub_bidang').first()

    @cached_property
    def pimpinan_instalasi(self):
        return self.instalasi.nama_pimpinan if self.instalasi else '-'

    @cached_property
    def pimpinan(self):
        return self.instalasi.sub_bidang.nama_pimpinan if self.instalasi and self.instalasi.sub_bidang else '-'


class PengajuanJadwalInstalasi(LoginRequiredMixin, UserPassesTestMixin, generic.View):
    def test_func(self):
        instalasi_id = self.kwargs['inst']
        user = self.request.user
        instalasi_pimpinan_list = UnitInstalasi.objects.filter(nama_pimpinan=user)
        if (user.is_staff or user.is_disiplin_admin) and instalasi_pimpinan_list.filter(pk=instalasi_id).exists():
            return True
        return False

    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        inst = self.kwargs.get('inst')
        messages.error(self.request, 'Anda tidak memiliki izin untuk melakukan pengajuan jadwal.')
        return redirect(reverse('disiplinsdm_urls:jadwal_pivot', kwargs={'inst':inst}))
    
    def post(self, request, *args, **kwargs):
        instalasi_id = kwargs.get('inst')
        bulan = kwargs.get('bulan')
        tahun = kwargs.get('tahun')
        instalasi = None
        try:
            instalasi=UnitInstalasi.objects.get(pk=instalasi_id)
        except UnitInstalasi.DoesNotExist:
            instalasi = None
        initial_jadwal = JenisSDMPerinstalasi.objects.filter(
            instalasi=instalasi, 
            bulan=bulan,
            tahun=tahun,
            status='draft'
        ).update(status='diajukan')
        messages.success(request, f'Jadwal instalasi {instalasi.instalasi} diajukan!')
        return redirect(reverse('disiplinsdm_urls:jadwal_pivot', kwargs={'inst':instalasi_id}))
    
        
class ApprovalJadwalInstalasi(LoginRequiredMixin, UserPassesTestMixin, generic.View):
    def test_func(self):
        user = self.request.user
        profil = getattr(user, 'profil_admin', None)
        if user.is_staff or user.is_disiplin_admin:
            if profil and profil.instalasi.exists():
                return True
            if profil and profil.sub_bidang.exists():
                return True
            if profil and profil.bidang.exists():
                return True
            if profil and profil.unor.exists():
                return True
            return False
        return False

    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        inst = self.kwargs.get('inst')
        messages.error(self.request, 'Anda tidak memiliki izin untuk melakukan approval jadwal ini.')
        return redirect(reverse('disiplinsdm_urls:jadwal_pivot', kwargs={'inst':inst}))
    
    def post(self, *args, **kwargs):
        instalasi_id = kwargs.get('inst')
        bulan = kwargs.get('bulan')
        tahun = kwargs.get('tahun')
        instalasi = None
        try:
            instalasi=UnitInstalasi.objects.get(pk=instalasi_id)
        except UnitInstalasi.DoesNotExist:
            instalasi = None
        with transaction.atomic():
            approve_instalasi_bulan(instalasi, bulan, tahun, self.request.user)
        return redirect(reverse('disiplinsdm_urls:jadwal_pivot', kwargs={'inst':instalasi_id}))


def generate_qr_with_logo_for_excel(data: str, logo_path: str, size=300) -> BytesIO:
    # Buat QR Code
    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_H  # Tingkat koreksi tinggi agar QR masih bisa dibaca meski ditimpa logo
    )
    qr.add_data(data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert('RGB')

    # Resize QR
    qr_img = qr_img.resize((size, size), Image.LANCZOS)

    # Buka logo
    logo = Image.open(logo_path)
    logo_size = size // 4  # Logo 1/4 dari ukuran QR
    logo = logo.resize((logo_size, logo_size), Image.LANCZOS)

    # Tempel logo ke tengah QR
    pos = ((qr_img.size[0] - logo.size[0]) // 2, (qr_img.size[1] - logo.size[1]) // 2)
    qr_img.paste(logo, pos, mask=logo if logo.mode == 'RGBA' else None)

    # Simpan ke BytesIO
    buffer = BytesIO()
    qr_img.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer


def hitung_standar_jam_kerja(model, bulan, tahun):
    _, total_hari = calendar.monthrange(tahun, bulan)
    jam_kerja = Decimal("0")
    for hari in range(1, total_hari + 1):
        tanggal = date(tahun, bulan, hari)
        if model.objects.filter(tanggal=tanggal).exists():
            continue
        weekday = tanggal.weekday()
        if weekday in [0, 1, 2, 3]:  # Senin–Kamis
            jam_kerja += Decimal("7")
        elif weekday == 4:  # Jumat
            jam_kerja += Decimal("6.5")
        elif weekday == 5:  # Sabtu
            jam_kerja += Decimal("4.5")
    return jam_kerja


def hitung_standar_jam_kerja_maks(model, bulan, tahun):
    _, total_hari = calendar.monthrange(tahun, bulan)
    jam_kerja = Decimal("0")
    for hari in range(1, total_hari + 1):
        tanggal = date(tahun, bulan, hari)
        if model.objects.filter(tanggal=tanggal).exists():
            continue
        weekday = tanggal.weekday()
        if weekday in [0, 1, 2, 3]:  # Senin–Kamis
            jam_kerja += Decimal("7")
        elif weekday == 4:  # Jumat
            jam_kerja += Decimal("6.5")
        elif weekday == 5:  # Sabtu
            jam_kerja += Decimal("5.5")
    return jam_kerja


def evaluasi_beban(jumlah_jam, standar_min, standar_max):
    if jumlah_jam < standar_min:
        return "Ringan 🔵"
    elif jumlah_jam <= standar_max:
        return "Ideal 🟢"
    else:
        return "Overload 🔴"


def draft_export_jadwal_excel(request, inst, bulan, tahun):
    jumlah_hari = monthrange(tahun, bulan)[1]
    nama_bulan = date(tahun, bulan, 1).strftime('%B')

    try:
        instalasi = UnitInstalasi.objects.get(pk=inst)
    except UnitInstalasi.DoesNotExist:
        instalasi = None

    metadata = JenisSDMPerinstalasi.objects.filter(
        bulan=bulan, tahun=tahun, instalasi=instalasi
    ).select_related('pegawai').prefetch_related('jadwaldinassdm_set__kategori_jadwal')

    # Siapkan workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jadwal {bulan}-{tahun}"

    judul = f"Jadwal Instalasi {instalasi.instalasi} Bulan {nama_bulan} {tahun}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=jumlah_hari + 4)
    judul_cell = ws.cell(row=1, column=1, value=judul)
    judul_cell.font = Font(size=14, bold=True)
    judul_cell.alignment = Alignment(horizontal='center')

    header = ["No", "Nama Pegawai"] + list(range(1, jumlah_hari + 1)) + ["Total Jam", "Evaluasi"]
    ws.append(header)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Loop data pegawai
    for idx, md in enumerate(metadata, start=1):
        baris = [idx, md.pegawai.full_name]
        jadwal_per_hari = {jadwal.tanggal.day: jadwal.kategori_jadwal.akronim if jadwal.kategori_jadwal else "-" for jadwal in md.jadwaldinassdm_set.all()}
        
        for tgl in range(1, jumlah_hari + 1):
            baris.append(jadwal_per_hari.get(tgl, "-"))

        total_jam = md.kurang_lebih_jam_kerja
        baris.append(total_jam)

        evaluasi = evaluasi_beban(
            Decimal(total_jam),
            Decimal(md.standar_min_efektif),
            Decimal(md.standar_max_efektif)
        )
        baris.append(evaluasi)

        ws.append(baris)

    # Styling border
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=jumlah_hari + 4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25

    # Tambahan info dan legend
    ws.append([])
    ws.append(["Keterangan Akronim:"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    keterangan_akronim = {
        "P": "Pagi", "S": "Siang", "M": "Malam", "Md": "Middle",
        "LP": "Lepas Piket", "L": "Libur", "LX": "Libur Extra", "C": "Cuti"
    }
    for kode, ket in keterangan_akronim.items():
        ws.append([kode, ket])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='center')

    # Response
    filename = f"jadwal_{bulan}_{tahun}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_jadwal_excel(request, inst, bulan, tahun):
    jumlah_hari = monthrange(tahun, bulan)[1]
    nama_bulan = date(tahun, bulan, 1).strftime('%B')

    try:
        instalasi = UnitInstalasi.objects.get(pk=inst)
    except UnitInstalasi.DoesNotExist:
        instalasi = None

    metadata = JenisSDMPerinstalasi.objects.filter(
        bulan=bulan, tahun=tahun, instalasi=instalasi
    ).select_related('pegawai')

    pegawai_list = [md.pegawai for md in metadata]

    jadwal_qs = ApprovedJadwalDinasSDM.objects.filter(
        tanggal__month=bulan,
        tanggal__year=tahun,
        pegawai__pegawai__in=pegawai_list
    ).select_related('kategori_jadwal', 'pegawai__pegawai')

    data = {}
    for md in metadata:
        data[md.pegawai.id] = {
            'nama': md.pegawai.full_name,
            'jadwal': defaultdict(str)
        }

    for j in jadwal_qs:
        hari = j.tanggal.day
        kategori = j.kategori_jadwal.akronim if j.kategori_jadwal else "-"
        slot = data[j.pegawai.pegawai.id]['jadwal']
        slot[hari] = f"{slot[hari]},{kategori}" if slot[hari] else kategori

    jam_pegawai = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
    for j in jadwal_qs:
        if not j.kategori_jadwal or not j.kategori_jadwal.waktu_datang or not j.kategori_jadwal.waktu_pulang:
            continue

        dt = datetime.combine(date.min, j.kategori_jadwal.waktu_datang)
        pt = datetime.combine(date.min, j.kategori_jadwal.waktu_pulang)
        if pt <= dt:
            pt += timedelta(days=1)
        durasi = Decimal(str((pt - dt).total_seconds() / 3600))
        jam_pegawai[j.pegawai.pegawai.id][j.tanggal.day] += durasi

    standar_min = hitung_standar_jam_kerja(HariLibur, bulan, tahun)
    standar_max = hitung_standar_jam_kerja_maks(HariLibur, bulan, tahun)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jadwal {bulan}-{tahun}"

    judul = f"Jadwal Instalasi {instalasi.instalasi} Bulan {nama_bulan} {tahun}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=jumlah_hari + 4)
    judul_cell = ws.cell(row=1, column=1, value=judul)
    judul_cell.font = Font(size=14, bold=True)
    judul_cell.alignment = Alignment(horizontal='center')

    header = ["No", "Nama Pegawai"] + list(range(1, jumlah_hari + 1)) + ["Total Jam", "Evaluasi"]
    ws.append(header)

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    for idx, (pid, row) in enumerate(data.items(), start=1):
        baris = [idx, row['nama']]
        total_jam = Decimal("0")
        for tgl in range(1, jumlah_hari + 1):
            isi = row['jadwal'].get(tgl, "-")
            jam = jam_pegawai[pid].get(tgl, Decimal("0"))
            total_jam += jam
            baris.append(isi)
        baris.append(float(total_jam))
        baris.append(evaluasi_beban(total_jam, standar_min, standar_max))
        ws.append(baris)

    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=jumlah_hari + 4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25

    ws.append([])
    ws.append([f"Standar jam kerja bulan ini:", f"{float(standar_min)} – {float(standar_max)} jam"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.append([])
    ws.append(["Keterangan Akronim:"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    keterangan_akronim = {
        "P": "Pagi", "S": "Siang", "M": "Malam", "Md": "Middle",
        "LP": "Lepas Piket", "L": "Libur", "LX": "Libur Extra", "C": "Cuti"
    }

    for kode, ket in keterangan_akronim.items():
        ws.append([kode, ket])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='center')
        
    # === Pengesahan Digital di Pojok Kanan Bawah ===
    tanggal = jadwal_qs.first()
    pimpinan = None
    try:
        unit = UnitInstalasi.objects.get(pk=inst)
        pimpinan = unit.sub_bidang.nama_pimpinan
    except UnitInstalasi.DoesNotExist:
        unit = None
    # Lokasi pengesahan (custom jika perlu)
    lokasi_pengesahan = "Pujut"
    tanggal_pengesahan = tanggal.updated_at.strftime("%d %B %Y")
    nama_pimpinan = pimpinan.full_name_2 if hasattr(pimpinan, 'full_name_2') else None

    # QR Code data
    qr_data = f"Jadwal Bulan {nama_bulan} {tahun} - Disahkan oleh {nama_pimpinan} ({lokasi_pengesahan}, {tanggal_pengesahan})"
    logo_path = finders.find('dist/img/logo_rsmandalika.png')  # Ganti sesuai path logo kamu
    # logo_path = '/var/www/html/prod/static/dist/img/logo_rsmandalika.png'

    qr_buffer = generate_qr_with_logo_for_excel(qr_data, logo_path)
    qr_image = XLImage(qr_buffer)
    qr_image.width = 150
    qr_image.height = 150

    # Posisi baris & kolom
    baris_awal = ws.max_row + 3
    kolom_qr = jumlah_hari
    
    col_letter = get_column_letter(kolom_qr)
    start_row = ws.max_row + 4

    # Isi teks pengesahan
    ws.cell(row=baris_awal, column=kolom_qr, value="Mengetahui,")
    ws.cell(row=baris_awal + 1, column=kolom_qr, value=f"{lokasi_pengesahan}, {tanggal_pengesahan}")
    ws.cell(row=baris_awal + 12, column=kolom_qr, value=nama_pimpinan)
    ws.cell(row=baris_awal + 12, column=kolom_qr).font = Font(bold=True)
    ws.cell(row=baris_awal + 12, column=kolom_qr).alignment = Alignment(horizontal='center')

    # Sisipkan QR Code
    qr_position = f"{col_letter}{start_row + 2}"
    ws.add_image(qr_image, qr_position)


    filename = f"jadwal_{bulan}_{tahun}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
    

class AjukanJadwalView(LoginRequiredMixin, generic.UpdateView):
    login_url = reverse_lazy('myaccount_urls:login_view')
    redirect_field_name = 'next'
    model = JenisSDMPerinstalasi
    form_class = PengajuanJadwalForm
    template_name = 'jadwal_piket/jadwal_pengajuan_persetujuan.html'
    
    def get_success_url(self):
        query_params = self.request.GET.copy()
        if query_params:  # hanya jika ada query string
            return f'{reverse("disiplinsdm_urls:jadwal_list")}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_list')

    def get_failure_url(self):
        pk = self.kwargs.get('pk')
        query_params = self.request.GET.copy()
        if query_params:
            return f'{reverse("disiplinsdm_urls:jadwal_auto_create", kwargs={"pk":pk})}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_auto_create', kwargs={'pk':pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obj = self.object
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        context['status'] = obj.status
        context['title'] = f'Detail Jadwal: { self.object.pegawai.full_name } ({ self.object.bulan }/{ self.object.tahun })'
        context['url'] = self.get_failure_url()

        queryset = obj.jadwaldinassdm_set.select_related('kategori_jadwal').order_by('tanggal')

        minggu_list = get_mingguan_lengkap(obj.bulan, obj.tahun)
        libur_set = set(HariLibur.objects.filter(
            tanggal__year=obj.tahun, tanggal__month=obj.bulan
        ).values_list('tanggal', flat=True))

        minggu_data = []
        total_bulanan = 0
        for minggu in minggu_list:
            hari_data = []
            total_mingguan = 0
            for jadwal in queryset:
                tgl = jadwal.tanggal
                if minggu[0] <= tgl <= minggu[-1]:
                    datang = jadwal.kategori_jadwal.waktu_datang if jadwal.kategori_jadwal else None
                    pulang = jadwal.kategori_jadwal.waktu_pulang if jadwal.kategori_jadwal else None
                    jam = 0
                    if datang and pulang:
                        mulai = datetime.combine(tgl, datang)
                        selesai = datetime.combine(tgl, pulang)
                        if selesai < mulai:
                            selesai += timedelta(days=1)
                        jam = round((selesai - mulai).total_seconds() / 3600, 1)
                    total_mingguan += jam
                    total_bulanan += jam

                    hari_data.append({
                        'tanggal': tgl,
                        'kategori': jadwal.kategori_jadwal,
                        'jam': jam,
                        'catatan': jadwal.catatan,
                        'libur': tgl.weekday() == 6 or tgl in libur_set
                    })
            minggu_data.append({
                'range': minggu,
                'data': hari_data,
                'total_jam': total_mingguan,
            })

        context.update({
            'minggu_data': minggu_data,
            'total_bulanan': total_bulanan,
            'standar_min': obj.standar_min_efektif,
            'standar_max': obj.standar_max_efektif,
            'selisih': obj.selisih_jam_kerja,
        })
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        if self.object.status not in ['draft', 'ditolak']:
            messages.warning(request, "Jadwal ini tidak dalam status draft atau ditolak.")
            return redirect(self.get_failure_url())

        jadwal_sdm = self.object

        with transaction.atomic():
            # Update status
            jadwal_sdm.status = 'diajukan'
            jadwal_sdm.save()
            messages.success(request, 'Pengajuan behasil dilakukan, informasikan ke atasan anda agar segera disetujui!')
        return redirect(self.get_success_url())


class SetujuiJadwalView(LoginRequiredMixin, UserPassesTestMixin, generic.UpdateView):
    login_url = reverse_lazy('myaccount_urls:login_view')
    redirect_field_name = 'next'
    model = JenisSDMPerinstalasi
    template_name = 'jadwal_piket/jadwal_pengajuan_persetujuan.html'
    fields = []
    
    def get_success_url(self):
        query_params = self.request.GET.copy()
        if query_params:  # hanya jika ada query string
            return f'{reverse("disiplinsdm_urls:jadwal_list")}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_list')
    
    def get_failure_url(self):
        pk = self.kwargs.get('pk')
        query_params = self.request.GET.copy()
        if query_params:  # hanya jika ada query string
            return f'{reverse("disiplinsdm_urls:jadwal_auto_create", kwargs={"pk":pk})}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_auto_create', kwargs={'pk':pk})

    def test_func(self):
        sub_bidang = SubBidang.objects.filter(nama_pimpinan=self.request.user).exists()
        if self.request.user.is_disiplin_admin:
            return True
        elif sub_bidang and (self.request.user.is_staff or self.request.user.is_disiplin_admin):
            return True
        return False

    def handle_no_permission(self):
        messages.error(self.request, "Anda tidak memiliki izin untuk menyetujui jadwal.")
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obj = self.object
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        context['status'] = obj.status if hasattr(obj, 'status') else 'draft'
        context['title'] = f'Detail Jadwal: { self.object.pegawai.full_name } ({ self.object.bulan }/{ self.object.tahun })'
        context['url'] = reverse_lazy('disiplinsdm_urls:jadwal_list')
        jadwal_list = obj.jadwaldinassdm_set.select_related('kategori_jadwal').order_by('tanggal')

        minggu_data = []
        minggu_list = get_mingguan_lengkap(obj.bulan, obj.tahun)
        libur_set = set(HariLibur.objects.filter(
            tanggal__year=obj.tahun,
            tanggal__month=obj.bulan
        ).values_list('tanggal', flat=True))

        total_bulanan = 0
        for minggu in minggu_list:
            hari_data = []
            total_mingguan = 0
            for jadwal in jadwal_list:
                if minggu[0] <= jadwal.tanggal <= minggu[-1]:
                    datang = jadwal.kategori_jadwal.waktu_datang if jadwal.kategori_jadwal else None
                    pulang = jadwal.kategori_jadwal.waktu_pulang if jadwal.kategori_jadwal else None
                    jam = 0
                    if datang and pulang:
                        mulai = datetime.combine(jadwal.tanggal, datang)
                        selesai = datetime.combine(jadwal.tanggal, pulang)
                        if selesai < mulai:
                            selesai += timedelta(days=1)
                        jam = round((selesai - mulai).total_seconds() / 3600, 1)
                    total_mingguan += jam
                    total_bulanan += jam
                    hari_data.append({
                        'tanggal': jadwal.tanggal,
                        'kategori': jadwal.kategori_jadwal,
                        'jam': jam,
                        'catatan': jadwal.catatan,
                        'libur': jadwal.tanggal.weekday() == 6 or jadwal.tanggal in libur_set
                    })
            minggu_data.append({
                'range': minggu,
                'data': hari_data,
                'total_jam': total_mingguan,
            })

        context.update({
            'minggu_data': minggu_data,
            'total_bulanan': total_bulanan,
            'standar_min': obj.standar_min_efektif,
            'standar_max': obj.standar_max_efektif,
            'selisih': obj.selisih_jam_kerja,
            'can_approve': self.request.user.is_staff or self.request.user.is_disiplin_admin,
        })
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        if self.object.status != 'diajukan':
            messages.warning(request, "Jadwal ini tidak dalam status 'diajukan'.")
            return redirect(self.get_failure_url())

        aksi = request.POST.get("aksi")
        alasan = request.POST.get("alasan_penolakan", "").strip()

        if aksi == 'tolak':
            if not alasan:
                messages.error(request, "Mohon isi alasan penolakan jika ingin menolak jadwal.")
                return redirect(self.get_failure_url())

            self.object.status = 'ditolak'
            self.object.alasan_penolakan = alasan
            self.object.save()

            messages.success(request, f"Jadwal {self.object.pegawai.full_name} berhasil ditolak.")
            return redirect(self.get_success_url())

        elif aksi == 'setujui':
            jadwal_list = self.object.jadwaldinassdm_set.filter(
                tanggal__month=self.object.bulan,
                tanggal__year=self.object.tahun
            )

            with transaction.atomic():
                self.object.status = 'disetujui'
                self.object.save()

                # Ambil semua jadwal approved yang sudah ada
                approved_existing = ApprovedJadwalDinasSDM.objects.filter(
                    pegawai=self.object,
                    tanggal__month=self.object.bulan,
                    tanggal__year=self.object.tahun
                )

                # Buat mapping: {(tanggal, kategori_id): Approved}
                approved_map = {
                    (a.tanggal, a.kategori_jadwal_id): a for a in approved_existing
                }

                # Buat mapping baru dari draft jadwal
                draft_map = {
                    (j.tanggal, j.kategori_jadwal_id): j for j in jadwal_list
                }

                to_create = []
                to_update = []

                # Update jika ada perubahan
                for key, jadwal in draft_map.items():
                    if key in approved_map:
                        approved = approved_map[key]
                        if (
                            approved.catatan != jadwal.catatan or
                            not approved.is_approved or
                            approved.approved_by != request.user
                        ):
                            approved.catatan = jadwal.catatan
                            approved.is_approved = True
                            approved.approved_by = request.user
                            to_update.append(approved)
                    else:
                        to_create.append(ApprovedJadwalDinasSDM(
                            pegawai=self.object,
                            tanggal=jadwal.tanggal,
                            kategori_jadwal=jadwal.kategori_jadwal,
                            catatan=jadwal.catatan,
                            is_approved=True,
                            approved_by=request.user
                        ))

                # Hapus approved yang sudah tidak ada di draft
                existing_keys = set(approved_map.keys())
                current_keys = set(draft_map.keys())
                removed_keys = existing_keys - current_keys
                if removed_keys:
                    ApprovedJadwalDinasSDM.objects.filter(
                        pegawai=self.object,
                        tanggal__in=[k[0] for k in removed_keys],
                        kategori_jadwal_id__in=[k[1] for k in removed_keys]
                    ).delete()

                if to_create:
                    ApprovedJadwalDinasSDM.objects.bulk_create(to_create)

                if to_update:
                    ApprovedJadwalDinasSDM.objects.bulk_update(
                        to_update,
                        fields=['catatan', 'is_approved', 'approved_by', 'updated_at']
                    )

            messages.success(request, f"Jadwal {self.object.pegawai.full_name} berhasil disetujui.")
            return redirect(self.get_success_url())


def get_jenis_hari(tanggal):
    weekday = tanggal.weekday()
    if weekday in [0, 1, 2, 3]:
        return "Senin s/d kamis"
    elif weekday == 4:
        return "Jumat"
    elif weekday == 5:
        return "Sabtu"
    elif weekday == 6:
        return "Ahad"
    
def generate_bulk_jadwal(pegawai_obj, bulan, tahun):
    minggu_list = get_mingguan_lengkap(bulan, tahun)
    libur_qs = set(HariLibur.objects.filter(tanggal__month=bulan, tanggal__year=tahun).values_list('tanggal', flat=True))
    detail_map = {
        "Senin s/d kamis": DetailKategoriJadwalDinas.objects.filter(hari="Senin s/d kamis").first(),
        "Jumat": DetailKategoriJadwalDinas.objects.filter(hari="Jumat").first(),
        "Sabtu": DetailKategoriJadwalDinas.objects.filter(hari="Sabtu").first(),
        "Ahad": DetailKategoriJadwalDinas.objects.filter(hari="Ahad").first()  # libur
    }

    jadwal_list = []
    for minggu in minggu_list:
        for tanggal in minggu:
            is_libur = tanggal in libur_qs or tanggal.weekday() == 6
            kategori = detail_map.get("Ahad")
            if not is_libur:
                jenis_hari = get_jenis_hari(tanggal)
                kategori = detail_map.get(jenis_hari)

            jadwal = JadwalDinasSDM(
                pegawai=pegawai_obj,
                tanggal=tanggal,
                kategori_jadwal=kategori
            )
            jadwal_list.append(jadwal)
    
    # Simpan semua jadwal sekaligus
    JadwalDinasSDM.objects.bulk_create(jadwal_list)
    return jadwal_list


def safe_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None
    
class JadwalListView(LoginRequiredMixin, generic.ListView):
    model = JenisSDMPerinstalasi
    template_name = 'jadwal_piket/jadwal_list.html'
    login_url = reverse_lazy('myaccount_urls:login_view')
    redirect_field_name = 'next'
    paginate_by = 10

    # ====================
    # Helper Methods
    # ====================
    def get_penempatan_object(self, nip):
        if not nip:
            return None
        return RiwayatPenempatan.objects.filter(pegawai__profil_user__nip=nip, status=True).last()

    def get_jenis_sdm(self, nip):
        if not nip:
            return None
        return RiwayatJabatan.objects.filter(pegawai__profil_user__nip=nip).last()

    def get_user(self, nip):
        if not nip:
            return None
        return Users.objects.filter(profil_user__nip=nip, is_active=True).first()

    def get_filter_params(self):
        get = self.request.GET.get
        today = date.today()

        try:
            bulan = int(get('bulan')) if get('bulan') else today.month
            if not (1 <= bulan <= 12):
                bulan = today.month
        except ValueError:
            bulan = today.month

        try:
            tahun = int(get('tahun')) if get('tahun') else today.year
        except ValueError:
            tahun = today.year

        return {
            'query': get('q'),
            'bulan': bulan,
            'tahun': tahun,
            'nip': get('nip'),
            'tanggal': get_date_from_string(get('tanggal') or today.isoformat()),
        }


    def get_active_instalasi(self):
        inst_param = self.request.GET.get('inst')
        if inst_param:
            return UnitInstalasi.objects.filter(pk=inst_param).first()

        user = self.request.user
        profil = getattr(user, 'profil_admin', None)

        if user.is_disiplin_admin:
            return UnitInstalasi.objects.first()
        elif profil:
            if profil.instalasi.exists():
                return profil.instalasi.first()
            if profil.sub_bidang.exists():
                sub_bidang_pks = profil.sub_bidang.values_list('pk', flat=True)
                return UnitInstalasi.objects.filter(sub_bidang__in=sub_bidang_pks).first()
            if profil.bidang.exists():
                bidang_pks = profil.bidang.values_list('pk', flat=True)
                return UnitInstalasi.objects.filter(sub_bidang__bidang__in=bidang_pks).first()
            if profil.unor.exists():
                unor_pks = profil.unor.values_list('pk', flat=True)
                return UnitInstalasi.objects.filter(sub_bidang__bidang__unor__in=unor_pks).first()
        else:
            instalasi = user.riwayat_penempatan.filter(status=True).first()
            if instalasi is not None:
                instalasi = instalasi.penempatan_level4
            else:
                instalasi = None
            return instalasi
        

    def get_user_queryset(self):
        user = self.request.user
        users = Users.objects.exclude(is_superuser=True, is_active=False).prefetch_related('riwayat_penempatan')
        profil = getattr(user, 'profil_admin', None)

        if user.is_disiplin_admin:
            return users
        if not profil:
            return users.none()

        if profil.instalasi.exists():
            return users.filter(riwayat_penempatan__penempatan_level4__in=profil.instalasi.values_list('pk', flat=True), riwayat_penempatan__status=True)
        if profil.sub_bidang.exists():
            return users.filter(riwayat_penempatan__penempatan_level3__in=profil.sub_bidang.values_list('pk', flat=True), riwayat_penempatan__status=True)
        if profil.bidang.exists():
            return users.filter(riwayat_penempatan__penempatan_level2__in=profil.bidang.values_list('pk', flat=True), riwayat_penempatan__status=True)
        if profil.unor.exists():
            return users.filter(riwayat_penempatan__penempatan_level1__in=profil.unor.values_list('pk', flat=True), riwayat_penempatan__status=True)
        return users.none()
    
    def get_instalasi_queryset(self):
        user = self.request.user
        instalasi_list = UnitInstalasi.objects.all().prefetch_related(
            'sub_bidang', 'sub_bidang__bidang', 'sub_bidang__bidang__unor'
        )
        profil = getattr(user, 'profil_admin', None)

        if user.is_disiplin_admin:
            return instalasi_list
        if not profil:
            return instalasi_list.none()

        # Tambahkan .distinct() di setiap return yang menggunakan filter riwayatpenempatan
        if profil.instalasi.exists():
            return instalasi_list.filter(
                pk__in=profil.instalasi.values_list('pk', flat=True), 
                riwayatpenempatan__status=True
            ).distinct() # <-- Mencegah duplikasi data akibat JOIN
            
        if profil.sub_bidang.exists():
            return instalasi_list.filter(
                sub_bidang__in=profil.sub_bidang.values_list('pk', flat=True), 
                riwayatpenempatan__status=True
            ).distinct()
            
        if profil.bidang.exists():
            return instalasi_list.filter(
                sub_bidang__bidang__in=profil.bidang.values_list('pk', flat=True), 
                riwayatpenempatan__status=True
            ).distinct()
            
        if profil.unor.exists():
            return instalasi_list.filter(
                sub_bidang__bidang__unor__in=profil.unor.values_list('pk', flat=True), 
                riwayatpenempatan__status=True
            ).distinct()
            
        return instalasi_list.none()
    
    def get_queryset(self):
        params = self.get_filter_params()
        instalasi = self.get_active_instalasi()
        queryset = JenisSDMPerinstalasi.objects.select_related('pegawai', 'jenis_sdm', 'instalasi').order_by('-id').exclude(
                Q(pegawai__is_active=False) |
                Q(pegawai__is_superuser=True)
            )
        if self.request.user.is_disiplin_admin:
            queryset = queryset
        if instalasi:
            queryset = queryset.filter(instalasi=instalasi)
        else:
            queryset = queryset.none()
        
        queryset = queryset.filter(bulan=params['bulan'], tahun=params['tahun'])
        if params['query']:
            q = params['query']
            queryset = queryset.filter(
                Q(pegawai__first_name__icontains=q) |
                Q(pegawai__last_name__icontains=q) |
                Q(jenis_sdm__profesi__profesi__icontains=q) |
                Q(instalasi__instalasi__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = self.get_filter_params()
        instalasi = self.get_active_instalasi()
        selelcted_instalasi = instalasi.instalasi if instalasi is not None else ''
        
        instalasi_list = self.get_instalasi_queryset()

        context['instalasi_list'] = instalasi_list
        context['instalasi_param'] = instalasi.pk if instalasi else None
        context['users'] = self.get_user_queryset()

        selected_user = self.get_user(params['nip'])
        penempatan = self.get_penempatan_object(params['nip'])
        jenis_sdm = self.get_jenis_sdm(params['nip'])
        jenis_sdm_nama = jenis_sdm.nama_jabatan if jenis_sdm else None

        form_initial = {
            'jenis_sdm': jenis_sdm_nama,
            'pegawai': selected_user,
            'unor': getattr(penempatan, 'penempatan_level1', None),
            'bidang': getattr(penempatan, 'penempatan_level2', None),
            'sub_bidang': getattr(penempatan, 'penempatan_level3', None),
            'instalasi': getattr(penempatan, 'penempatan_level4', None),
            'bulan': params['tanggal'].month,
            'tahun': params['tanggal'].year,
        }

        current_year = datetime.now().year
        querydict = self.request.GET.copy()
        querydict.pop('page', None)

        context.update({
            'form': JenisSDMPerinstalasiForm(initial=form_initial),
            'bulan_list': [(i, month_name[i]) for i in range(1, 13)],
            'tahun_list': list(range(current_year - 5, current_year + 6)),
            'bulan': params['bulan'],
            'tahun': params['tahun'],
            'selected_user': selected_user,
            'preserved_query': querydict.urlencode(),
            'query': params['query'],
            'searchform': SearchForm(self.request.GET or None),
            'instalasi':instalasi,
            'title_page': 'disiplin',
            'riwayat': 'active',
            'selected': 'disiplin',
            'title': f"Jadwal Kerja Unit/Instalasi {selelcted_instalasi}",
        })
        return context

    def post(self, request, *args, **kwargs):
        if not request.user.is_staff and not request.user.is_disiplin_admin:
            messages.warning(request, 'Maaf anda tidak berhak menambahkan jadwal pegawai!')
            return redirect(reverse('disiplinsdm_urls:jadwal_list'))

        form = JenisSDMPerinstalasiForm(request.POST)
        pegawai = form.data.get('pegawai')
        bulan = form.data.get('bulan')
        tahun = form.data.get('tahun')

        if pegawai and bulan and tahun:
            existing = JenisSDMPerinstalasi.objects.filter(pegawai__id=pegawai, bulan=bulan, tahun=tahun).first()
            if existing:
                messages.info(request, 'Jadwal sudah ada untuk pegawai ini pada bulan dan tahun tersebut.')
                return redirect(reverse('disiplinsdm_urls:jadwal_auto_create', kwargs={'pk': existing.pk}))

        if form.is_valid():
            with transaction.atomic():
                obj = form.save()
                generate_bulk_jadwal(obj, int(bulan), int(tahun))
            return redirect(reverse('disiplinsdm_urls:jadwal_auto_create', kwargs={'pk': obj.pk}))

        return render(request, self.template_name, {
            'form': form,
            'error': 'Form tidak valid.'
        })
    

class JadwalDinasFormsetUpdateView(LoginRequiredMixin, UserPassesTestMixin, generic.UpdateView):#view pembuatan multipiket 1 hari
    login_url = reverse_lazy('myaccount_urls:login_view')
    redirect_field_name = 'next'
    model = JenisSDMPerinstalasi
    fields = []
    template_name = 'jadwal_piket/jadwal_dinas_form.html'

    def get_success_url(self):
        query_params = self.request.GET.copy()
        if query_params:  # hanya jika ada query string
            return f'{reverse("disiplinsdm_urls:jadwal_list")}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_list')
    
    def get_failure_url(self):
        pk = self.kwargs.get('pk')
        query_params = self.request.GET.copy()
        if query_params:
            return f'{reverse("disiplinsdm_urls:jadwal_auto_create", kwargs={"pk":pk})}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_auto_create', kwargs={'pk':pk})
    
    def test_func(self):
        user = self.request.user
        if user.is_staff or user.is_disiplin_admin:
            profil = user.profil_admin if hasattr(user, 'profil_admin') else None
            if profil and profil.instalasi.exists():
                return True
            if profil and profil.sub_bidang.exists():
                return True
            if profil and profil.bidang.exists():
                return True
            if profil and profil.unor.exists():
                return True
            return False
        return False
    
    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        messages.error(self.request, 'Anda tidak memiliki izin untuk membuat atau mengedit jadwal.')
        return redirect(self.get_success_url())
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Cek apakah POST berasal dari klik tombol "+ Shift"
        if request.POST.get('action') == 'tambah_shift':
            shift_tanggal = request.POST.get('shift_tanggal') or request.GET.get('shift_tanggal')
            
            if shift_tanggal:
                try:
                    tanggal_obj = datetime.strptime(shift_tanggal, "%Y-%m-%d").date()
                    label_hari = {
                        0: 'Senin s/d kamis', 1: 'Senin s/d kamis', 2: 'Senin s/d kamis',
                        3: 'Senin s/d kamis', 4: 'Jumat', 5: 'Sabtu', 6: 'Minggu'
                    }
                    label = label_hari[tanggal_obj.weekday()]
                    
                    kategori_default = DetailKategoriJadwalDinas.objects.filter(
                        kategori_dinas__kategori_dinas='Reguler',
                        hari=label
                    ).first()

                    # Pembuatan data baru yang aman di dalam request POST
                    JadwalDinasSDM.objects.create(
                        pegawai=self.object,
                        tanggal=tanggal_obj,
                        kategori_jadwal=kategori_default
                    )
                    messages.success(request, f"Shift baru ditambahkan untuk {tanggal_obj.strftime('%d %b %Y')}")
                except Exception as e:
                    messages.error(request, f"Gagal menambahkan shift: {e}")
            
            # Setelah data berhasil dibuat, langsung redirect (PRG Pattern)
            # Ini akan mengarah ke URL bersih menggunakan request GET sehingga loop terhenti
            return redirect(self.get_failure_url())
        
        # Jika POST berasal dari tombol utama (Simpan Draft / Ajukan), 
        # biarkan Django UpdateView yang mengurusnya melalui form_valid/form_invalid
        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = JadwalDinasSDM.objects.filter(
            pegawai=self.object,
            tanggal__year=self.object.tahun,
            tanggal__month=self.object.bulan
        ).order_by('tanggal')

        if self.request.POST:
            formset = update_jadwal_formset(data=self.request.POST, instance=self.object, queryset=queryset)
        else:
            formset = update_jadwal_formset(instance=self.object, queryset=queryset)

        minggu_formsets, total = self.build_mingguan_context(formset)

        context['formset'] = formset
        context['minggu_formsets'] = minggu_formsets
        context['total_bulanan'] = total
        context['url'] = reverse('disiplinsdm_urls:jadwal_list')
        context['title'] = 'Atur jadwal pegawai'
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        query_params = self.request.GET.copy()

        if formset.is_valid():
            with transaction.atomic():
                self.object = form.save(commit=False)
                self.object.status = 'draft'
                self.object.save()
                formset.instance = self.object
                for f in formset.forms:
                    if not f.cleaned_data.get('kategori_jadwal'):
                        initial_value = f.initial.get('kategori_jadwal')
                        if initial_value:
                            f.instance.kategori_jadwal_id = initial_value
                formset.save()
            if 'ajukan' in self.request.POST:
                messages.success(self.request, "Silahkan cek kembali detail data yang telah dibuat sebelum disimpan!")
                if query_params:
                    return redirect(f'{reverse("disiplinsdm_urls:ajukan_jadwal", kwargs={"pk": self.object.pk})}?{query_params.urlencode()}')
                return redirect(reverse('disiplinsdm_urls:ajukan_jadwal', kwargs={'pk': self.object.pk}))
            else:
                messages.success(self.request, "Data berhasil disimpan sebagai draft.")
                return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)
        
    def form_invalid(self, form):
        messages.error(self.request, 'Maaf data gagal disimpan!')
        context = self.get_context_data(form=form)
        context['query_params'] = self.request.GET.copy()
        return self.render_to_response(context)

    def build_mingguan_context(self, formset):
        minggu_list = get_mingguan_lengkap(self.object.bulan, self.object.tahun)
        minggu_formsets = []
        total = 0

        libur_set = set(HariLibur.objects.filter(
            tanggal__year=self.object.tahun,
            tanggal__month=self.object.bulan
        ).values_list('tanggal', flat=True))

        label_hari = {
            0: 'Senin s/d kamis',
            1: 'Senin s/d kamis',
            2: 'Senin s/d kamis',
            3: 'Senin s/d kamis',
            4: 'Jumat',
            5: 'Sabtu',
        }

        kategori_libur = DetailKategoriJadwalDinas.objects.filter(
            kategori_jadwal='Libur'
        ).first()

        from collections import defaultdict
        for minggu in minggu_list:
            tanggal_forms = defaultdict(list)
            for f in formset.forms:
                tgl = f.instance.tanggal
                if tgl and minggu[0] <= tgl <= minggu[-1]:
                    is_libur = tgl.weekday() == 6 or tgl in libur_set
                    f.is_hari_libur = is_libur

                    # Auto-assign kategori_jadwal jika kosong
                    if not f.instance.kategori_jadwal:
                        if is_libur:
                            f.initial['kategori_jadwal'] = kategori_libur
                        else:
                            label = label_hari.get(tgl.weekday())
                            default_kat = DetailKategoriJadwalDinas.objects.filter(
                                kategori_dinas__kategori_dinas='Reguler',
                                hari=label
                            ).first()
                            if default_kat:
                                f.initial['kategori_jadwal'] = default_kat

                    tanggal_forms[tgl].append(f)

            minggu_data = []
            for tgl, forms in sorted(tanggal_forms.items()):
                jam = hitung_total_jam(forms, DetailKategoriJadwalDinas)
                total += jam
                minggu_data.append({
                    'tanggal': tgl,
                    'forms': forms,
                    'total_jam': jam
                })

            minggu_formsets.append({
                'range': minggu,
                'data_harian': minggu_data,
                'total_jam': sum(h['total_jam'] for h in minggu_data)
            })

        return minggu_formsets, total

           
class JadwalDinasFormsetUpdateView2(generic.UpdateView):#view pembuatan jadwal hanya 1 piket 1 hari
    model = JenisSDMPerinstalasi
    fields = []
    template_name = 'jadwal_piket/jadwal_dinas_form.html'
    success_url = reverse_lazy('disiplinsdm_urls:jadwal_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            formset = update_jadwal_formset(data=self.request.POST, instance=self.object)
        else:
            formset = update_jadwal_formset(instance=self.object)
        minggu_formsets, total = self.build_mingguan_context(formset)
        context['formset'] = formset
        context['minggu_formsets'] = minggu_formsets
        context['total_bulanan'] = total
        context['url'] = reverse('disiplinsdm_urls:jadwal_list')
        context['title'] = 'Atur jadwal pegawai'
        context['riwayat'] ='active'
        context['selected'] ='disiplin'
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            with transaction.atomic():
                # Terapkan nilai initial yang tidak terkirim dalam POST
                for f in formset.forms:
                    if not f.cleaned_data.get('kategori_jadwal'):
                        initial_value = f.initial.get('kategori_jadwal')
                        if initial_value:
                            f.instance.kategori_jadwal_id = initial_value
                formset.save()
                messages.success(self.request, 'Data berhasil disimpan!')
            return super().form_valid(form)
        else:
            messages.error(self.request, 'Maaf data gagal disimpan!')
            print('formset: ', formset.errors)
            print('form: ', form.errors)
            return self.form_invalid(form)
        
    def build_mingguan_context(self, formset):
        minggu_list = get_mingguan_lengkap(self.object.bulan, self.object.tahun)
        minggu_formsets = []
        total = 0

        libur_set = set(HariLibur.objects.filter(
            tanggal__year=self.object.tahun,
            tanggal__month=self.object.bulan
        ).values_list('tanggal', flat=True))

        label_hari = {
            0: 'Senin s/d kamis',
            1: 'Senin s/d kamis',
            2: 'Senin s/d kamis',
            3: 'Senin s/d kamis',
            4: 'Jumat',
            5: 'Sabtu',
        }

        kategori_libur = DetailKategoriJadwalDinas.objects.filter(
            kategori_jadwal='Libur'
        ).first()

        for minggu in minggu_list:
            forms = []
            for f in formset.forms:
                tgl = f.instance.tanggal
                if tgl and minggu[0] <= tgl <= minggu[-1]:
                    is_libur = tgl.weekday() == 6 or tgl in libur_set
                    f.is_hari_libur = is_libur
                    if not f.instance.kategori_jadwal:
                        if is_libur:
                            f.initial['kategori_jadwal'] = kategori_libur
                        else:
                            label = label_hari.get(tgl.weekday())
                            default_kat = DetailKategoriJadwalDinas.objects.filter(
                                kategori_dinas__kategori_dinas='Reguler',
                                hari=label
                            ).first()
                            if default_kat:
                                f.initial['kategori_jadwal'] = default_kat
                    forms.append(f)        
                    
            jam = hitung_total_jam(forms, DetailKategoriJadwalDinas)
            total += jam
            minggu_formsets.append({
                'range': minggu,
                'formset': forms,
                'total_jam': jam
            })

        return minggu_formsets, total

    
class JadwalDinasDetailView(LoginRequiredMixin, generic.DetailView):#view untuk multipiket perhari
    model = JenisSDMPerinstalasi
    template_name = 'jadwal_piket/jadwal_detail.html'
    context_object_name = 'object'
    
    def back_url(self):
        query_params = self.request.GET.copy()
        if query_params:
            return f'{reverse("disiplinsdm_urls:jadwal_list")}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obj = self.object
        jadwal_list = obj.jadwaldinassdm_set.select_related('kategori_jadwal').order_by('tanggal')

        minggu_data = []
        minggu_list = get_mingguan_lengkap(obj.bulan, obj.tahun)
        libur_set = set(HariLibur.objects.filter(
            tanggal__year=obj.tahun,
            tanggal__month=obj.bulan
        ).values_list('tanggal', flat=True))

        total_bulanan = 0
        for minggu in minggu_list:
            hari_data = []
            total_mingguan = 0
            for jadwal in jadwal_list:
                if minggu[0] <= jadwal.tanggal <= minggu[-1]:
                    datang = jadwal.kategori_jadwal.waktu_datang if jadwal.kategori_jadwal else None
                    pulang = jadwal.kategori_jadwal.waktu_pulang if jadwal.kategori_jadwal else None
                    jam = 0
                    if datang and pulang:
                        mulai = datetime.combine(jadwal.tanggal, datang)
                        selesai = datetime.combine(jadwal.tanggal, pulang)
                        if selesai < mulai:
                            selesai += timedelta(days=1)
                        jam = round((selesai - mulai).total_seconds() / 3600, 1)
                    total_mingguan += jam
                    total_bulanan += jam
                    hari_data.append({
                        'tanggal': jadwal.tanggal,
                        'kategori': jadwal.kategori_jadwal,
                        'jam': jam,
                        'catatan': jadwal.catatan,
                        'libur': jadwal.tanggal.weekday() == 6 or jadwal.tanggal in libur_set
                    })
            minggu_data.append({
                'range': minggu,
                'data': hari_data,
                'total_jam': total_mingguan,
            })

        context.update({
            'minggu_data': minggu_data,
            'total_bulanan': total_bulanan,
            'standar_min': obj.standar_min_efektif,
            'standar_max': obj.standar_max_efektif,
            'selisih': obj.selisih_jam_kerja,
            'can_approve': self.request.user.is_staff or self.request.user.is_disiplin_admin,
        })
        context['url'] = reverse('disiplinsdm_urls:jadwal_list')
        context['title'] = 'Lihat Jadwal Pegawai'
        context['title_page'] = 'Lihat Jadwal'
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        return context
    

class SalinJadwalView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'jadwal_piket/jadwal_salin_form.html'
    success_url_name = 'disiplinsdm_urls:jadwal_list'
    failure_url_name = 'disiplinsdm_urls:salin_jadwal'
    # NOTE: pastikan nama namespace ini konsisten di urls.py

    # ===== Utilities =====
    def _url_with_query(self, urlname):
        base = reverse(urlname)
        qp = self.request.GET.copy()
        return f'{base}?{qp.urlencode()}' if qp else base

    def get_success_url(self):
        return self._url_with_query(self.success_url_name)

    def get_failure_url(self):
        return self._url_with_query(self.failure_url_name)

    # ===== Permissions =====
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_disiplin_admin

    def handle_no_permission(self):
        messages.error(self.request, 'Anda tidak memiliki izin untuk menyalin data.')
        return redirect(self.get_failure_url())

    # ===== Data helpers =====
    def get_instalasi(self, pegawai):
        if not pegawai:
            return None
        # Jika pegawai bisa berupa pk/int, amankan:
        qs = RiwayatPenempatan.objects.filter(pegawai=pegawai, status=True)
        return qs.select_related(
            'penempatan_level1','penempatan_level2','penempatan_level3','penempatan_level4'
        ).first()

    def get_jadwal_meta(self, pegawai, bulan, tahun):
        if not all([pegawai, bulan, tahun]):
            return None
        return (JenisSDMPerinstalasi.objects
                .filter(pegawai=pegawai, bulan=bulan, tahun=tahun)
                .select_related('unor','bidang','sub_bidang','instalasi','jenis_sdm')
                .first())

    def get_pegawai(self, pk):
        if not pk:
            return None
        return Users.objects.filter(pk=pk).first()

    def _parse_params(self, request):
        sumber_pk = request.GET.get('sumber')
        tujuan_pk = request.GET.get('tujuan')
        tgl_str  = request.GET.get('tanggal')

        tanggal = get_date_from_string(tgl_str) if tgl_str else timezone.now().date()
        sumber  = self.get_pegawai(sumber_pk)
        tujuan  = self.get_pegawai(tujuan_pk)
        return sumber, tujuan, tanggal

    def _users_queryset_for_login(self, login_user):
        base = (Users.objects
                .exclude(is_superuser=True, is_active=False)
                .prefetch_related('riwayat_penempatan'))

        users = base.none()
        if login_user.is_disiplin_admin:
            users = base
        elif login_user.is_staff and not login_user.is_disiplin_admin:
            pa = getattr(login_user, 'profil_admin', None)
            if not pa:
                return base.none()

            filt = Q(riwayat_penempatan__status=True)
            if getattr(pa, 'instalasi', None) and pa.instalasi.exists():
                filt &= Q(riwayat_penempatan__penempatan_level4__in=pa.instalasi.values_list('pk', flat=True))
            elif getattr(pa, 'sub_bidang', None) and pa.sub_bidang.exists():
                filt &= Q(riwayat_penempatan__penempatan_level3__in=pa.sub_bidang.values_list('pk', flat=True))
            elif getattr(pa, 'bidang', None) and pa.bidang.exists():
                filt &= Q(riwayat_penempatan__penempatan_level2__in=pa.bidang.values_list('pk', flat=True))
            elif getattr(pa, 'unor', None) and pa.unor.exists():
                filt &= Q(riwayat_penempatan__penempatan_level1__in=pa.unor.values_list('pk', flat=True))
            else:
                # fallback aman: tidak melihat siapa pun
                return base.none()

            users = base.filter(filt).distinct()
        return users

    # ===== HTTP methods =====
    def get(self, request):
        login_user = request.user
        sumber, tujuan, tanggal = self._parse_params(request)
        bulan, tahun = tanggal.month, tanggal.year

        users = self._users_queryset_for_login(login_user)

        selected_instalasi = self.get_instalasi(sumber)
        selected_jadwal    = self.get_jadwal_meta(tujuan, bulan, tahun)

        initial = {
            'instalasi': selected_instalasi,
            'sumber': sumber,
            'tujuan': tujuan,
            'bulan': bulan,
            'tahun': tahun,
            'tanggal': tanggal,
        }

        form = SalinJadwalForm(initial=initial)

        context = {
            'users': users,
            'form': form,
            'title': 'Copy jadwal pegawai',
            'selected_instalasi': selected_instalasi,
            'selected_sumber': sumber,
            'selected_tujuan': tujuan,
            'selected_jadwal': selected_jadwal,
            'url': reverse(self.success_url_name),
            'riwayat': 'active',
            'selected': 'disiplin',
        }
        return render(request, self.template_name, context)

    def post(self, request):
        form = SalinJadwalForm(request.POST)
        if not form.is_valid():
            return redirect(self.get_failure_url())

        sumber = form.cleaned_data['sumber']
        tujuan = form.cleaned_data['tujuan']
        bulan = form.cleaned_data['bulan']
        tahun = form.cleaned_data['tahun']

        sumber_inst = (JenisSDMPerinstalasi.objects
                       .filter(pegawai=sumber, bulan=bulan, tahun=tahun)
                       .select_related('unor','bidang','sub_bidang','instalasi','jenis_sdm')
                       .first())

        if not sumber_inst:
            messages.warning(request, "Pegawai sumber belum memiliki metadata jadwal bulan ini.")
            return render(request, self.template_name, {'form': form})

        with transaction.atomic():
            tujuan_inst, created = JenisSDMPerinstalasi.objects.get_or_create(
                pegawai=tujuan,
                bulan=bulan,
                tahun=tahun,
                defaults={
                    'unor': sumber_inst.unor,
                    'bidang': sumber_inst.bidang,
                    'sub_bidang': sumber_inst.sub_bidang,
                    'instalasi': sumber_inst.instalasi,
                    'jenis_sdm': sumber_inst.jenis_sdm,
                }
            )

            if not created:
                messages.warning(
                    request,
                    f"Jadwal pegawai atas nama {tujuan_inst.pegawai.full_name} sudah ada!"
                )
                # pastikan namespace url benar:
                return redirect(reverse('disiplinsdm_urls:jadwal_auto_create', kwargs={'pk': tujuan_inst.pk}))

            # Ambil jadwal sumber (hanya field perlu)
            jadwal_sumber = (JadwalDinasSDM.objects
                             .filter(pegawai=sumber_inst)
                             .only('tanggal', 'kategori_jadwal'))

            jadwal_baru = []
            for j in jadwal_sumber:
                try:
                    tanggal_baru = j.tanggal.replace(month=bulan, year=tahun)
                except ValueError:
                    # mis. 31 → Februari
                    continue
                jadwal_baru.append(JadwalDinasSDM(
                    pegawai=tujuan_inst,
                    tanggal=tanggal_baru,
                    kategori_jadwal=j.kategori_jadwal
                ))

            JadwalDinasSDM.objects.bulk_create(jadwal_baru, ignore_conflicts=True)

        messages.success(
            request,
            f"Berhasil menyalin {len(jadwal_baru)} entri jadwal dari {sumber.full_name} ke {tujuan.full_name}."
        )
        return redirect(self.get_success_url())



class SalinJadwalInstalasiView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'jadwal_piket/jadwal_salin_instalasi_form.html'
    
    def get_success_url(self):
        query_params = self.request.GET.copy()
        if query_params:
            return f'{reverse("disiplinsdm_urls:jadwal_list")}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_list')
    
    def get_failure_url(self):
        query_params = self.request.GET.copy()
        if query_params:
            return f'{reverse("disiplinsdm_urls:salin_jadwal_instalasi")}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:salin_jadwal_instalasi')
    
    def test_func(self):
        if self.request.user.is_staff or self.request.user.is_disiplin_admin:
            return True
        return False

    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        messages.error(self.request, 'Anda tidak memiliki izin untuk menyalin data.')
        return redirect(self.get_success_url())

    def get(self, request):
        form = SalinJadwalInstalasiForm(user=self.request.user)
        context = {
            'title':'Silahkan buat jadwal instalasi anda berdasarkan jadwal pada bulan sebelumnya',
            'url' : self.get_success_url(),
            'form': form
        }
        return render(request, self.template_name, context)

    def post(self, request):
        user = self.request.user
        form = SalinJadwalInstalasiForm(request.POST, user=user)
        if form.is_valid():
            instalasi = form.cleaned_data['instalasi']
            bulan_sumber = int(form.cleaned_data['bulan_sumber'])
            tahun_sumber = form.cleaned_data['tahun_sumber']
            bulan_tujuan = int(form.cleaned_data['bulan'])
            tahun_tujuan = form.cleaned_data['tahun']

            sumber_jadwal = JenisSDMPerinstalasi.objects.filter(
                instalasi=instalasi,
                bulan=bulan_sumber,
                tahun=tahun_sumber
            ).prefetch_related('jadwaldinassdm_set', 'pegawai')

            total_jadwal_dibuat = 0
            for item in sumber_jadwal:
                # Cari instalasi aktif pegawai saat ini (bisa saja sudah pindah)
                riwayat_aktif = RiwayatPenempatan.objects.filter(
                    pegawai=item.pegawai, status=True
                ).first()

                instalasi_aktif = riwayat_aktif.penempatan_level4 if riwayat_aktif else item.instalasi

                tujuan_obj, _ = JenisSDMPerinstalasi.objects.get_or_create(
                    pegawai=item.pegawai,
                    instalasi=instalasi_aktif,
                    bulan=bulan_tujuan,
                    tahun=tahun_tujuan,
                    defaults={
                        'unor': item.unor,
                        'bidang': item.bidang,
                        'sub_bidang': item.sub_bidang,
                        'jenis_sdm': item.jenis_sdm,
                    }
                )

                # Hapus jadwal lama tujuan
                JadwalDinasSDM.objects.filter(pegawai=tujuan_obj).delete()

                # Persiapkan bulk_create
                daftar_jadwal = []
                for jadwal in item.jadwaldinassdm_set.all():
                    try:
                        tanggal_baru = jadwal.tanggal.replace(month=bulan_tujuan, year=tahun_tujuan)
                        daftar_jadwal.append(JadwalDinasSDM(
                            pegawai=tujuan_obj,
                            tanggal=tanggal_baru,
                            kategori_jadwal=jadwal.kategori_jadwal
                        ))
                    except ValueError:
                        continue  # lewati misal 31 Feb

                JadwalDinasSDM.objects.bulk_create(daftar_jadwal)
                total_jadwal_dibuat += len(daftar_jadwal)

            messages.success(request, f"Berhasil menyalin {total_jadwal_dibuat} entri jadwal ke bulan {bulan_tujuan}/{tahun_tujuan}.")
            return redirect(self.get_success_url())
        messages.error(request, 'Form tidak valid. Silahkan periksa kembali.')
        return redirect(self.get_failure_url())
    

class JadwalUpdateView(LoginRequiredMixin, UserPassesTestMixin, generic.UpdateView):
    model = JenisSDMPerinstalasi
    login_url = reverse_lazy('myaccount_urls:login_view')
    redirect_field_name='next'
    form_class = JenisSDMPerinstalasiBasicForm
    template_name = 'kehadirankegiatan/form.html'
    
    def test_func(self):
        if self.request.user.is_staff or self.request.user.is_disiplin_admin:
            return True
        return False

    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        messages.error(self.request, 'Anda tidak memiliki izin untuk edit data.')
        return redirect(self.get_success_url())
    
    def get_success_url(self):
        query_params = self.request.GET.copy()
        if query_params:  # hanya jika ada query string
            return f'{reverse("disiplinsdm_urls:jadwal_list")}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_list')

    def form_invalid(self, form):
        messages.error(self.request, 'Maaf data gagal disimpan!')
        # Simpan query params untuk mengembalikan ke halaman yang sama
        query_params = self.request.GET.copy()
        context = self.get_context_data(form=form)
        context['query_params'] = query_params
        return self.render_to_response(context)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit jadwal pegawai'
        context['url'] =reverse('disiplinsdm_urls:jadwal_list')
        return context
    
    
class DeleteJadwalView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        if self.request.user.is_staff or self.request.user.is_disiplin_admin:
            return True
        return False
    
    def get_success_url(self):
        query_params = self.request.GET.copy()
        if query_params:  # hanya jika ada query string
            return f'{reverse("disiplinsdm_urls:jadwal_list")}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_list')

    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        messages.error(self.request, 'Anda tidak memiliki izin untuk menghapus data.')
        return redirect(self.get_success_url())
    
    def get_object(self, id):
        try:
            data = JenisSDMPerinstalasi.objects.get(id=id)
            return data
        except JenisSDMPerinstalasi.DoesNotExist:
            return None
        
    def get(self, request, *args, **kwargs):
        id_jadwal = kwargs.get('id')
        instance = self.get_object(id_jadwal)
        context={
            'delete_jadwal':True,
            'data':instance,
            'url':reverse_lazy('disiplinsdm_urls:jadwal_list'),
            'title': 'Delete jadwal pegawai',
            'sub_page':'Riwayat',
            'title_page':'disiplin',
            'riwayat':'active',
            'selected':'disiplin'
        }
        return render(request, 'jadwal_piket/validasi_delete_jadwal.html', context)
    
    def post(self, request, **kwargs):
        id_jadwal = kwargs.get('id')
        instance = self.get_object(id_jadwal)
        instance.delete()
        return redirect(self.get_success_url())
    
    
class VerifikasiJadwalView(FormView):
    template_name = 'jadwal_piket/verifikasi_jadwal.html'
    form_class = PersetujuanForm

    def dispatch(self, request, *args, **kwargs):
        self.pengajuan = get_object_or_404(JenisSDMPerinstalasi, pk=self.kwargs['pk'], status='diajukan')
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['initial'] = {
            'alasan': self.pengajuan.alasan_penolakan if self.pengajuan.alasan_penolakan else ''
        }
        return kwargs
    
    def get_success_url(self):
        query_params = self.request.GET.copy()
        if query_params:  # hanya jika ada query string
            return f'{reverse("disiplinsdm_urls:jadwal_list")}?{query_params.urlencode()}'
        return reverse('disiplinsdm_urls:jadwal_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pengajuan = self.pengajuan

        draft_set = JadwalDinasSDM.objects.filter(pegawai=pengajuan)
        approved_set = ApprovedJadwalDinasSDM.objects.filter(pegawai=pengajuan)
        approved_set = ApprovedJadwalDinasSDM.objects.filter(pegawai=pengajuan)
        total_jam_disetujui = 0

        for jadwal in approved_set:
            kategori = jadwal.kategori_jadwal
            if kategori and kategori.waktu_datang and kategori.waktu_pulang:
                datang = datetime.combine(datetime.today(), kategori.waktu_datang)
                pulang = datetime.combine(datetime.today(), kategori.waktu_pulang)
                if pulang < datang:
                    # Shift malam, tambahkan 1 hari ke pulang
                    pulang += timedelta(days=1)
                durasi = pulang - datang
                total_jam_disetujui += durasi.total_seconds() / 3600  # konversi ke jam

        tanggal_list = sorted(set(draft_set.values_list('tanggal', flat=True)) | set(approved_set.values_list('tanggal', flat=True)))
        perbandingan = []

        for tanggal in tanggal_list:
            draft = draft_set.filter(tanggal=tanggal).first()
            approved = approved_set.filter(tanggal=tanggal).first()
            perbandingan.append({
                'tanggal': tanggal,
                'jadwal_draft': draft.kategori_jadwal.kategori_jadwal if draft and draft.kategori_jadwal else '-',
                'jadwal_approved': approved.kategori_jadwal.kategori_jadwal if approved and approved.kategori_jadwal else '-',
                'status': '✅ Sama' if draft and approved and draft.kategori_jadwal == approved.kategori_jadwal else '❌ Berbeda',
                'catatan': draft.catatan or approved.catatan or '-'
            })

        context.update({
            'title': f'Verifikasi Jadwal: { pengajuan.pegawai.full_name } ({ pengajuan.bulan }/{ pengajuan.tahun })',
            'riwayat': 'active',
            'selected': 'disiplin',
            'url': reverse('disiplinsdm_urls:jadwal_list'),
            'pengajuan': pengajuan,
            'perbandingan_jadwal': perbandingan,
            'jam_aktual': pengajuan.kurang_lebih_jam_kerja,
            'jam_min': pengajuan.standar_min_efektif,
            'jam_max': pengajuan.standar_max_efektif,
            'selisih': pengajuan.selisih_jam_kerja,
            'jam_disetujui': round(total_jam_disetujui, 1),
        })
        return context

    def form_valid(self, form):
        alasan = form.cleaned_data['alasan']
        
        aksi = self.request.POST.get("aksi")
        if aksi == 'tolak':
            if not alasan:
                messages.error(self.request, "Mohon isi alasan penolakan jika ingin menolak jadwal.")
                return super().form_invalid(form)

            self.pengajuan.status = 'ditolak'
            self.pengajuan.alasan_penolakan = alasan
            self.pengajuan.save()

            messages.success(self.request, f"Jadwal berhasil ditolak.")
            return super().form_valid(form)

        elif aksi == 'setujui':
            with transaction.atomic():
                self.pengajuan.status = 'disetujui'
                self.pengajuan.alasan_penolakan = alasan
                self.pengajuan.save()
                for jadwal in JadwalDinasSDM.objects.filter(pegawai=self.pengajuan):
                    ApprovedJadwalDinasSDM.objects.update_or_create(
                        pegawai=self.pengajuan,
                        tanggal=jadwal.tanggal,
                        defaults={
                            'kategori_jadwal':jadwal.kategori_jadwal, 
                            'catatan':jadwal.catatan,
                            'is_approved':True,
                            'approved_by':self.request.user,
                        }
                    )
            return super().form_valid(form)
        
    def form_invalid(self, form):
        print("Form invalid")
        print(form.errors)
        return super().form_invalid(form)


class KehadiranSpesialisListView(LoginRequiredMixin, generic.ListView):
    login_url = reverse_lazy('myaccount_urls:login_view')
    redirect_field_name = 'next'
    model = DaftarKegiatanPegawai
    template_name = 'kehadirankegiatan/kehadiran_spesialis_list.html'
    
    def get_date_params(self):
        """Parse tanggal dari query string dan kembalikan bulan dan tahun."""
        tgl = self.request.GET.get('tanggal')
        get_tanggal = get_date_from_string(tgl)
        return get_tanggal, get_tanggal.month, get_tanggal.year
    
    def get_instalasi(self):
        """Ambil dan validasi parameter instalasi dari query string."""
        inst_id = self.request.GET.get('inst')
        if inst_id and inst_id.strip():
            try:
                return UnitInstalasi.objects.get(id=inst_id)
            except UnitInstalasi.DoesNotExist:
                return None
        return None
    
    def get_queryset_for_user(self, bulan, tahun, instalasi):
        """Kembalikan queryset berdasarkan role user."""
        base_filter = {
            'bulan': bulan,
            'tahun': tahun,
            'pegawai__profil_user__is_dokter_spesialis':True,
        }
        user = self.request.user
        # Jika user adalah superuser, filter berdasarkan instalasi jika ada
        if user.is_disiplin_admin:
            if instalasi:
                base_filter['instalasi'] = instalasi
            return DaftarKegiatanPegawai.objects.filter(**base_filter)

        elif user.is_staff:
            profil = user.profil_admin if hasattr(user, 'profil_admin') else None
            if instalasi:
                if isinstance(instalasi, QuerySet):
                    instalasi = instalasi.pk
                    base_filter['instalasi'] = instalasi
                else:
                    base_filter['instalasi'] = instalasi
            elif profil and profil.instalasi.exists():
                base_filter['instalasi__in'] = profil.instalasi.values_list('pk', flat=True)
            elif profil and profil.sub_bidang.exists():
                base_filter['sub_bidang__in'] = profil.sub_bidang.values_list('pk', flat=True)
            elif profil and profil.bidang.exists():
                base_filter['bidang__in'] = profil.bidang.values_list('pk', flat=True)
            return DaftarKegiatanPegawai.objects.filter(**base_filter)

        # Default untuk user biasa (pegawai)
        base_filter['pegawai'] = user
        return DaftarKegiatanPegawai.objects.filter(**base_filter)

    def get_queryset(self):
        tanggal, bulan, tahun = self.get_date_params()
        instalasi = self.get_instalasi()
        queryset = self.get_queryset_for_user(bulan, tahun, instalasi)
        return queryset.order_by('pegawai__first_name', 'pegawai__last_name')
    
    def get_instalasi_list(self):
        instalasi = None
        user = self.request.user
        if user.is_disiplin_admin:
            instalasi = UnitInstalasi.objects.filter(jenissdmperinstalasi__isnull=False).order_by('instalasi').distinct()

        elif user.is_staff:
            profil = user.profil_admin if hasattr(user, 'profil_admin') else None
            if profil and profil.instalasi.exists():
                instalasis = profil.instalasi.values_list('pk', flat=True)
                instalasi = UnitInstalasi.objects.filter(pk__in=instalasis)
            elif profil and profil.sub_bidang.exists():
                sub_bidang_ids = profil.sub_bidang.values_list('pk', flat=True)
                instalasi = UnitInstalasi.objects.filter(sub_bidang__in=sub_bidang_ids)
            elif profil and profil.bidang.exists():
                instalasi = UnitInstalasi.objects.filter(sub_bidang__bidang__in=profil.bidang.values_list('pk', flat=True))
        return instalasi
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tanggal, bulan, tahun = self.get_date_params()
        context.update({
            'instalasi_list': self.get_instalasi_list(),
            'bulan': bulan,
            'tahun': tahun,
            'title': 'Daftar Kehadiran Pegawai',
            'page': 'Home',
            'sub_page': 'Riwayat',
            'title_page': 'Disiplin',
            'riwayat': 'active',
            'selected': 'disiplin',
            'url': reverse('disiplinsdm_urls:jadwal_list'),
        })
        return context

    
class PenilaianKehadiranApelPagi(FormView):
    template_name = "kehadirankegiatan/form.html"
    form_class = ProsesKehadiranForm
    success_url = reverse_lazy('disiplinsdm_urls:kehadiran_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': 'Proses Penilaian Kehadiran Apel Pagi',
            'selected': 'disiplin',
            'url': self.success_url,
        })
        return context

    def form_valid(self, form):
        tanggal = form.cleaned_data['tanggal']
        # Panggil Service
        jumlah = ApelPagiService.proses_penilaian_apel_massal(tanggal)

        messages.success(self.request, f"Berhasil memproses {jumlah} pegawai yang apel pagi pada {tanggal}.")
        return super().form_valid(form)


class KehadiranUtamaView(LoginRequiredMixin, BridgeSyncService, View):
    login_url = reverse_lazy('myaccount_urls:login_view')
    model = DaftarKegiatanPegawai
    success_url = reverse_lazy('disiplinsdm_urls:kehadiran_list') # Sesuaikan name url view ini
    
    def get_date_params(self):
        """Mengurai tanggal dari query string dan mengembalikan bulan dan tahun."""
        tgl = self.request.GET.get('tanggal')
        get_tanggal = get_date_from_string(tgl)
        return get_tanggal, get_tanggal.month, get_tanggal.year

    def get_instalasi(self):
        """Mengambil dan memvalidasi parameter instalasi dari query string."""
        inst_id = self.request.GET.get('inst')
        if inst_id and inst_id.strip():
            try:
                return UnitInstalasi.objects.get(id=inst_id)
            except UnitInstalasi.DoesNotExist:
                return None
        return None

    def get_queryset_for_user(self, bulan, tahun, instalasi):
        """Mengembalikan queryset dasar berdasarkan role user dan filter."""
        base_filter = {
            'bulan': bulan,
            'tahun': tahun,
            'kegiatan__slug': 'absen-datang',
            'pegawai__profil_user__is_dokter_spesialis': False,
        }
        
        user = self.request.user
        if user.is_disiplin_admin:
            if instalasi:
                base_filter['instalasi'] = instalasi
            return self.model.objects.filter(**base_filter)

        elif user.is_staff and hasattr(user, 'profil_admin'):
            profil = user.profil_admin
            if instalasi:
                 base_filter['instalasi'] = instalasi
            elif profil.instalasi.exists():
                base_filter['instalasi__in'] = profil.instalasi.all()
            elif profil.sub_bidang.exists():
                base_filter['sub_bidang__in'] = profil.sub_bidang.values_list('pk', flat=True)
            elif profil.bidang.exists():
                base_filter['bidang__in'] = profil.bidang.values_list('pk', flat=True)
            return self.model.objects.filter(**base_filter)

        base_filter['pegawai'] = user
        return self.model.objects.filter(**base_filter)

    def get_instalasi_list(self):
        """Menyediakan daftar instalasi untuk filter dropdown berdasarkan hak akses."""
        user = self.request.user
        if user.is_disiplin_admin:
            return UnitInstalasi.objects.filter(jenissdmperinstalasi__isnull=False).order_by('instalasi').distinct()

        elif user.is_staff and hasattr(user, 'profil_admin'):
            profil = user.profil_admin
            if profil.instalasi.exists():
                return profil.instalasi.all().order_by('instalasi')
            elif profil.sub_bidang.exists():
                return UnitInstalasi.objects.filter(sub_bidang__in=profil.sub_bidang.values_list('pk', flat=True))
            elif profil.bidang.exists():
                return UnitInstalasi.objects.filter(sub_bidang__bidang__in=profil.bidang.values_list('pk', flat=True))
        return None

    def get(self, request, *args, **kwargs):
        """Menangani pemuatan halaman daftar kehadiran, filter, pencarian, dan export excel."""
        tanggal, bulan, tahun = self.get_date_params()
        instalasi = self.get_instalasi()
        base_queryset = self.get_queryset_for_user(bulan, tahun, instalasi)

        # --- Subquery Definitions ---
        def create_count_subquery(filter_kwargs):
            return KehadiranKegiatan.objects.filter(
                pegawai_id=OuterRef('pk'), **filter_kwargs
            ).values('pegawai_id').annotate(
                total=Count(TruncDate('tanggal'), distinct=True)
            ).values('total')

        hadir_subquery = create_count_subquery({'hadir': True})
        tidak_hadir_subquery = create_count_subquery({'hadir': False})
        izin_subquery = create_count_subquery({'alasan__alasan__iexact': 'Izin'})
        sakit_subquery = create_count_subquery({'alasan__alasan__iexact': 'Sakit'})
        
        status_terlambat = ['Terlambat Ringan', 'Terlambat Sedang', 'Terlambat Berat', 'Terlambat']
        terlambat_subquery = create_count_subquery({'status_ketepatan__in': status_terlambat})

        # --- Terapkan Anotasi ---
        annotated_queryset = base_queryset.annotate(
            rekap_hadir=Coalesce(Subquery(hadir_subquery, output_field=IntegerField()), 0),
            rekap_tidak_hadir=Coalesce(Subquery(tidak_hadir_subquery, output_field=IntegerField()), 0),
            rekap_terlambat=Coalesce(Subquery(terlambat_subquery, output_field=IntegerField()), 0),
            rekap_izin=Coalesce(Subquery(izin_subquery, output_field=IntegerField()), 0),
            rekap_sakit=Coalesce(Subquery(sakit_subquery, output_field=IntegerField()), 0)
        ).select_related('pegawai', 'instalasi').order_by('pegawai__first_name', 'pegawai__last_name')

        # --- LOGIKA EXPORT EXCEL (Jika dipicu) ---
        if request.GET.get('export') == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rekap Kehadiran"
            ws.append(['No', 'Nama Pegawai', 'Instalasi', 'Hadir', 'Terlambat', 'Absen', 'Izin', 'Sakit'])
            
            for index, item in enumerate(annotated_queryset, start=1):
                ws.append([
                    index, item.pegawai.get_full_name() if item.pegawai else '',
                    item.instalasi.instalasi if item.instalasi else '',
                    item.rekap_hadir, item.rekap_terlambat, item.rekap_tidak_hadir,
                    item.rekap_izin, item.rekap_sakit
                ])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Rekap_Kehadiran_{bulan}_{tahun}.xlsx"'
            wb.save(response)
            return response

        # --- INTEGRASI PAGINASI MANUAL ---
        paginate_by = 25
        paginator = Paginator(annotated_queryset, paginate_by)
        page_number = request.GET.get('page', 1)
        
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Pertahankan parameter filter untuk navigasi link halaman template
        preserved_filters = request.GET.copy()
        if 'page' in preserved_filters:
            del preserved_filters['page']
        if 'export' in preserved_filters:
            del preserved_filters['export']

        context = {
            'object_list': page_obj.object_list, # Kompatibel dengan template ListView lama
            'page_obj': page_obj,
            'is_paginated': page_obj.has_other_pages(),
            'form': PenilaianKehadiranForm(), # Form penilaian manual disatukan di sini
            'instalasi_list': self.get_instalasi_list(),
            'bulan': bulan,
            'tahun': tahun,
            'title': 'Daftar Kehadiran & Penilaian Pegawai',
            'riwayat': 'active',
            'selected': 'disiplin',
            'preserved_filters': preserved_filters.urlencode()
        }
        return render(request, 'kehadirankegiatan/kehadiran_list.html', context)

    def post(self, request, *args, **kwargs):
        """Menangani eksekusi pengiriman data form penilaian massal."""
        form = PenilaianKehadiranForm(request.POST)
        if form.is_valid():
            tanggal = form.cleaned_data['tanggal']
            
            # Eksekusi Service Penilaian Massal
            jumlah = KehadiranService.proses_kehadiran_massal(tanggal)
            
            messages.success(
                request, 
                f"Berhasil memproses {jumlah} pegawai yang memiliki jadwal pada {tanggal}."
            )
            # Dapatkan parameter filter aktif saat ini dari URL asal / HTTP_REFERER
            # agar filter user tidak hilang setelah form di-submit
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', self.success_url))
        
        # Jika form tidak valid, arahkan kembali dengan pesan error
        messages.error(request, "Gagal memproses penilaian. Pastikan format tanggal benar.")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', self.success_url))


class DetailKehadiranView(LoginRequiredMixin, generic.ListView):
    login_url = reverse_lazy('myaccount_urls:login_view')
    model = DaftarKegiatanPegawai
    template_name = 'kehadirankegiatan/kehadiran_detail_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pegawai_id = self.kwargs.get('pk')
        queryset = super().get_queryset().filter(pegawai_id=pegawai_id)
        pegawai = queryset.annotate(tgl=TruncDate('kehadirankegiatan__tanggal'))
        
        riwayat_kehadiran = (
            pegawai.values('tgl')  # Group by berdasarkan Tanggal
            .annotate(
                # Solusi: Definisikan jumlah_record di awal agar bisa dibaca oleh Case/When di bawahnya
                jumlah_record=Count('kehadirankegiatan__id'),
                
                # Ambil jam terkecil sebagai jam datang jika ada yang hadir
                jam_datang=Case(
                    When(kehadirankegiatan__hadir=True, then=Min('kehadirankegiatan__tanggal__time')),
                    default=Value(None, output_field=TimeField())
                ),

                # Gunakan alias 'jumlah_record' yang sudah didefinisikan di atas
                jam_pulang=Case(
                    When(
                        kehadirankegiatan__hadir=True, 
                        jumlah_record__gt=1,  # <--- SEKARANG AMAN, Menggunakan alias yang valid
                        then=Max('kehadirankegiatan__tanggal__time')
                    ),
                    default=Value(None, output_field=TimeField())
                ),

                status=Case(
                    When(kehadirankegiatan__hadir=True, then=Value('Hadir')),
                    default=Value('Tidak Hadir'),
                    output_field=CharField()
                ),

                ketepatan=Min('kehadirankegiatan__status_ketepatan'),
                alasan=Max('kehadirankegiatan__alasan__alasan'),
                keterangan=Max('kehadirankegiatan__ket'),
            )
            .order_by('-tgl')
        )

        context['riwayat_kehadiran'] = riwayat_kehadiran
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        context['title'] = 'Detail Kehadiran Pegawai'
        context['url'] = reverse_lazy('disiplinsdm_urls:kehadiran_list')
        
        return context

    
class KehadiranCreateView(LoginRequiredMixin, UserPassesTestMixin, generic.CreateView):
    model = DaftarKegiatanPegawai
    form_class = DaftarKegiatanPegawaiForm
    template_name = 'kehadirankegiatan/kehadiran_form.html'
    login_url = reverse_lazy('myaccount_urls:login_view')
    redirect_field_name = 'next'
    # success_url = reverse_lazy('disiplinsdm_urls:kehadiran_list')
    
    def test_func(self):
        if self.request.user.is_disiplin_admin:
            return True
        return False

    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        messages.error(self.request, 'Anda tidak memiliki izin untuk menambah kehadiran data.')
        return redirect(reverse('disiplinsdm_urls:kehadiran_list'))
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        tanggal = date.today()
        tgl = self.request.GET.get('tanggal')
        if tgl is not None:
            tanggal = get_date_from_string(tgl)
        if self.request.user.is_disiplin_admin:
            kwargs['request'] = self.request
            kwargs['tanggal'] = tanggal
        else:
            kwargs['initial'] = {
                'pegawai': self.request.user,
            }
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        formset = kehadiran_formset(data=self.request.POST or None)
        context['formset']=formset
        context['url'] = reverse_lazy('disiplinsdm_urls:kehadiran_list')
        context['title'] = 'Daftar Kehadiran Pegawai'
        context['page'] = 'Home'
        context['sub_page'] = 'Riwayat'
        context['title_page'] = 'Disiplin'
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                formset.instance = self.object
                formset.save()
                messages.success(self.request, 'Data berhasil disimpan!')
            return super().form_valid(form)
        else:
            messages.error(self.request, 'Maaf data gagal disimpan!')
            return self.form_invalid(form)
        
    def get_success_url(self):
        url = reverse('disiplinsdm_urls:kehadiran_update', kwargs={'pk':self.object.pk})
        return url
        
        
class KehadiranUpdateView(LoginRequiredMixin, UserPassesTestMixin, generic.UpdateView):
    model = DaftarKegiatanPegawai
    form_class = DaftarKegiatanPegawaiForm
    template_name = 'kehadirankegiatan/kehadiran_update_form.html'
    login_url = reverse_lazy('myaccount_urls:login_view')
    redirect_field_name = 'next'
    
    def test_func(self):
        if self.request.user.is_disiplin_admin:
            return True
        return False

    def handle_no_permission(self):
        # Bisa redirect atau tampilkan pesan khusus
        messages.error(self.request, 'Anda tidak memiliki izin untuk mengedit kehadiran data.')
        return redirect(reverse('disiplinsdm_urls:kehadiran_list'))
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        tanggal = date.today()
        tgl = self.request.GET.get('tanggal')
        if tgl is not None:
            tanggal = get_date_from_string(tgl)
        if self.request.user.is_disiplin_admin:
            kwargs['request'] = self.request
            kwargs['tanggal'] = tanggal
        else:
            kwargs['initial'] = {
                'pegawai': self.request.user,
            }
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        formset = kehadiran_formset(data=self.request.POST or None, instance=self.object)
        context['formset']=formset
        context['url'] = reverse_lazy('disiplinsdm_urls:kehadiran_list')
        context['title'] = 'Daftar Kehadiran Pegawai'
        context['page'] = 'Home'
        context['sub_page'] = 'Riwayat'
        context['title_page'] = 'Disiplin'
        context['riwayat'] = 'active'
        context['selected'] = 'disiplin'
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                formset.instance = self.object
                formset.save()
                messages.success(self.request, 'Data berhasil disimpan!')
            return super().form_valid(form)
        else:
            messages.error(self.request, 'Maaf data gagal disimpan!')
            return self.form_invalid(form)
        
    def form_invalid(self, form):
        print('error form: ', form.errors)
        messages.error(self.request, 'Maaf data gagal tersimpan!')
        return super().form_invalid(form)
        
    def get_success_url(self):
        url = reverse('disiplinsdm_urls:kehadiran_update', kwargs={'pk':self.object.pk})
        return url


class RekapPiketListView(generic.ListView):
    model = JadwalDinasSDM
    template_name = 'jadwal_piket/rekap_piket.html'
    context_object_name = 'daftar_jadwal'
    paginated_by = 20

    def get_active_instalasi(self):
        """Logika penentuan penempatan sesuai hierarki"""
        inst_param = self.request.GET.get('inst')
        if inst_param:
            return UnitInstalasi.objects.filter(pk=inst_param).first()

        user = self.request.user
        profil = getattr(user, 'profil_admin', None)

        if user.is_disiplin_admin:
            return None # Superuser default melihat semua jika tanpa param
        
        if profil:
            if profil.instalasi.exists():
                return profil.instalasi.first()
            if profil.sub_bidang.exists():
                return profil.sub_bidang.first()
            if profil.bidang.exists():
                return profil.bidang.first()
            if profil.unor.exists():
                return profil.unor.first()
        
        # Pegawai biasa
        penempatan = user.riwayat_penempatan.filter(status=True).first()
        return penempatan.penempatan_level4 if penempatan else None

    def get_queryset(self):
        tanggal_raw = self.request.GET.get('tanggal')
    
        if tanggal_raw:
            # Ambil hanya 10 karakter pertama (YYYY-MM-DD) 
            # untuk membuang sampah seperti '?detail=Pagi' jika terselip
            clean_date = tanggal_raw[:10] 
            try:
                self.target_date = date.fromisoformat(clean_date)
            except ValueError:
                self.target_date = date.today()
        else:
            self.target_date = date.today()

        # Ambil filter detail jika ada (untuk memfilter tabel)
        self.filter_detail = self.request.GET.get('detail') # Pagi, Siang, dll
        self.obj_penempatan = self.get_active_instalasi()

        # 2. Query Dasar
        queryset = JadwalDinasSDM.objects.filter(tanggal=self.target_date)

        # 3. Filter Berdasarkan Hierarki Penempatan
        if self.obj_penempatan:
            if isinstance(self.obj_penempatan, UnitInstalasi):
                queryset = queryset.filter(pegawai__instalasi=self.obj_penempatan)
            elif isinstance(self.obj_penempatan, SubBidang):
                queryset = queryset.filter(pegawai__sub_bidang=self.obj_penempatan)
            elif isinstance(self.obj_penempatan, Bidang):
                queryset = queryset.filter(pegawai__bidang=self.obj_penempatan)
            # Anda bisa tambah elif untuk Unor jika perlu
        elif not self.request.user.is_disiplin_admin:
            # Jika user biasa tanpa data penempatan, hanya lihat jadwal sendiri
            queryset = queryset.filter(pegawai__pegawai=self.request.user)
            
        if self.filter_detail == 'Pagi':
            queryset = queryset.filter(kategori_jadwal__kategori_jadwal__icontains='pagi')
        elif self.filter_detail == 'Siang':
            queryset = queryset.filter(kategori_jadwal__kategori_jadwal__icontains='siang')
        elif self.filter_detail == 'Malam':
            queryset = queryset.filter(kategori_jadwal__kategori_jadwal__icontains='malam')
        elif self.filter_detail == 'Middle':
            queryset = queryset.filter(kategori_jadwal__kategori_jadwal__icontains='middle')
        elif self.filter_detail == 'Lain-lain':
            queryset = queryset.exclude(
                Q(kategori_jadwal__kategori_jadwal__icontains='pagi') |
                Q(kategori_jadwal__kategori_jadwal__icontains='siang') |
                Q(kategori_jadwal__kategori_jadwal__icontains='malam') |
                Q(kategori_jadwal__kategori_jadwal__icontains='middle')
            )

        return queryset.select_related('pegawai__pegawai', 'kategori_jadwal')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qs = self.get_queryset()

        # LOGIKA GROUPING SHIFT (URUTAN SANGAT PENTING)
        rekap_piket = (
            qs.annotate(
                grup_shift=Case(
                    When(kategori_jadwal__kategori_jadwal__icontains='pagi', then=Value('Pagi')),
                    When(kategori_jadwal__kategori_jadwal__icontains='siang', then=Value('Siang')),
                    When(kategori_jadwal__kategori_jadwal__icontains='malam', then=Value('Malam')),
                    When(kategori_jadwal__kategori_jadwal__icontains='middle', then=Value('Middle')),
                    default=Value('Lain-lain'),
                    output_field=CharField(),
                )
            )
            .values('grup_shift') # Group By Alias
            .annotate(total=Count('id')) # Count per Group
            .order_by('grup_shift')
        )
        
        # Ambil semua parameter URL saat ini (copy agar bisa dimodifikasi)
        querydict = self.request.GET.copy()
        querydict.pop('page', None)

        context.update({
            
            'rekap_piket': rekap_piket,
            'total_pegawai_masuk': qs.count(),
            'target_date': self.target_date,
            'active_loc': self.obj_penempatan,
            'all_instalasi': UnitInstalasi.objects.all() if self.request.user.is_disiplin_admin else None,
            'preserved_query': querydict.urlencode(),
            'selected': 'disiplin',
            'riwayat': 'active', 
        })
        return context
    
## LOGIKA PRESENSI MESIN BARU (HIKVISION)

class SinkronisasiLogView(View):
    """
    Generic View untuk memproses sinkronisasi saat tombol diklik.
    Hanya menerima request POST untuk keamanan.
    """
    
    def post(self, request, *args, **kwargs):
        result = self._proses_sinkronisasi()
        
        # Simpan hasil statistik ke session untuk ditampilkan di halaman result
        request.session['sinkronisasi_result'] = result
        
        # Set flash message berdasarkan hasil
        if result['new_data'] > 0:
            messages.success(request, f"Berhasil menyimpan {result['new_data']} data fingerprint baru.")
        elif result['total_fetched'] > 0 and result['skipped_exists'] > 0:
            messages.info(request, "Semua data sudah ter-sync. Tidak ada data baru.")
            
        if result['skipped_no_mapping'] > 0:
            messages.warning(request, f"{result['skipped_no_mapping']} data dilewati karena ID mesin belum ada di tabel Mapping.")
            
        if result['errors']:
            messages.error(request, f"Terjadi {len(result['errors'])} error saat proses sinkronisasi.")
            
        return redirect('sinkronisasi-result')
    
    def _proses_sinkronisasi(self):
        stats = {
            'new_data': 0,
            'skipped_exists': 0,
            'skipped_no_mapping': 0,
            'total_fetched': 0,
            'errors': []
        }
        
        url = settings.API_FINGERPRINT_URL
        
        # LOOPING UNTUK HANDLE PAGINATION
        while url:
            try:
                response = requests.get(url, timeout=settings.API_FINGERPRINT_TIMEOUT)
                response.raise_for_status()
            except requests.RequestException as e:
                stats['errors'].append(f'Gagal koneksi ke API: {str(e)}')
                break
            
            data = response.json()
            results = data.get('results', [])
            
            # LOOPING SETIAP ITEM LOG
            for item in results:
                stats['total_fetched'] += 1
                
                # 1. Ambil ID Pegawai dari Mesin (Field "id" pada API)
                pegawai_mesin_id = str(item.get('id', '')).strip()
                if not pegawai_mesin_id:
                    continue
                
                # 2. Cari Mapping ke SIMADU berdasarkan ID Mesin
                try:
                    mapping = MappingMesinAbsensi.objects.select_related('pegawai').get(
                        mesin_id=pegawai_mesin_id
                    )
                except MappingMesinAbsensi.DoesNotExist:
                    stats['skipped_no_mapping'] += 1
                    continue
                
                # 3. Parse Datetime
                datetime_str = item.get('datetime', '')
                try:
                    parsed_datetime = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
                except (ValueError, TypeError):
                    stats['errors'].append(f'Format datetime error: {datetime_str}')
                    continue
                
                # 4. Ambil dan Validasi Direction
                direction = item.get('direction', '').strip().upper()
                if direction not in ['IN', 'OUT']:
                    continue
                
                # 5. CEK DUPLIKASI (Pegawai + Waktu Presisi Detik + Arah)
                if LogKehadiran.objects.filter(
                    mapping=mapping,
                    datetime=parsed_datetime,
                    direction=direction
                ).exists():
                    stats['skipped_exists'] += 1
                    continue
                
                # 6. INSERT DATA BARU SAJA
                try:
                    LogKehadiran.objects.create(
                        mapping=mapping,
                        datetime=parsed_datetime,
                        direction=direction,
                        devicename=item.get('devicename', '').strip(),
                        personname=item.get('personname', '').strip(),
                    )
                    stats['new_data'] += 1
                except IntegrityError:
                    # Double safety dari database constraint (jika ada race condition)
                    stats['skipped_exists'] += 1
                except Exception as e:
                    stats['errors'].append(f'Gagal simpan log (ID Mesin: {pegawai_mesin_id}): {str(e)}')
            
            # Pindah ke URL halaman berikutnya jika ada
            url = data.get('next')
            
        return stats


class SinkronisasiResultView(TemplateView):
    """
    Generic Template View untuk menampilkan halaman hasil sinkronisasi.
    """
    template_name = 'absensi/sinkronisasi_result.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ambil data statistik dari session, jika tidak ada kosongkan
        context['result'] = self.request.session.get('sinkronisasi_result', {})
        return context
    
# Dashboard pengecekan kehadiran pegawai berdasarkan data log kehadiran dari mesin baru
class LogKehadiranView(BridgeSyncService, View):
    def set_default_date(self):
        """Fungsi pembantu untuk menetapkan waktu default (Hari ini)"""
        today = timezone.now().date()
        self.hari = today.day
        self.bulan = today.month
        self.tahun = today.year
        self.full_date_str = today.strftime('%Y-%m-%d') # '2026-06-11'
        
    def get(self, request, *args, **kwargs):
        # 1. Ambil string tanggal dari input type="date" (Format: YYYY-MM-DD)
        date_param = self.request.GET.get('tanggal_pilih')
        
        if date_param:
            try:
                # Konversi string 'YYYY-MM-DD' menjadi objek datetime
                parsed_date = datetime.strptime(date_param, '%Y-%m-%d').date()
                self.hari = parsed_date.day
                self.bulan = parsed_date.month
                self.tahun = parsed_date.year
                self.full_date_str = date_param # Simpan untuk dioper ke template
            except ValueError:
                # Antisipasi jika format string tidak valid, fallback ke hari ini
                self.set_default_date()
        else:
            # Jika form baru pertama dimuat (kosong), fallback ke hari ini
            self.set_default_date()
            
        # Mapping nama bulan untuk judul file
        nama_bulan = {
            1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
            7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
        }.get(self.bulan, '')
            
        devicename = request.GET.get('devicename')
        pegawai = request.GET.get('pegawai')
        export_mode = request.GET.get('export') # Ambil parameter export
        
        # 1. Bangun QuerySet dasar (Berlaku untuk HTML maupun Excel)
        # ==============================================================
        # OPTIMASI QUERY DENGAN PREFETCH UNTUK PENEMPATAN AKTIF
        # ==============================================================
        # 1. Ambil hanya penempatan yang statusnya aktif, serta join level 1-4 nya
        active_penempatan_qs = RiwayatPenempatan.objects.filter(
            status=True
        ).select_related(
            'penempatan_level1', 'penempatan_level2', 
            'penempatan_level3', 'penempatan_level4'
        )

        # 2. Subquery untuk status evaluasi
        subquery_evaluasi = LogAktivitasAbsen.objects.filter(
            absensi_harian__pegawai_id=OuterRef('mapping__pegawai_id'),
            waktu=OuterRef('datetime')
        )
        
        # 3. Bangun Queryset Utama
        queryset = LogKehadiran.objects.select_related(
            'mapping__pegawai'
        ).prefetch_related(
            # Gunakan Prefetch untuk menyimpan penempatan aktif ke atribut custom _active_penempatans
            Prefetch('mapping__pegawai__riwayat_penempatan', queryset=active_penempatan_qs, to_attr='active_penempatans')
        ).annotate(
            is_evaluated=Exists(subquery_evaluasi)
        ).order_by('-datetime')
        
        if devicename:
            queryset = queryset.filter(devicename=devicename)
        if pegawai:
            queryset = queryset.filter(mapping__pegawai=pegawai)
        if self.hari:
            queryset = queryset.filter(datetime__day=self.hari)
        if self.bulan:
            queryset = queryset.filter(datetime__month=self.bulan)
        if self.tahun:
            queryset = queryset.filter(datetime__year=self.tahun)
            
        # 2. LOGIKA EXPORT EXCEL (Jika ?export=excel ada di URL)
        if export_mode == 'excel':
            # Membuat workbook excel baru di memori
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Log Kehadiran"
            
            # Mengaktifkan gridlines agar garis tabel bawaan excel tetap terlihat
            ws.views.sheetView[0].showGridLines = True

            # Styles tambahan untuk status evaluasi (Opsional agar makin rapi)
            font_sudah = Font(name='Arial', size=10, color='1E4620', bold=True) # Hijau Tua
            fill_sudah = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Hijau Lembut
            
            font_belum = Font(name='Arial', size=10, color='721C24') # Merah Tua
            fill_belum = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid') # Merah Lembut
            
            # Styles komponen tabel
            font_judul = Font(name='Arial', size=14, bold=True)
            font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            font_data = Font(name='Arial', size=10)
            
            fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Biru Navy Corporate
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left = Alignment(horizontal='left', vertical='center')
            
            border_tipis = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            # Menulis Judul Dokumen (Baris 1 & 2)
            ws['A1'] = "LOG PRESENSI DARI MESIN ABSENSI"
            ws['A1'].font = font_judul
            ws['A2'] = f"Periode: {nama_bulan} {self.tahun}"
            ws['A2'].font = Font(name='Arial', size=11, italic=True)
            
            # TAMBAHKAN KOLOM PENEMPATAN DI HEADER
            headers = ['No', 'Nama Pegawai', 'Penempatan', 'Waktu/Datetime', 'Arah (IN/OUT)', 'Nama Perangkat', 'Status Evaluasi']
            ws.append([]) 
            ws.append(headers) 
            
            for cell in ws[4]:
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
            ws.row_dimensions[4].height = 25
            
            for index, log in enumerate(queryset, start=1):
                dt_naive = timezone.make_naive(log.datetime) if timezone.is_aware(log.datetime) else log.datetime
                dt_str = dt_naive.strftime('%Y-%m-%d %H:%M:%S')
                status_bool = getattr(log, 'is_evaluated', False)
                status_teks = "Sudah Dinilai" if status_bool else "Belum Dinilai"
                
                # MENGAMBIL PENEMPATAN UNTUK EXCEL
                penempatan_str = "N/A"
                if hasattr(log.mapping.pegawai, 'active_penempatans') and log.mapping.pegawai.active_penempatans:
                    # Ambil property penempatan dari objek RiwayatPenempatan yang aktif
                    penempatan_str = log.mapping.pegawai.active_penempatans[0].penempatan
                
                ws.append([
                    index,
                    log.personname,
                    penempatan_str, # Sisipkan di sini
                    dt_str,
                    log.direction,
                    log.devicename,
                    status_teks
                ])
                
                current_row = ws.max_row
                ws.row_dimensions[current_row].height = 20
                
                for col_idx, cell in enumerate(ws[current_row], start=1):
                    cell.font = font_data
                    cell.border = border_tipis
                    if col_idx in [1, 5]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
                        
                    if col_idx == 7: # Kolom status bergeser ke 7
                        if status_bool:
                            cell.font = font_sudah
                            cell.fill = fill_sudah
                        else:
                            cell.font = font_belum
                            cell.fill = fill_belum

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row < 4: continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Log_Kehadiran_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            return response
        
        devicenames = queryset.filter(devicename__isnull=False).values_list('devicename', flat=True).order_by().distinct()
        pegawais = queryset.filter(mapping__pegawai__isnull=False).values('mapping__pegawai__id', 'mapping__pegawai__first_name', 'mapping__pegawai__last_name').order_by().distinct()
        
        # =================================================================
        # AMBIL DATA MASTER STRUKTUR ORGANISASI UNTUK TOMBOL LANGKAH 2
        # =================================================================
        master_instalasi = UnitInstalasi.objects.all().order_by('instalasi')
        master_sub_bidang = SubBidang.objects.all().order_by('sub_bidang')
        master_bidang = Bidang.objects.all().order_by('bidang')
        master_unor = UnitOrganisasi.objects.all().order_by('unor') # jika ada tingkat Unor
        
        # 4. PAGINASI
        items_per_page = 50  
        paginator = Paginator(queryset, items_per_page)
        page_number = request.GET.get('page', 1)
        
        try:
            log_kehadiran_page = paginator.page(page_number)
        except PageNotAnInteger:
            log_kehadiran_page = paginator.page(1)
        except EmptyPage:
            log_kehadiran_page = paginator.page(paginator.num_pages)
            
        # 5. MASUKKAN KE CONTEXT TEMPLATE
        context = {
            'current_date_filter': self.full_date_str,
            'form': ProsesKehadiranForm(),
            'log_kehadiran_list': log_kehadiran_page, 
            'title': 'Log Kehadiran dari Mesin Absensi',
            'selected': 'disiplin',
            'riwayat': 'active',
            'devicenames': devicenames,
            'pegawais': pegawais,
            
            # Variabel context baru yang dibaca oleh template loop {% for inst in master_instalasi %}
            'master_instalasi': master_instalasi,
            'master_sub_bidang': master_sub_bidang,
            'master_bidang': master_bidang,
            'master_unor': master_unor,
            'title': 'Log Kehadiran Dari Mesin Absen',
            'title_page':'Log Kehadiran'
        }
        return render(request, 'kehadirankegiatan/log_kehadiran_list.html', context)
    
    
    def post(self, request, *args, **kwargs):
        form = ProsesKehadiranForm(request.POST)
        action = request.POST.get('action')
        
        # =================================================================
        # AKSI 1: TARIK DATA LOG DARI MESIN (Menggunakan form validasi tanggal default)
        # =================================================================
        if action == 'pull_bridge_logs' and form.is_valid():
            target_date = form.cleaned_data['tanggal']
            total_synced, total_ignored = self.run_daily_sync(target_date)
            messages.success(request, f"Sukses menarik log mesin tanggal {target_date}. Total baru: {total_synced}.")
            
        # =================================================================
        # AKSI 2: EVALUASI BERBASIS STRUKTUR & TANGGAL KHUSUS EVALUASI
        # =================================================================
        elif action == 'evaluate_attendance':
            # Ambil tanggal khusus dari input Langkah 2
            tanggal_eval_raw = request.POST.get('tanggal_evaluasi')
            
            if not tanggal_eval_raw:
                messages.error(request, "Tanggal target evaluasi Langkah 2 wajib diisi.")
                return redirect(reverse('disiplinsdm_urls:log-kehadiran'))
                
            target_date = datetime.strptime(tanggal_eval_raw, '%Y-%m-%d').date()
            
            # Tangkap parameter struktur organisasi
            instalasi_id = request.POST.get('instalasi_id')
            sub_bidang_id = request.POST.get('sub_bidang_id')
            bidang_id = request.POST.get('bidang_id')
            unor_id = request.POST.get('unor_id')
            
            instalasi_id = int(instalasi_id) if instalasi_id and instalasi_id.isdigit() else None
            sub_bidang_id = int(sub_bidang_id) if sub_bidang_id and sub_bidang_id.isdigit() else None
            bidang_id = int(bidang_id) if bidang_id and bidang_id.isdigit() else None
            unor_id = int(unor_id) if unor_id and unor_id.isdigit() else None
            
            # Eksekusi orchestrator
            success, batch_message = AttendanceOrchestrator.execute_by_structure(
                target_date=target_date,
                instalasi_id=instalasi_id,
                sub_bidang_id=sub_bidang_id,
                bidang_id=bidang_id,
                unor_id=unor_id
            )
            
            # print('hasil proses penilaian: ', success)
            
            if success:
                messages.success(request, batch_message)
            else:
                messages.warning(request, batch_message)
                
        else:
            messages.error(request, "Terjadi kesalahan pemrosesan form atau input tidak valid.")
            
        return redirect(reverse('disiplinsdm_urls:log-kehadiran'))

def sync_dashboard(request):
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    page = int(request.GET.get('page', 1))
    personname = request.GET.get('personname', '').strip()

    # 1. Ambil data dari Flask Bridge
    # Kita asumsikan fetch_unsynced_by_date mengembalikan (data_list, metadata)
    data_list, metadata = BridgeSyncService.fetch_unsynced_by_date(
        limit=1000, 
        date_from=date_from, 
        date_to=date_to,
        personname=personname,
    )

    # 2. Ambil Mapping Pegawai
    mappings = {m.mesin_id: m for m in MappingMesinAbsensi.objects.select_related('pegawai')}

    # 3. Kumpulkan Key Unik untuk Filter (ID + Datetime)
    # Gunakan set tuple untuk pencarian O(1) yang super cepat
    target_keys = set()
    for log in data_list:
        if 'id' in log and 'datetime' in log:
            target_keys.add((str(log['id']), log['datetime']))

    # 4. Ambil data yang SUDAH ADA di LogKehadiran SIMADU
    # Kita hanya filter berdasarkan keys yang ada di data_list untuk efisiensi
    existing_logs = {
        (str(m_id), dt.strftime('%Y-%m-%dT%H:%M:%S') if hasattr(dt, 'strftime') else str(dt))
        for m_id, dt in LogKehadiran.objects.filter(
            mapping__mesin_id__in=[str(lk[0]) for lk in target_keys]
        ).values_list('mapping__mesin_id', 'datetime')
    }

    filtered_data = []
    for log in data_list:
        mesin_id = str(log['id'])
        log_datetime = log['datetime']
        
        # LOGIKA UTAMA: Jika sudah ada di LogKehadiran, skip (tidak perlu tampil di dashboard sync)
        if (mesin_id, log_datetime) in existing_logs:
            continue

        mapping = mappings.get(mesin_id)
        
        filtered_data.append({
            'raw_log': log,
            'raw_log_json': json.dumps(log),
            'mapping': mapping,
            'is_ready': True if mapping else False,
            # Tambahan: info status dari Bridge (biasanya False karena kita panggil endpoint 'unsynced')
            'is_in_bridge_tracker': log.get('is_synced', False) 
        })

    # 5. Handle Parameter URL untuk Pagination agar tidak hilang
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    preserved_filters = query_params.urlencode()

    # 6. Django Pagination
    paginator = Paginator(filtered_data, 50)
    page_obj = paginator.get_page(page)

    context = {
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
        'personname': personname,
        'preserved_filters': preserved_filters,
        'total_unsynced': len(filtered_data),
        'title':'Data Belum Sinkron (Terbaru di Atas)',
    }

    return render(request, 'kehadirankegiatan/sync_table.html', context)


def sync_individual_api(request):
    """Endpoint untuk tombol 'Sinkronkan' per baris di UI"""
    if request.method == 'POST':
        log_data = json.loads(request.body)
        # Eksekusi hanya untuk 1 data ini
        count = BridgeSyncService.execute_sync([log_data])
        return JsonResponse({'success': count > 0})
    
@login_required
@require_POST
def process_single_sync(request):
    try:
        # 1. Ambil data JSON yang dikirim oleh fetch JavaScript
        data = json.loads(request.body)
        
        if not data:
            return JsonResponse({'success': False, 'message': 'Data kosong'}, status=400)

        # 2. Bungkus data ke dalam list karena execute_sync menerima format list/batch
        # Data yang diterima adalah satu raw_log (dict)
        logs_to_process = [data]
        
        # 3. Jalankan sinkronisasi menggunakan service yang sudah ada
        success_count, ignored_count = BridgeSyncService.execute_sync(logs_to_process)
        
        if success_count > 0:
            return JsonResponse({
                'success': True, 
                'message': f'Berhasil sinkronisasi: {data.get("personname")}'
            })
        else:
            return JsonResponse({
                'success': False, 
                'message': 'Gagal sinkronisasi. Mungkin mapping belum ada.'
            }, status=400)

    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Server Error: {str(e)}'
        }, status=500)


class SyncAnalisisPresensiPerInstalasi(View):
    def get(self, request, *args, **kwargs):
        instalasi_id = kwargs.get('instalasi_id')
        sub_bidang_id = kwargs.get('sub_bidang_id')
        bidang_id = kwargs.get('bidang_id')
        penempatan_id = None

        target_date_str = request.GET.get('tanggal')
        success_count = 0
        ignored_count = 0
        
        try:
            target_date = date.fromisoformat(target_date_str)
        except (ValueError, TypeError):
            messages.error(request, 'Format tanggal tidak valid. Gunakan YYYY-MM-DD.')
            return redirect(reverse('disiplinsdm_urls:kehadiran_list'))
        
        # 1. Tentukan ID mana yang aktif (is not None)
        if instalasi_id is not None:
            penempatan_id = instalasi_id
            filter_params = {'instalasi_id': instalasi_id}
        elif sub_bidang_id is not None:
            penempatan_id = sub_bidang_id
            filter_params = {'sub_bidang_id': sub_bidang_id}
        elif bidang_id is not None:
            penempatan_id = bidang_id
            filter_params = {'bidang_id': bidang_id}
        else:
            penempatan_id = None
            filter_params = {}

        # 2. Panggil fungsi cukup SATU KALI menggunakan **kwargs
        if filter_params:
            success_count, ignored_count = AttendanceOrchestrator.execute_by_structure(
                target_date, **filter_params
            )

        messages.success(request, f'Sinkronisasi selesai untuk Instalasi ID {penempatan_id} pada {target_date}. Berhasil: {success_count}, Dilewati: {ignored_count}')
        return redirect(reverse('disiplinsdm_urls:kehadiran_list') + f'?inst={penempatan_id}&tanggal={target_date_str}')
    
    
########################################### VIEW BARU UNTUK MODEL BARU ##############################
from .models import AbsensiHarian, LogAktivitasAbsen

class RekapPresensiBulananView(LoginRequiredMixin, generic.ListView):
    login_url = reverse_lazy('myaccount_urls:login_view')
    model = Users
    template_name = 'kehadirankegiatan/rekap_kehadiran_bulanan.html'
    context_object_name = 'daftar_pegawai'
    paginate_by = 50 

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Users.objects.none()

        # 1. Base Queryset dengan filter awal (Pegawai Aktif & Bukan Superuser)
        queryset = Users.objects.filter(is_active=True).exclude(is_superuser=True).select_related('profil_user')
        
        # 2. Ambil parameter bulan & tahun dari request URL (GET)
        try:
            bulan = int(self.request.GET.get('bulan', ''))
            tahun = int(self.request.GET.get('tahun', ''))
        except ValueError:
            bulan = date.today().month
            tahun = date.today().year
        
        hari_ini = date.today()
        
        # Tentukan Batas Awal dan Akhir Bulan untuk Filter Query
        awal_bulan = date(tahun, bulan, 1)
        akhir_bulan = (awal_bulan + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        # 3. Tentukan batas hari yang akan dinilai di dalam bulan tersebut
        if tahun == hari_ini.year and bulan == hari_ini.month:
            batas_hari_penilaian = hari_ini.day
        else:
            batas_hari_penilaian = calendar.monthrange(tahun, bulan)[1]
            if date(tahun, bulan, 1) > hari_ini:
                batas_hari_penilaian = 0

        # =========================================================================
        # OPTIMASI TAMBAHAN: AMBIL HARI LIBUR NASIONAL BULAN TERPILIH
        # =========================================================================
        libur_nasional_set = set(
            HariLibur.objects.filter(
                tanggal__range=(awal_bulan, akhir_bulan)
            ).values_list('tanggal', flat=True)
        )
                
        # Subquery untuk memeriksa keberadaan plotting jadwal kerja pegawai
        jadwal_exists_subquery = JadwalDinasSDM.objects.filter(
            pegawai__pegawai_id=OuterRef('pk'),
            tanggal__range=(awal_bulan, akhir_bulan),
        )

        # Prefetch data transaksional bulanan
        queryset = queryset.annotate(
            has_jadwal_bulan_ini=Exists(jadwal_exists_subquery)
        ).prefetch_related(
            Prefetch(
                'riwayat_penempatan', 
                queryset=RiwayatPenempatan.objects.filter(status=True).select_related(
                    'penempatan_level1', 'penempatan_level2', 'penempatan_level3', 'penempatan_level4'
                ), 
                to_attr='sk_penempatan_aktif'
            ),
            Prefetch(
                'absensiharian_set', 
                # Ikut sertakan prefetch tabel ApprovedJadwalDinasSDM melalui nama relation-nya (contoh: 'approved_jadwal')
                # jika Anda ingin menarik detail tipe dinas langsung per hari dari database.
                queryset=AbsensiHarian.objects.filter(tanggal__range=(awal_bulan, akhir_bulan)).prefetch_related('logs'),
                to_attr='transaksi_absen_bulan_ini'
            )
        ).distinct().order_by('first_name', 'last_name')

        queryset_list = list(queryset)
        
        # 4. JALANKAN KALKULASI DI LEVEL PYTHON (SANGAT RINGAN & CEPAT)
        for pegawai in queryset_list:
            absen_db_dict = {absen.tanggal: absen for absen in pegawai.transaksi_absen_bulan_ini}

            # Inisialisasi counter indikator kedisiplinan
            total_hadir = 0
            total_apel = 0
            total_alpa = 0
            total_izin = 0
            total_terlambat = 0
            total_pulang_cepat = 0
            total_libur = 0

            # Evaluasi kehadiran mundur dari hari penilaian sampai tanggal 1
            for day in range(batas_hari_penilaian, 0, -1):
                loop_date = date(tahun, bulan, day)
                
                # Deteksi aturan kalender global secara realtime
                is_minggu = loop_date.weekday() == 6
                is_libur_nasional = loop_date in libur_nasional_set
                is_tanggal_merah = is_minggu or is_libur_nasional

                if loop_date in absen_db_dict:
                    absen_hari_ini = absen_db_dict[loop_date]
                    status = str(absen_hari_ini.status_final).strip().upper()
                    
                    if status in ['HADIR', 'DINAS']:
                        total_hadir += 1
                        
                        for log in absen_hari_ini.logs.all():
                            tipe_log = log.tipe
                            status_ketepatan = str(log.status_ketepatan).strip().upper() if log.status_ketepatan else ""

                            if tipe_log == 'APEL':
                                total_apel += 1

                            if tipe_log in ['DATANG', 'APEL'] and 'TERLAMBAT' in status_ketepatan:
                                total_terlambat += 1

                            if tipe_log == 'PULANG' and 'CEPAT' in status_ketepatan:
                                total_pulang_cepat += 1
                                
                    elif status == 'IZIN':
                        total_izin += 1
                    elif status == 'LIBUR':
                        total_libur += 1 
                    elif status == 'ALPA':
                        total_alpa += 1
                else:
                    # =========================================================================
                    # SOLUSI INTI: EVALUASI KONDISI RECORD ABSENSI KOSONG (TANPA TRANSAKSI)
                    # =========================================================================
                    # Jika hari tersebut adalah tanggal merah (Minggu/Nasional) dan pegawai tidak punya jadwal 
                    # ATAU memiliki jadwal selain bertipe 'Piket', maka dianggap sebagai HARI LIBUR
                    if is_tanggal_merah:
                        # Perlindungan: Staf piket yang harusnya masuk tapi tidak dibuatkan record oleh unit tetap Alpa.
                        # Di sini kita asumsikan jika tidak ada record transaksi, default-nya diarahkan sebagai Libur Kalender.
                        total_libur += 1
                    else:
                        # Jika hari kerja efektif biasa tapi datanya bolong, mutlak divonis Mangkir (Alpa)
                        total_alpa += 1

            # Tempelkan atribut summary ke objek pegawai agar bisa langsung dirender oleh template
            pegawai.summary_hadir = total_hadir
            pegawai.summary_apel = total_apel
            pegawai.summary_alpa = total_alpa
            pegawai.summary_izin = total_izin
            pegawai.summary_terlambat = total_terlambat
            pegawai.summary_pulang_cepat = total_pulang_cepat
            pegawai.summary_libur = total_libur

        # =========================================================================
        # 5. LOGIKA SORTING
        # =========================================================================
        def sorting_key(pegawai):
            if pegawai.sk_penempatan_aktif:
                sk = pegawai.sk_penempatan_aktif[0]
                nama_instalasi = sk.penempatan_level4.instalasi if sk.penempatan_level4 else ""
            else:
                nama_instalasi = "ZZZZ" 

            kehadiran = -pegawai.summary_hadir
            return (nama_instalasi, kehadiran)

        queryset_list.sort(key=sorting_key)
        return queryset_list

    def get_context_data(self, **kwargs):
        # Ambil konteks dasar dari ListView
        context = super().get_context_data(**kwargs)
        
        # Ambil kembali parameter bulan dan tahun dari request URL (atau set default hari ini)
        try:
            bulan = int(self.request.GET.get('bulan', ''))
            tahun = int(self.request.GET.get('tahun', ''))
        except ValueError:
            bulan = date.today().month
            tahun = date.today().year
        
        # Sediakan data pilihan bulan untuk filter select form di HTML
        nama_bulan = {
            1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
            7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
        }
        
        # Kirim parameter ke context template
        context['selected_bulan'] = bulan
        context['selected_tahun'] = tahun
        context['daftar_bulan'] = nama_bulan
        context['title_page'] = 'Rekap Presensi Bulanan'
        context['label_periode'] = f"{nama_bulan.get(bulan)} {tahun}"
        
        return context

    def post(self, request, *args, **kwargs):
        """
        Menangani Aksi Trigger Tombol Sinkronisasi & Penilaian Otomatis (Orchestrator)
        Menerapkan teknik Rolling Window untuk menjamin keamanan data Shift Malam.
        """
        action = request.POST.get('action')
        target_date_str = request.POST.get('target_date') # Format dari input HTML5: YYYY-MM-DD

        if action == 'trigger_orchestrator' and target_date_str:
            try:
                # 1. Parsing tanggal dasar yang dipilih oleh admin
                chosen_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
                
                # 2. Tentukan window penanganan (Tanggal Awal dan Tanggal Akhir)
                # Jika admin memproses 'hari ini', window-nya adalah [Kemarin, Hari Ini]
                if chosen_date == timezone.now().date():
                    date_awal = chosen_date - timedelta(days=1)
                    date_akhir = chosen_date
                else:
                    # Jika admin memproses tanggal masa lalu, window-nya adalah [Tanggal Terpilih, Esoknya]
                    # Ini memastikan log pulang shift malam di keesokan harinya ikut ditarik dan dievaluasi
                    date_awal = chosen_date
                    date_akhir = chosen_date + timedelta(days=1)

                # 3. Jalankan Eksekusi Beruntun (Rolling Window)
                # Tahap 1: Jalankan tanggal awal (untuk mengunci/mengawinkan sisa log shift malam)
                success_1, msg_1 = NewAttendanceOrchestrator.execute_by_structure(target_date=date_awal, is_final_stage = True)
            
                # Tahap 2: Jalankan tanggal akhir (untuk membuka/memperbarui draft status berjalan)
                success_2, msg_2 = NewAttendanceOrchestrator.execute_by_structure(target_date=date_akhir, is_final_stage = False)

                # 4. Berikan feedback yang informatif kepada Admin di UI
                if success_1 and success_2:
                    messages.success(
                        request, 
                        f"Sukses menjalankan sinkronisasi berantai! "
                        f"Fase 1 ({date_awal.strftime('%d-%m-%Y')}): {msg_1} | "
                        f"Fase 2 ({date_akhir.strftime('%d-%m-%Y')}): {msg_2}"
                    )
                elif not success_1:
                    messages.error(request, f"Gagal pada Sinkronisasi Fase 1 ({date_awal.strftime('%d-%m-%Y')}): {msg_1}")
                else:
                    messages.error(request, f"Gagal pada Sinkronisasi Fase 2 ({date_akhir.strftime('%d-%m-%Y')}): {msg_2}")
                    
            except Exception as e:
                messages.error(request, f"Terjadi kesalahan sistem saat pemrosesan batch: {str(e)}")
        else:
            messages.warning(request, "Aksi atau tanggal target tidak valid.")

        # Redirect kembali ke halaman rekap membawa parameter filter aktif agar halaman tidak blank
        current_month = timezone.now().month
        current_year = timezone.now().year
        return redirect(
            f"{request.path}?bulan={request.GET.get('bulan', current_month)}&tahun={request.GET.get('tahun', current_year)}"
        )


class DetailPresensiPegawaiView(LoginRequiredMixin, generic.DetailView):
    login_url = reverse_lazy('myaccount_urls:login_view')
    model = Users
    template_name = 'kehadirankegiatan/detail_kehadiran_pegawai.html'
    context_object_name = 'pegawai'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pegawai = self.get_object()

        # 1. Ambil parameter filter bulan & tahun
        try:
            bulan = int(self.request.GET.get('bulan', ''))
            tahun = int(self.request.GET.get('tahun', ''))
        except ValueError:
            bulan = date.today().month
            tahun = date.today().year
        
        hari_ini = date.today()
        
        # Tentukan Batas Awal dan Akhir Bulan untuk Filter Kalender Libur Nasional
        awal_bulan = date(tahun, bulan, 1)
        akhir_bulan = (awal_bulan + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        # =========================================================================
        # OPTIMASI DAFTAR TANGGAL MERAH NASIONAL (BULAN TERPILIH)
        # =========================================================================
        libur_nasional_dict = {
            hl.tanggal: hl.keterangan 
            for hl in HariLibur.objects.filter(tanggal__range=(awal_bulan, akhir_bulan))
        }

        # 2. Tarik rekap absensi harian beserta seluruh log aktivitasnya (Prefetched)
        absensi_bulan_ini = AbsensiHarian.objects.filter(
            pegawai=pegawai,
            tanggal__month=bulan,
            tanggal__year=tahun
        ).prefetch_related('logs').order_by('-tanggal')

        # Map ke dictionary berbasis tanggal agar pencarian data instan
        absen_dict = {absen.tanggal: absen for absen in absensi_bulan_ini}

        # 3. Rentang tanggal penilaian (Mundur dari hari ini/akhir bulan)
        if tahun == hari_ini.year and bulan == hari_ini.month:
            total_hari = hari_ini.day
        else:
            total_hari = calendar.monthrange(tahun, bulan)[1]
            if date(tahun, bulan, 1) > hari_ini:
                total_hari = 0

        riwayat_harian = []

        # 4. Konstruksi baris tanggal per hari
        for day in range(total_hari, 0, -1):
            loop_date = date(tahun, bulan, day)
            
            # Deteksi status kalender harian di level Python
            is_minggu = loop_date.weekday() == 6
            is_libur_nasional = loop_date in libur_nasional_dict
            
            # Kerangka default untuk mengantisipasi pegawai yang tidak punya data/jadwal
            data_hari = {
                'id': '',
                'tanggal': loop_date,
                'status_final': 'Belum Presensi',
                'status_css': 'secondary',
                'keterangan_tambahan': '',
                'list_detak_log': [], 
                'has_logs': False
            }

            if loop_date in absen_dict:
                absen_hari_ini = absen_dict[loop_date]
                status_raw = str(absen_hari_ini.status_final).strip().upper() if absen_hari_ini.status_final else ""
                data_hari['id'] = absen_hari_ini.pk
                
                if status_raw == 'HADIR':
                    data_hari['status_final'] = 'HADIR'
                    data_hari['status_css'] = 'success'
                elif status_raw == 'ALPA':
                    data_hari['status_final'] = 'MANGKIR'
                    data_hari['status_css'] = 'danger'
                    data_hari['keterangan_tambahan'] = 'Teridentifikasi TK karena tidak ada log presensi sampai batas evaluasi dilakukan'
                elif status_raw == 'IZIN':
                    data_hari['status_final'] = 'IZIN / CUTI'
                    data_hari['status_css'] = 'info'
                elif status_raw == 'DINAS':
                    data_hari['status_final'] = 'DINAS LUAR'
                    data_hari['status_css'] = 'primary'
                elif status_raw == 'LIBUR':
                    data_hari['status_final'] = 'HARI LIBUR'
                    data_hari['status_css'] = 'dark' 
                    data_hari['keterangan_tambahan'] = absen_hari_ini.keterangan or 'Libur sesuai dengan jadwal dinas terdaftar'
                else:
                    # =========================================================================
                    # INTERVENSI VIEW: JIKA STATUS KOSONG DI DATABASE TAPI KALENDER ADALAH LIBUR
                    # =========================================================================
                    if is_libur_nasional:
                        nama_libur = libur_nasional_dict.get(loop_date)
                        data_hari['status_final'] = 'HARI LIBUR'
                        data_hari['status_css'] = 'dark'
                        data_hari['keterangan_tambahan'] = f'Libur Otomatis: {nama_libur} (Terdapat log presensi)'
                    elif is_minggu:
                        data_hari['status_final'] = 'HARI LIBUR'
                        data_hari['status_css'] = 'dark'
                        data_hari['keterangan_tambahan'] = 'Libur Otomatis: Hari Ahad / Minggu (Terdapat log presensi)'
                    else:
                        # Benar-benar hari kerja aktif yang belum diputuskan
                        data_hari['status_final'] = absen_hari_ini.status_final or 'Belum Presensi'
                        data_hari['status_css'] = 'secondary'
                        data_hari['keterangan_tambahan'] = 'Sudah terevaluasi dan belum diputuskan status kehadirannya'

                # Ambil detak transaksi log (Child)
                logs_child = absen_hari_ini.logs.all()
                if logs_child: 
                    data_hari['has_logs'] = True
                    for log in logs_child:
                        ketepatan_raw = str(log.status_ketepatan).strip().upper() if log.status_ketepatan else ""
                        
                        if 'TEPAT WAKTU' in ketepatan_raw or 'LUAR JADWAL' in ketepatan_raw:
                            bg_box = 'success text-white'
                        elif 'TERLAMBAT' in ketepatan_raw:
                            bg_box = 'warning text-dark'
                        elif 'CEPAT PULANG' in ketepatan_raw or 'CEPAT' in ketepatan_raw:
                            bg_box = 'danger text-white'
                        else:
                            bg_box = 'light text-muted'

                        data_hari['list_detak_log'].append({
                            'label_tipe': log.get_tipe_display() if hasattr(log, 'get_tipe_display') else log.tipe,
                            'jam': log.waktu.time() if hasattr(log.waktu, 'time') else log.waktu,
                            'ketepatan': log.status_ketepatan,
                            'bg_box': bg_box,
                        })
            else:
                # =========================================================================
                # PENYELAMATAN DATA KOSONG TANPA RECORD ABSENSI SAMA SEKALI
                # =========================================================================
                if is_libur_nasional:
                    nama_libur = libur_nasional_dict.get(loop_date)
                    data_hari['status_final'] = 'HARI LIBUR'
                    data_hari['status_css'] = 'dark'
                    data_hari['keterangan_tambahan'] = f'Libur Otomatis: {nama_libur}'
                elif is_minggu:
                    data_hari['status_final'] = 'HARI LIBUR'
                    data_hari['status_css'] = 'dark'
                    data_hari['keterangan_tambahan'] = 'Libur Otomatis: Hari Ahad / Minggu'
                else:
                    data_hari['status_final'] = 'MANGKIR'
                    data_hari['status_css'] = 'danger'
                    data_hari['keterangan_tambahan'] = 'Teridentifikasi TK karena tidak ada data jadwal atau data log absensi bolong di hari kerja'

            riwayat_harian.append(data_hari)

        # 5. Siapkan parameter nama bulan untuk label header
        nama_bulan = {
            1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
            7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
        }

        context['riwayat_harian'] = riwayat_harian
        context['selected_bulan'] = bulan
        context['selected_tahun'] = tahun
        context['label_periode'] = f"{nama_bulan.get(bulan)} {tahun}"
        context['title_page'] = f"Jurnal Aktivitas Presensi - {pegawai.first_name} {pegawai.last_name}"

        return context
    
    

class UpdatePresensiPegawaiView(LoginRequiredMixin, UserPassesTestMixin,  generic.UpdateView):
    login_url = reverse_lazy('myaccount_urls:login_view')
    model = AbsensiHarian
    template_name = 'kehadirankegiatan/kehadiran_formset.html'
    fields = ('status_final', 'keterangan')
    
    def test_func(self):   
        return self.request.user.is_disiplin_admin
    
    def handle_no_permission(self):
        messages.error(
            self.request,
            "Anda tidak memiliki akses ke halaman ini."
        )
        return redirect(reverse('disiplinsdm_urls:rekap_kehadiran_bulanan'))
    
    def get_success_url(self):
        bulan = self.request.GET.get('bulan', '')
        tahun = self.request.GET.get('tahun', '')
        id_pegawai = self.object.pegawai_id
        jenissdmperinstalasi = JenisSDMPerinstalasi.objects.filter(pegawai_id=id_pegawai, bulan=bulan, tahun=tahun).first()
        if jenissdmperinstalasi is not None:
            pk = jenissdmperinstalasi.pk
        else:
            pk = id_pegawai
        url = reverse('disiplinsdm_urls:detail_presensi_pegawai', kwargs={'pk':pk})
        if bulan and tahun:
            return f"{url}?bulan={bulan}&tahun={tahun}"
        return url
    
    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Update Status Presensi Pegawai'
        context['url'] = self.get_success_url()
        
        # MASUKKAN FORMSET KE CONTEXT
        # Jika ada error validation (POST), kirim data POST-an tadi ke formset agar inputan tidak hilang
        if self.request.POST:
            context['log_formset'] = LogAktivitasFormSet(self.request.POST, instance=self.object)
        else:
            context['log_formset'] = LogAktivitasFormSet(instance=self.object)
            
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        log_formset = context['log_formset']
        
        # Validasi ganda: Cek form utama DAN formset anak wajib valid
        if form.is_valid() and log_formset.is_valid():
            self.object = form.save()          # Simpan data AbsensiHarian (Parent)
            log_formset.instance = self.object
            log_formset.save()                 # Simpan data rincian LogAktivitasAbsen (Child)
            
            messages.success(self.request, "Status presensi harian dan rincian log berhasil diperbarui.")
            return redirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))
    

class DownloadRekapPresensiExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # 1. Ambil Parameter Filter Periode
        bulan = int(request.GET.get('bulan', timezone.now().month))
        tahun = int(request.GET.get('tahun', timezone.now().year))
        
        nama_bulan = {
            1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
            7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
        }.get(bulan, '')

        # 2. Ambil Queryset dengan Filter Keamanan Struktur
        queryset = JenisSDMPerinstalasi.objects.filter(bulan=bulan, tahun=tahun).select_related(
            'pegawai', 'unor', 'bidang', 'sub_bidang', 'instalasi'
        )
        
        user = request.user
        profil = getattr(user, 'profil_admin', None)
        
        if user.is_disiplin_admin:
            queryset = queryset
        elif profil and profil.is_pejabat:
            if profil.instalasi.exists():
                queryset = queryset.filter(instalasi__in=profil.instalasi.all())
            elif profil.sub_bidang.exists():
                queryset = queryset.filter(sub_bidang__in=profil.sub_bidang.all())
            elif profil.bidang.exists():
                queryset = queryset.filter(sub_bidang__bidang__in=profil.bidang.all())
            elif profil.unor.exists():
                queryset = queryset.filter(unor__in=profil.unor.all())
        else:
            riwayat = user.riwayat_penempatan.filter(status=True).values('penempatan_level4_id').first()
            if riwayat and riwayat['penempatan_level4_id']:
                queryset = queryset.filter(instalasi_id=riwayat['penempatan_level4_id'])
            else:
                queryset = queryset.none()

        # PERUBAHAN DI SINI: Tambahkan anotasi total_libur
        queryset = queryset.annotate(
            total_hadir=Count('pegawai__absensiharian__id', distinct=True, filter=Q(
                pegawai__absensiharian__tanggal__month=bulan, pegawai__absensiharian__tanggal__year=tahun, pegawai__absensiharian__status_final='HADIR'
            )),
            total_tk=Count('pegawai__absensiharian__id', distinct=True, filter=Q(
                pegawai__absensiharian__tanggal__month=bulan, pegawai__absensiharian__tanggal__year=tahun, pegawai__absensiharian__status_final='ALPA'
            )),
            total_izin=Count('pegawai__absensiharian__id', distinct=True, filter=Q(
                pegawai__absensiharian__tanggal__month=bulan, pegawai__absensiharian__tanggal__year=tahun, pegawai__absensiharian__status_final='IZIN'
            )),
            total_libur=Count('pegawai__absensiharian__id', distinct=True, filter=Q(
                pegawai__absensiharian__tanggal__month=bulan, pegawai__absensiharian__tanggal__year=tahun, pegawai__absensiharian__status_final='LIBUR'
            )),
            total_terlambat=Count('pegawai__absensiharian__logs__id', distinct=True, filter=Q(
                pegawai__absensiharian__tanggal__month=bulan, pegawai__absensiharian__tanggal__year=tahun, pegawai__absensiharian__logs__tipe='DATANG', pegawai__absensiharian__logs__status_ketepatan__in=['Terlambat Ringan', 'Terlambat Sedang', 'Terlambat Berat']
            )),
            total_cepat_pulang=Count('pegawai__absensiharian__logs__id', distinct=True, filter=Q(
                pegawai__absensiharian__tanggal__month=bulan, pegawai__absensiharian__tanggal__year=tahun, pegawai__absensiharian__logs__tipe='PULANG', pegawai__absensiharian__logs__status_ketepatan='Cepat Pulang'
            )),
        ).order_by('pegawai__first_name', 'pegawai__last_name')

        # 3. Proses Pembuatan File Excel (openpyxl)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rekap Presensi"
        ws.views.sheetView[0].showGridLines = True

        # Styles komponen tabel
        font_judul = Font(name='Arial', size=14, bold=True)
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        font_data = Font(name='Arial', size=10)
        
        fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center')
        
        border_tipis = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        ws['A1'] = "REKAPITULASI PRESENSI BULANAN PEGAWAI"
        ws['A1'].font = font_judul
        ws['A2'] = f"Periode: {nama_bulan} {tahun}"
        ws['A2'].font = Font(name='Arial', size=11, italic=True)
        
        # PERUBAHAN DI SINI: Tambah 'Libur' ke daftar header (index baru kolom ke-5)
        headers = ['No', 'Nama Pegawai', 'Instalasi/Unit', 'Hadir', 'Izin', 'Libur', 'Alpa (TK)', 'Terlambat', 'Cepat Pulang']
        ws.append([]) 
        ws.append(headers) 
        
        for cell in ws[4]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        ws.row_dimensions[4].height = 25

        # 4. Memasukkan Data Kinerja Pegawai ke Tabel
        for idx, row_data in enumerate(queryset, start=1):
            nama_lengkap = f"{row_data.pegawai.first_name} {row_data.pegawai.last_name}".strip()
            nama_instalasi = row_data.instalasi.instalasi if row_data.instalasi else "-"
            
            # PERUBAHAN DI SINI: Menyisipkan row_data.total_libur ke posisi kolom yang pas
            row_values = [
                idx,
                nama_lengkap,
                nama_instalasi,
                row_data.total_hadir,
                row_data.total_izin,
                row_data.total_libur,
                row_data.total_tk,
                row_data.total_terlambat,
                row_data.total_cepat_pulang
            ]
            ws.append(row_values)
            
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 20
            
            for col_idx, cell in enumerate(ws[current_row], start=1):
                cell.font = font_data
                cell.border = border_tipis
                # PERUBAHAN DI SINI: Menambahkan index kolom 9 karena jumlah total kolom bertambah menjadi 9
                if col_idx in [1, 4, 5, 6, 7, 8, 9]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        # 5. Auto-Fit Lebar Kolom
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < 4: 
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # 6. Setup Response HTTP Browser
        filename = f"Rekap_Presensi_{bulan}_{tahun}.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
