import calendar
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from xml.sax.saxutils import escape

from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models import OuterRef, Q, Subquery
from django.http import Http404, HttpResponse
from django.shortcuts import render
from django.template.defaultfilters import slugify
from django.utils import timezone
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from strukturorg.models import PejabatStruktur, UnitInstalasi
from dokumen.models import RiwayatPengangkatan, STATUSPEGAWAI
from myaccount.models import Users

from .access import (
    filter_installations_for_discipline_admin,
    filter_users_for_discipline_admin,
    is_discipline_admin,
)
from .models import (
    AbsensiHarian,
    ApprovedJadwalDinasSDM,
    HariLibur,
    JenisSDMPerinstalasi,
)


NAMA_BULAN = {
    1: 'Januari',
    2: 'Februari',
    3: 'Maret',
    4: 'April',
    5: 'Mei',
    6: 'Juni',
    7: 'Juli',
    8: 'Agustus',
    9: 'September',
    10: 'Oktober',
    11: 'November',
    12: 'Desember',
}

STATUS_CODE = {
    'HADIR': 'H',
    'ALPA': 'A',
    'IZIN': 'I',
    'DINAS': 'D',
    'LIBUR': 'L',
}

STATUS_COLORS = {
    'H': colors.HexColor('#D9EAD3'),
    'A': colors.HexColor('#F4CCCC'),
    'I': colors.HexColor('#CFE2F3'),
    'D': colors.HexColor('#D9D2E9'),
    'L': colors.HexColor('#E7E6E6'),
    '-': colors.white,
}


SUMMARY_HEADERS = ('Hadir', 'TK', 'Izin', 'Dinas', 'Libur', '-', 'Terlambat', 'Cepat Pulang', 'Akumulasi')
AUTOMATED_DOCUMENT_NOTICE = (
    'Dokumen ini digenerate secara otomatis dari aplikasi SIMADU BERDANSA -- '
    'Sistem informasi terpadu sumber daya manusia RS Mandalika Provinsi NTB'
)


def format_minutes(total_minutes):
    """Format akumulasi menit sebagai jam:menit tanpa terpotong setelah 24 jam."""
    total_minutes = max(0, int(round(total_minutes or 0)))
    hours, minutes = divmod(total_minutes, 60)
    return f'{hours:02d}:{minutes:02d}'


def format_indonesian_date(value):
    return f'{value.day} {NAMA_BULAN[value.month]} {value.year}'


def as_local_naive(value, current_timezone=None):
    """Samakan datetime DB aware/naive untuk kalkulasi jadwal lokal."""
    if timezone.is_aware(value):
        value = timezone.localtime(
            value,
            current_timezone or timezone.get_current_timezone(),
        )
        return value.replace(tzinfo=None)
    return value


def current_local_date():
    current = timezone.now()
    if timezone.is_aware(current):
        return timezone.localtime(current).date()
    return current.date()


class DownloadPresensiBulananPDFView(LoginRequiredMixin, View):
    """Filter dan unduh matriks presensi bulanan untuk satu instalasi."""

    login_url = 'myaccount_urls:login_view'
    template_name = 'kehadirankegiatan/rekap_presensi_pdf_filter.html'

    def get_allowed_installations(self):
        user = self.request.user
        queryset = UnitInstalasi.objects.select_related(
            'sub_bidang__bidang__unor'
        ).order_by('instalasi')

        if is_discipline_admin(user):
            return filter_installations_for_discipline_admin(queryset, user)

        installation_ids = self.request.user.riwayat_penempatan.filter(
            status=True,
            penempatan_level4__isnull=False,
        ).values_list('penempatan_level4_id', flat=True)
        return queryset.filter(pk__in=installation_ids)

    def get_period(self):
        now = timezone.now().date()
        try:
            bulan = int(self.request.GET.get('bulan', now.month))
            tahun = int(self.request.GET.get('tahun', now.year))
        except (TypeError, ValueError):
            return now.month, now.year, 'Bulan atau tahun tidak valid.'

        if bulan not in NAMA_BULAN or not 2023 <= tahun <= 2100:
            return now.month, now.year, 'Periode harus berada pada bulan 1–12 dan tahun 2023–2100.'
        return bulan, tahun, None

    def get_context(self, bulan, tahun, error=None):
        return {
            'title_page': 'Download Presensi Bulanan',
            'daftar_instalasi': self.get_allowed_installations(),
            'daftar_bulan': NAMA_BULAN,
            'selected_instalasi': self.request.GET.get('instalasi', ''),
            'selected_cakupan': self.request.GET.get('cakupan', 'instalasi'),
            'selected_status_pegawai': self.request.GET.get('status_pegawai', ''),
            'daftar_status_pegawai': STATUSPEGAWAI,
            'can_filter_status': is_discipline_admin(self.request.user),
            'selected_bulan': bulan,
            'selected_tahun': tahun,
            'error': error,
        }

    def get(self, request, *args, **kwargs):
        bulan, tahun, error = self.get_period()
        download_format = request.GET.get('download')
        is_download = download_format in {'pdf', 'xlsx'}

        if not is_download:
            return render(request, self.template_name, self.get_context(bulan, tahun, error))

        if error:
            return render(
                request,
                self.template_name,
                self.get_context(bulan, tahun, error),
                status=400,
            )

        cakupan = request.GET.get('cakupan', 'instalasi')
        if cakupan == 'status':
            if not is_discipline_admin(request.user):
                raise PermissionDenied('Filter status pegawai lintas instalasi hanya untuk admin disiplin.')

            status_pegawai = request.GET.get('status_pegawai', '')
            valid_statuses = {value for value, _label in STATUSPEGAWAI}
            if status_pegawai not in valid_statuses:
                error = 'Pilih status pegawai yang akan dibuatkan laporan.'
                return render(
                    request,
                    self.template_name,
                    self.get_context(bulan, tahun, error),
                    status=400,
                )
            scope = {
                'mode': 'status',
                'status_pegawai': status_pegawai,
                'label': f'Status Pegawai: {status_pegawai} (Semua Instalasi)',
                'slug': f'status-{status_pegawai}',
            }
        else:
            try:
                instalasi_id = int(request.GET.get('instalasi', ''))
            except (TypeError, ValueError):
                error = 'Pilih instalasi yang akan dibuatkan laporan.'
                return render(
                    request,
                    self.template_name,
                    self.get_context(bulan, tahun, error),
                    status=400,
                )

            instalasi = self.get_allowed_installations().filter(pk=instalasi_id).first()
            if instalasi is None:
                raise Http404('Instalasi tidak ditemukan atau berada di luar hak akses Anda.')
            scope = {
                'mode': 'instalasi',
                'instalasi': instalasi,
                'label': f'Instalasi: {instalasi.instalasi}',
                'slug': instalasi.instalasi,
            }

        if download_format == 'xlsx':
            return self.build_excel(scope, bulan, tahun)
        return self.build_pdf(scope, bulan, tahun)

    def get_employees(self, scope, bulan, tahun):
        if scope['mode'] == 'status':
            latest_status = RiwayatPengangkatan.objects.filter(
                pegawai_id=OuterRef('pk'),
            ).order_by(
                '-tgl_srt_putusan',
                '-pk',
            ).values('status_pegawai')[:1]
            employees = Users.objects.filter(
                is_active=True,
            ).exclude(
                is_superuser=True,
            ).annotate(
                latest_employment_status=Subquery(latest_status),
            ).filter(
                latest_employment_status=scope['status_pegawai'],
            ).select_related(
                'profil_user',
            ).order_by(
                'first_name',
                'last_name',
                'pk',
            )
            return list(filter_users_for_discipline_admin(
                employees,
                self.request.user,
            ))

        instalasi = scope['instalasi']
        metadata = JenisSDMPerinstalasi.objects.filter(
            instalasi=instalasi,
            bulan=bulan,
            tahun=tahun,
            pegawai__is_active=True,
        ).select_related(
            'pegawai',
            'pegawai__profil_user',
        ).order_by(
            'pegawai__first_name',
            'pegawai__last_name',
            'pegawai_id',
        )

        employees = []
        seen_user_ids = set()
        for item in metadata:
            if item.pegawai_id not in seen_user_ids:
                seen_user_ids.add(item.pegawai_id)
                employees.append(item.pegawai)
        return employees

    def get_report_data(self, scope, bulan, tahun):
        jumlah_hari = calendar.monthrange(tahun, bulan)[1]
        awal_bulan = date(tahun, bulan, 1)
        akhir_bulan = date(tahun, bulan, jumlah_hari)
        employees = self.get_employees(scope, bulan, tahun)
        employee_ids = [employee.pk for employee in employees]

        attendance_rows = list(AbsensiHarian.objects.filter(
                pegawai_id__in=employee_ids,
                tanggal__range=(awal_bulan, akhir_bulan),
            ).prefetch_related('logs'))
        attendance_map = {
            (attendance.pegawai_id, attendance.tanggal):
                (attendance.status_final or '').strip().upper()
            for attendance in attendance_rows
        }

        schedule_map = {}
        schedules = ApprovedJadwalDinasSDM.objects.filter(
            pegawai__pegawai_id__in=employee_ids,
            tanggal__range=(awal_bulan, akhir_bulan),
            is_approved=True,
        ).select_related('pegawai', 'kategori_jadwal').order_by('pk')
        for schedule in schedules:
            schedule_map[(schedule.pegawai.pegawai_id, schedule.tanggal)] = schedule

        summaries = {
            employee_id: {
                'HADIR': 0, 'ALPA': 0, 'IZIN': 0, 'DINAS': 0,
                'LIBUR': 0, 'BELUM': 0,
                'late_minutes': 0, 'early_minutes': 0,
            }
            for employee_id in employee_ids
        }
        for employee_id in employee_ids:
            for day in range(1, jumlah_hari + 1):
                status = attendance_map.get(
                    (employee_id, date(tahun, bulan, day)), ''
                )
                summaries[employee_id][status if status in STATUS_CODE else 'BELUM'] += 1

        current_timezone = timezone.get_current_timezone()
        for attendance in attendance_rows:
            schedule = schedule_map.get((attendance.pegawai_id, attendance.tanggal))
            category = getattr(schedule, 'kategori_jadwal', None)
            if category is None:
                continue

            logs = list(attendance.logs.all())
            arrival_logs = [log for log in logs if log.tipe == 'DATANG']
            if not arrival_logs:
                arrival_logs = [log for log in logs if log.tipe == 'APEL']
            departure_logs = [log for log in logs if log.tipe == 'PULANG']

            if arrival_logs and category.waktu_datang:
                actual_arrival = as_local_naive(
                    min(log.waktu for log in arrival_logs),
                    current_timezone,
                )
                scheduled_arrival = datetime.combine(
                    attendance.tanggal,
                    category.waktu_datang,
                )
                summaries[attendance.pegawai_id]['late_minutes'] += max(
                    0,
                    (actual_arrival - scheduled_arrival).total_seconds() / 60,
                )

            if departure_logs and category.waktu_pulang:
                actual_departure = as_local_naive(
                    max(log.waktu for log in departure_logs),
                    current_timezone,
                )
                departure_date = attendance.tanggal
                if (
                    category.waktu_datang
                    and category.waktu_pulang <= category.waktu_datang
                ):
                    departure_date += timedelta(days=1)
                scheduled_departure = datetime.combine(
                    departure_date,
                    category.waktu_pulang,
                )
                summaries[attendance.pegawai_id]['early_minutes'] += max(
                    0,
                    (scheduled_departure - actual_departure).total_seconds() / 60,
                )
        holidays = set(HariLibur.objects.filter(
            tanggal__range=(awal_bulan, akhir_bulan)
        ).values_list('tanggal', flat=True))
        installations_by_employee = defaultdict(list)
        placement_rows = JenisSDMPerinstalasi.objects.filter(
            pegawai_id__in=employee_ids,
            bulan=bulan,
            tahun=tahun,
            instalasi__isnull=False,
        ).select_related('instalasi').order_by('pegawai_id', 'instalasi__instalasi')
        for placement in placement_rows:
            installation_name = placement.instalasi.instalasi
            if installation_name not in installations_by_employee[placement.pegawai_id]:
                installations_by_employee[placement.pegawai_id].append(installation_name)

        installation_labels = {
            employee_id: ', '.join(names)
            for employee_id, names in installations_by_employee.items()
        }
        return (
            jumlah_hari,
            employees,
            attendance_map,
            holidays,
            installation_labels,
            summaries,
        )

    def get_signature_officer(self, scope):
        """Ambil pejabat aktif Kepala Bagian Tata Usaha beserta gelarnya."""
        candidates = PejabatStruktur.objects.filter(
            is_active=True,
            pejabat__is_active=True,
        ).filter(
            Q(nama_jabatan__icontains='tata usaha')
            | Q(bidang__bidang__icontains='tata usaha')
        ).select_related(
            'pejabat',
            'pejabat__profil_user',
            'bidang',
            'bidang__unor',
        )

        if scope['mode'] == 'instalasi':
            unor_id = scope['instalasi'].sub_bidang.bidang.unor_id
            scoped_officer = candidates.filter(bidang__unor_id=unor_id).first()
            if scoped_officer:
                return scoped_officer
        return candidates.order_by('-tanggal_mulai', '-pk').first()

    def build_pdf(self, scope, bulan, tahun):
        jumlah_hari, employees, attendance_map, holidays, installation_labels, summaries = self.get_report_data(
            scope,
            bulan,
            tahun,
        )

        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=8 * mm,
            rightMargin=8 * mm,
            topMargin=10 * mm,
            bottomMargin=12 * mm,
            title=f'Rekap Presensi {scope["label"]} {NAMA_BULAN[bulan]} {tahun}',
            author='SIMADU',
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'AttendancePDFTitle',
            parent=styles['Title'],
            fontName='Helvetica-Bold',
            fontSize=12,
            leading=14,
            alignment=TA_CENTER,
            spaceAfter=2 * mm,
        )
        subtitle_style = ParagraphStyle(
            'AttendancePDFSubtitle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
        )
        name_style = ParagraphStyle(
            'AttendancePDFName',
            parent=styles['Normal'],
            fontSize=6,
            leading=7,
            alignment=TA_LEFT,
        )
        signature_style = ParagraphStyle(
            'AttendancePDFSignature',
            parent=styles['Normal'],
            fontSize=8,
            leading=11,
            alignment=TA_CENTER,
        )
        header_style = ParagraphStyle(
            'AttendancePDFHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=5,
            leading=6,
            alignment=TA_CENTER,
        )

        story = [
            Paragraph('REKAPITULASI PRESENSI BULANAN PEGAWAI', title_style),
            Paragraph(
                f'Cakupan: <b>{escape(scope["label"])}</b><br/>'
                f'Periode: <b>{NAMA_BULAN[bulan]} {tahun}</b>',
                subtitle_style,
            ),
            Spacer(1, 4 * mm),
        ]

        weekday_labels = ('Sen', 'Sel', 'Rab', 'Kam', 'Jum', 'Sab', 'Min')
        header = [
            Paragraph('No', header_style),
            Paragraph('Nama Pegawai / NIP', header_style),
            Paragraph('Instalasi', header_style),
        ]
        for day in range(1, jumlah_hari + 1):
            current_date = date(tahun, bulan, day)
            header.append(Paragraph(
                f'{day}<br/><font size="4">{weekday_labels[current_date.weekday()]}</font>',
                header_style,
            ))
        header.extend(Paragraph(label, header_style) for label in SUMMARY_HEADERS)

        table_data = [header]
        for number, employee in enumerate(employees, start=1):
            nip = getattr(getattr(employee, 'profil_user', None), 'nip', None) or '-'
            row = [
                str(number),
                Paragraph(
                    f'<b>{escape(employee.full_name or str(employee))}</b>'
                    f'<br/><font size="5">NIP. {escape(str(nip))}</font>',
                    name_style,
                ),
                Paragraph(
                    escape(installation_labels.get(employee.pk, '-')),
                    name_style,
                ),
            ]
            for day in range(1, jumlah_hari + 1):
                status = attendance_map.get((employee.pk, date(tahun, bulan, day)), '')
                row.append(STATUS_CODE.get(status, '-'))
            summary = summaries[employee.pk]
            row.extend([
                summary['HADIR'],
                summary['ALPA'],
                summary['IZIN'],
                summary['DINAS'],
                summary['LIBUR'],
                summary['BELUM'],
                format_minutes(summary['late_minutes']),
                format_minutes(summary['early_minutes']),
                format_minutes(summary['late_minutes'] + summary['early_minutes']),
            ])
            table_data.append(row)

        if not employees:
            table_data.append([
                '',
                Paragraph('Tidak ada data pegawai pada instalasi dan periode ini.', name_style),
                '',
            ] + ['-' for _ in range(jumlah_hari + len(SUMMARY_HEADERS))])

        available_width = landscape(A4)[0] - document.leftMargin - document.rightMargin
        number_width = 7 * mm
        name_width = 39 * mm
        installation_width = 31 * mm
        summary_widths = [6 * mm] * 6 + [11 * mm] * 3
        day_width = (
            available_width - number_width - name_width - installation_width
            - sum(summary_widths)
        ) / jumlah_hari
        table = LongTable(
            table_data,
            colWidths=(
                [number_width, name_width, installation_width]
                + [day_width] * jumlah_hari
                + summary_widths
            ),
            repeatRows=1,
            hAlign='CENTER',
        )

        table_commands = [
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#777777')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 1),
            ('RIGHTPADDING', (0, 0), (-1, -1), 1),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]

        for day in range(1, jumlah_hari + 1):
            current_date = date(tahun, bulan, day)
            column = day + 2
            if current_date.weekday() == 6 or current_date in holidays:
                table_commands.extend([
                    ('BACKGROUND', (column, 0), (column, 0), colors.HexColor('#A61C00')),
                    ('TEXTCOLOR', (column, 0), (column, 0), colors.white),
                ])

        for row_index, row in enumerate(table_data[1:], start=1):
            for column_index, status_code in enumerate(row[3:], start=3):
                table_commands.append((
                    'BACKGROUND',
                    (column_index, row_index),
                    (column_index, row_index),
                    STATUS_COLORS.get(status_code, colors.white),
                ))
                if status_code == 'A':
                    table_commands.append((
                        'TEXTCOLOR',
                        (column_index, row_index),
                        (column_index, row_index),
                        colors.HexColor('#990000'),
                    ))

        table.setStyle(TableStyle(table_commands))
        story.append(table)
        story.append(Spacer(1, 3 * mm))

        legend_data = [[
            Paragraph('<b>Legenda:</b>', name_style),
            'H = Hadir',
            'A = Alpa',
            'I = Izin/Cuti',
            'D = Dinas',
            'L = Libur',
            '- = Belum Dinilai',
        ]]
        legend = Table(legend_data, hAlign='LEFT')
        legend.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(legend)
        story.append(Spacer(1, 5 * mm))

        officer = self.get_signature_officer(scope)
        officer_name = (
            officer.pejabat.full_name_2
            if officer is not None
            else 'Pejabat belum ditetapkan'
        )
        officer_title = (
            officer.nama_jabatan
            if officer is not None and officer.nama_jabatan
            else 'Kepala Bagian Tata Usaha'
        )
        signature = Table(
            [[
                '',
                Paragraph(
                    f'Pujut, {format_indonesian_date(current_local_date())}'
                    f'<br/>{escape(officer_title)},'
                    f'<br/><br/><br/><br/><b><u>{escape(officer_name)}</u></b>',
                    signature_style,
                ),
            ]],
            colWidths=[available_width * 0.64, available_width * 0.36],
            hAlign='RIGHT',
        )
        signature.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(signature)

        def add_page_number(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 6)
            canvas.setFillColor(colors.HexColor('#555555'))
            canvas.drawCentredString(
                landscape(A4)[0] / 2,
                8.5 * mm,
                AUTOMATED_DOCUMENT_NOTICE,
            )
            canvas.drawString(
                document.leftMargin,
                5.5 * mm,
                f'Dicetak dari SIMADU pada {timezone.now().strftime("%d-%m-%Y %H:%M")}',
            )
            canvas.drawRightString(
                landscape(A4)[0] - document.rightMargin,
                5.5 * mm,
                f'Halaman {doc.page}',
            )
            canvas.restoreState()

        document.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
        pdf = buffer.getvalue()
        buffer.close()

        safe_scope = slugify(scope['slug']) or 'presensi'
        filename = f'presensi-{safe_scope}-{tahun}-{bulan:02d}.pdf'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def build_excel(self, scope, bulan, tahun):
        jumlah_hari, employees, attendance_map, holidays, installation_labels, summaries = self.get_report_data(
            scope,
            bulan,
            tahun,
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Presensi Bulanan'
        last_column = 4 + jumlah_hari + len(SUMMARY_HEADERS)
        last_column_letter = get_column_letter(last_column)

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        worksheet['A1'] = 'REKAPITULASI PRESENSI BULANAN PEGAWAI'
        worksheet['A1'].font = Font(name='Arial', size=14, bold=True, color='1F1F1F')
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet.row_dimensions[1].height = 24

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
        worksheet['A2'] = f'Cakupan: {scope["label"]}'
        worksheet['A2'].font = Font(name='Arial', size=10, bold=True)
        worksheet['A2'].alignment = Alignment(horizontal='center')

        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)
        worksheet['A3'] = f'Periode: {NAMA_BULAN[bulan]} {tahun}'
        worksheet['A3'].font = Font(name='Arial', size=10)
        worksheet['A3'].alignment = Alignment(horizontal='center')

        header_row = 5
        headers = (
            ['No', 'Nama Pegawai', 'NIP', 'Instalasi']
            + list(range(1, jumlah_hari + 1))
            + list(SUMMARY_HEADERS)
        )
        thin_border = Border(
            left=Side(style='thin', color='808080'),
            right=Side(style='thin', color='808080'),
            top=Side(style='thin', color='808080'),
            bottom=Side(style='thin', color='808080'),
        )
        header_fill = PatternFill('solid', fgColor='1F4E78')
        holiday_fill = PatternFill('solid', fgColor='A61C00')
        status_fills = {
            'H': PatternFill('solid', fgColor='D9EAD3'),
            'A': PatternFill('solid', fgColor='F4CCCC'),
            'I': PatternFill('solid', fgColor='CFE2F3'),
            'D': PatternFill('solid', fgColor='D9D2E9'),
            'L': PatternFill('solid', fgColor='E7E6E6'),
            '-': PatternFill('solid', fgColor='FFFFFF'),
        }

        for column, value in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=column, value=value)
            cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill

            if 5 <= column <= 4 + jumlah_hari:
                current_date = date(tahun, bulan, column - 4)
                if current_date.weekday() == 6 or current_date in holidays:
                    cell.fill = holiday_fill
                cell.value = f'{current_date.day}\n{("Sen", "Sel", "Rab", "Kam", "Jum", "Sab", "Min")[current_date.weekday()]}'
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.row_dimensions[header_row].height = 30
        first_data_row = header_row + 1

        for row_number, employee in enumerate(employees, start=first_data_row):
            sequence = row_number - header_row
            nip = getattr(getattr(employee, 'profil_user', None), 'nip', None) or '-'
            row_values = [
                sequence,
                employee.full_name or str(employee),
                str(nip),
                installation_labels.get(employee.pk, '-'),
            ]

            for day in range(1, jumlah_hari + 1):
                status = attendance_map.get((employee.pk, date(tahun, bulan, day)), '')
                row_values.append(STATUS_CODE.get(status, '-'))
            summary = summaries[employee.pk]
            row_values.extend([
                summary['HADIR'],
                summary['ALPA'],
                summary['IZIN'],
                summary['DINAS'],
                summary['LIBUR'],
                summary['BELUM'],
                format_minutes(summary['late_minutes']),
                format_minutes(summary['early_minutes']),
                format_minutes(summary['late_minutes'] + summary['early_minutes']),
            ])

            for column, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=row_number, column=column, value=value)
                cell.font = Font(name='Arial', size=9)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='left' if column in {2, 3, 4} else 'center',
                    vertical='center',
                )
                if 5 <= column <= 4 + jumlah_hari:
                    cell.fill = status_fills.get(value, status_fills['-'])
                    if value == 'A':
                        cell.font = Font(name='Arial', size=9, bold=True, color='990000')

        if not employees:
            worksheet.merge_cells(
                start_row=first_data_row,
                start_column=1,
                end_row=first_data_row,
                end_column=last_column,
            )
            empty_cell = worksheet.cell(
                row=first_data_row,
                column=1,
                value='Tidak ada data pegawai pada instalasi dan periode ini.',
            )
            empty_cell.alignment = Alignment(horizontal='center')
            empty_cell.font = Font(name='Arial', size=9, italic=True)

        last_data_row = header_row + max(len(employees), 1)
        if employees:
            validation = DataValidation(
                type='list',
                formula1='"H,A,I,D,L,-"',
                allow_blank=False,
            )
            validation.error = 'Gunakan kode H, A, I, D, L, atau -.'
            validation.errorTitle = 'Status tidak valid'
            validation.prompt = 'H=Hadir, A=Alpa, I=Izin/Cuti, D=Dinas, L=Libur, -=Belum Dinilai'
            validation.promptTitle = 'Kode status presensi'
            worksheet.add_data_validation(validation)
            last_day_letter = get_column_letter(4 + jumlah_hari)
            validation.add(f'E{first_data_row}:{last_day_letter}{last_data_row}')

        legend_row = last_data_row + 2
        worksheet.cell(row=legend_row, column=1, value='Legenda:').font = Font(
            name='Arial',
            size=9,
            bold=True,
        )
        legend_items = [
            'H = Hadir',
            'A = Alpa',
            'I = Izin/Cuti',
            'D = Dinas',
            'L = Libur',
            '- = Belum Dinilai',
        ]
        for offset, legend_text in enumerate(legend_items, start=2):
            cell = worksheet.cell(row=legend_row, column=offset, value=legend_text)
            cell.font = Font(name='Arial', size=9)

        officer = self.get_signature_officer(scope)
        officer_name = (
            officer.pejabat.full_name_2
            if officer is not None
            else 'Pejabat belum ditetapkan'
        )
        officer_title = (
            officer.nama_jabatan
            if officer is not None and officer.nama_jabatan
            else 'Kepala Bagian Tata Usaha'
        )
        signature_start_row = legend_row + 3
        signature_start_column = max(5, last_column - 4)
        worksheet.merge_cells(
            start_row=signature_start_row,
            start_column=signature_start_column,
            end_row=signature_start_row,
            end_column=last_column,
        )
        worksheet.cell(
            row=signature_start_row,
            column=signature_start_column,
            value=f'Pujut, {format_indonesian_date(current_local_date())}',
        )
        worksheet.merge_cells(
            start_row=signature_start_row + 1,
            start_column=signature_start_column,
            end_row=signature_start_row + 1,
            end_column=last_column,
        )
        worksheet.cell(
            row=signature_start_row + 1,
            column=signature_start_column,
            value=f'{officer_title},',
        )
        worksheet.merge_cells(
            start_row=signature_start_row + 6,
            start_column=signature_start_column,
            end_row=signature_start_row + 6,
            end_column=last_column,
        )
        signature_name_cell = worksheet.cell(
            row=signature_start_row + 6,
            column=signature_start_column,
            value=officer_name,
        )
        signature_name_cell.font = Font(name='Arial', size=10, bold=True, underline='single')
        for row in (signature_start_row, signature_start_row + 1, signature_start_row + 6):
            worksheet.cell(row=row, column=signature_start_column).alignment = Alignment(
                horizontal='center',
                vertical='center',
            )

        worksheet.column_dimensions['A'].width = 6
        worksheet.column_dimensions['B'].width = 32
        worksheet.column_dimensions['C'].width = 22
        worksheet.column_dimensions['D'].width = 26
        for column in range(5, 5 + jumlah_hari):
            worksheet.column_dimensions[get_column_letter(column)].width = 5
        for column in range(5 + jumlah_hari, last_column + 1):
            worksheet.column_dimensions[get_column_letter(column)].width = (
                13 if column > last_column - 3 else 9
            )

        worksheet.freeze_panes = 'E6'
        worksheet.auto_filter.ref = f'A{header_row}:{last_column_letter}{last_data_row}'
        worksheet.sheet_view.showGridLines = False
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.print_title_rows = f'{header_row}:{header_row}'
        worksheet.print_area = (
            f'A1:{last_column_letter}{signature_start_row + 6}'
        )
        worksheet.oddFooter.center.text = AUTOMATED_DOCUMENT_NOTICE
        worksheet.oddFooter.center.size = 8
        worksheet.oddFooter.center.font = 'Arial'

        safe_scope = slugify(scope['slug']) or 'presensi'
        filename = f'presensi-{safe_scope}-{tahun}-{bulan:02d}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
