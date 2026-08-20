from datetime import date, datetime, time

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from io import BytesIO

from myaccount.models import Users
from dokumen.models import RiwayatPengangkatan
from strukturorg.models import (
    Bidang,
    InstansiDaerah,
    SatuanKerjaInduk,
    SubBidang,
    UnitInstalasi,
    UnitOrganisasi,
    PejabatStruktur,
)

from .models import (
    AbsensiHarian,
    ApprovedJadwalDinasSDM,
    DetailKategoriJadwalDinas,
    JenisSDMPerinstalasi,
    KategoriJadwalDinas,
    LogAktivitasAbsen,
)
from .services import NewAttendanceOrchestrator


class DownloadPresensiBulananPDFViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = Users.objects.create_superuser(
            email='admin-presensi-pdf@example.com',
            first_name='Admin',
            last_name='Presensi',
            password='test-password',
        )
        cls.employee = Users.objects.create_user(
            email='pegawai-presensi-pdf@example.com',
            first_name='Pegawai',
            last_name='Contoh',
            password='test-password',
        )
        cls.outsider = Users.objects.create_user(
            email='luar-presensi-pdf@example.com',
            first_name='Pegawai',
            last_name='Luar',
            password='test-password',
        )

        instansi = InstansiDaerah.objects.create(instansi='Pemerintah Daerah')
        satker = SatuanKerjaInduk.objects.create(
            instansi_daerah=instansi,
            satuan_kerja='Rumah Sakit',
        )
        cls.unor = UnitOrganisasi.objects.create(
            satker_induk=satker,
            unor='RS Mandalika',
        )
        bidang = Bidang.objects.create(unor=cls.unor, bidang='Pelayanan')
        bagian_tata_usaha = Bidang.objects.create(
            unor=cls.unor,
            bidang='Bagian Tata Usaha',
        )
        sub_bidang = SubBidang.objects.create(bidang=bidang, sub_bidang='Keperawatan')
        cls.instalasi = UnitInstalasi.objects.create(
            sub_bidang=sub_bidang,
            instalasi='Instalasi Pengujian',
        )
        PejabatStruktur.objects.create(
            bidang=bagian_tata_usaha,
            pejabat=cls.admin,
            nama_jabatan='Kepala Bagian Tata Usaha',
            tanggal_mulai=date(2025, 1, 1),
        )

        employee_metadata = JenisSDMPerinstalasi.objects.create(
            pegawai=cls.employee,
            unor=cls.unor,
            bidang=bidang,
            sub_bidang=sub_bidang,
            instalasi=cls.instalasi,
            bulan=1,
            tahun=2026,
            status='disetujui',
        )
        attendance = AbsensiHarian.objects.create(
            pegawai=cls.employee,
            tanggal=date(2026, 1, 2),
            unor=cls.unor,
            bidang=bidang,
            sub_bidang=sub_bidang,
            instalasi=cls.instalasi,
            status_final='HADIR',
        )
        schedule_category = KategoriJadwalDinas.objects.create(
            kategori_dinas='REG',
        )
        schedule_detail = DetailKategoriJadwalDinas.objects.create(
            kategori_dinas=schedule_category,
            hari='Senin s/d kamis',
            kategori_jadwal='Pagi',
            waktu_datang=time(8, 0),
            waktu_pulang=time(17, 0),
        )
        ApprovedJadwalDinasSDM.objects.create(
            pegawai=employee_metadata,
            tanggal=date(2026, 1, 2),
            kategori_jadwal=schedule_detail,
            is_approved=True,
            approved_by=cls.admin,
        )
        LogAktivitasAbsen.objects.create(
            absensi_harian=attendance,
            tipe='DATANG',
            waktu=timezone.make_aware(datetime(2026, 1, 2, 8, 45)),
            status_ketepatan='Terlambat Berat',
        )
        LogAktivitasAbsen.objects.create(
            absensi_harian=attendance,
            tipe='PULANG',
            waktu=timezone.make_aware(datetime(2026, 1, 2, 16, 30)),
            status_ketepatan='Cepat Pulang',
        )
        RiwayatPengangkatan.objects.create(
            pegawai=cls.employee,
            status_pegawai='PNS',
            no_srt_putusan='SK-PNS-001',
            tgl_srt_putusan=date(2025, 1, 1),
        )
        RiwayatPengangkatan.objects.create(
            pegawai=cls.outsider,
            status_pegawai='Kontrak',
            no_srt_putusan='SK-KONTRAK-001',
            tgl_srt_putusan=date(2025, 1, 1),
        )

    def test_filter_page_lists_allowed_installation(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse('disiplinsdm_urls:download_presensi_pdf'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Instalasi Pengujian')
        self.assertContains(response, 'Download Presensi Bulanan')
        self.assertContains(response, 'Berdasarkan Status Pegawai')
        self.assertContains(response, 'Kontrak')

    def test_download_returns_pdf_for_selected_installation_and_period(self):
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse('disiplinsdm_urls:download_presensi_pdf'),
            {
                'download': 'pdf',
                'instalasi': self.instalasi.pk,
                'bulan': 1,
                'tahun': 2026,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('attachment;', response['Content-Disposition'])
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_installation_outside_user_scope_is_not_downloadable(self):
        self.client.force_login(self.outsider)
        response = self.client.get(
            reverse('disiplinsdm_urls:download_presensi_pdf'),
            {
                'download': 'pdf',
                'instalasi': self.instalasi.pk,
                'bulan': 1,
                'tahun': 2026,
            },
        )

        self.assertEqual(response.status_code, 404)

    def test_download_returns_editable_daily_excel(self):
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse('disiplinsdm_urls:download_presensi_pdf'),
            {
                'download': 'xlsx',
                'instalasi': self.instalasi.pk,
                'bulan': 1,
                'tahun': 2026,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook['Presensi Bulanan']
        self.assertEqual(worksheet['B6'].value, 'Pegawai Contoh')
        self.assertEqual(worksheet['F6'].value, 'H')
        self.assertEqual(worksheet['AJ5'].value, 'Hadir')
        self.assertEqual(worksheet['AJ6'].value, 1)
        self.assertEqual(worksheet['AP6'].value, '00:45')
        self.assertEqual(worksheet['AQ6'].value, '00:30')
        self.assertEqual(worksheet['AR6'].value, '01:15')
        self.assertIn('Pujut,', worksheet['AN11'].value)
        self.assertEqual(worksheet['AN12'].value, 'Kepala Bagian Tata Usaha,')
        self.assertEqual(worksheet['AN17'].value, cls.admin.full_name_2)
        self.assertIn(
            'Dokumen ini digenerate secara otomatis dari aplikasi SIMADU BERDANSA',
            worksheet.oddFooter.center.text,
        )
        self.assertEqual(worksheet.protection.sheet, False)

    def test_admin_can_download_all_installations_by_latest_employee_status(self):
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse('disiplinsdm_urls:download_presensi_pdf'),
            {
                'download': 'xlsx',
                'cakupan': 'status',
                'status_pegawai': 'PNS',
                'bulan': 1,
                'tahun': 2026,
            },
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook['Presensi Bulanan']
        employee_names = [
            worksheet.cell(row=row, column=2).value
            for row in range(6, worksheet.max_row + 1)
        ]
        self.assertIn('Pegawai Contoh', employee_names)
        self.assertNotIn('Pegawai Luar', employee_names)
        self.assertIn('status-pns', response['Content-Disposition'])

    def test_non_admin_cannot_download_cross_installation_status_report(self):
        self.client.force_login(self.outsider)
        response = self.client.get(
            reverse('disiplinsdm_urls:download_presensi_pdf'),
            {
                'download': 'pdf',
                'cakupan': 'status',
                'status_pegawai': 'Kontrak',
                'bulan': 1,
                'tahun': 2026,
            },
        )

        self.assertEqual(response.status_code, 403)

    def test_invalid_period_is_rejected(self):
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse('disiplinsdm_urls:download_presensi_pdf'),
            {
                'download': 'pdf',
                'instalasi': self.instalasi.pk,
                'bulan': 13,
                'tahun': 2026,
            },
        )

        self.assertEqual(response.status_code, 400)

    def test_past_normal_day_without_schedule_becomes_alpa_with_explanation(self):
        target_date = date(2026, 1, 3)
        success, _message = NewAttendanceOrchestrator.execute_by_structure(
            target_date=target_date,
            instalasi_id=self.instalasi.pk,
            is_final_stage=True,
        )

        self.assertTrue(success)
        attendance = AbsensiHarian.objects.get(
            pegawai=self.employee,
            tanggal=target_date,
        )
        self.assertEqual(attendance.status_final, 'ALPA')
        self.assertIn('jadwal dinas pegawai belum dibuat', attendance.keterangan)

    def test_reassessment_does_not_overwrite_real_tapping_when_schedule_missing(self):
        target_date = date(2026, 1, 5)
        attendance = AbsensiHarian.objects.create(
            pegawai=self.employee,
            tanggal=target_date,
            unor=self.unor,
            instalasi=self.instalasi,
            status_final='HADIR',
        )
        LogAktivitasAbsen.objects.create(
            absensi_harian=attendance,
            tipe='DATANG',
            waktu=datetime(2026, 1, 5, 7, 30),
            status_ketepatan='Tepat Waktu',
            devicename='Mesin Utama',
        )

        success, _message = NewAttendanceOrchestrator.execute_by_structure(
            target_date=target_date,
            instalasi_id=self.instalasi.pk,
            is_final_stage=True,
        )

        self.assertTrue(success)
        attendance.refresh_from_db()
        self.assertEqual(attendance.status_final, 'HADIR')
        self.assertIn('log presensi', attendance.keterangan)
