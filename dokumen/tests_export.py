from datetime import date, timedelta
from io import BytesIO

from django.contrib.auth.models import Group
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from jenissdm.models import JenisSDM
from myaccount.models import ProfilSDM, Users
from myaccount.roles import ADMIN_DOKUMEN

from .document_registry import DOCUMENT_TYPES
from .export_views import csv_safe
from .models import (
    DokumenSDM,
    RiwayatJabatan,
    RiwayatPendidikan,
    RiwayatProfesi,
    RiwayatSIPProfesi,
)


class DocumentExportCSVTests(TestCase):
    password = 'Password-Export-123!'

    @classmethod
    def setUpTestData(cls):
        cls.employee = cls.create_employee(
            'pegawai-export@example.com',
            'Pegawai Export',
            '199001',
        )
        cls.other_employee = cls.create_employee(
            'pegawai-export-lain@example.com',
            'Pegawai Lain',
            '199002',
        )
        cls.admin = cls.create_employee(
            'admin-export@example.com',
            'Admin Export',
            '199003',
        )
        group, _ = Group.objects.get_or_create(name=ADMIN_DOKUMEN)
        cls.admin.groups.add(group)

        document_type = DokumenSDM.objects.create(
            nama='Riwayat Pendidikan',
            url='pendidikan',
        )
        RiwayatPendidikan.objects.create(
            pegawai=cls.employee,
            dokumen=document_type,
            pendidikan='Data Pendidikan Milik Sendiri',
            nama_sek='Universitas Sendiri',
        )
        RiwayatPendidikan.objects.create(
            pegawai=cls.other_employee,
            dokumen=document_type,
            pendidikan='Data Pendidikan Milik Pegawai Lain',
            nama_sek='Universitas Lain',
        )

    @classmethod
    def create_employee(cls, email, first_name, nip):
        user = Users.objects.create_user(
            email=email,
            first_name=first_name,
            last_name='',
            password=cls.password,
        )
        ProfilSDM.objects.create(
            user=user,
            nip=nip,
            no_hp='08123456789',
            email_pribadi=email,
        )
        return user

    def get_export(self, user, params=None, document_type='pendidikan'):
        self.client.force_login(user)
        return self.client.get(
            reverse(
                'riwayat_urls:document_export_csv',
                kwargs={'document_type': document_type},
            ),
            params or {},
        )

    @staticmethod
    def get_content(response):
        return b''.join(response.streaming_content).decode('utf-8-sig')

    def test_regular_user_exports_only_own_documents(self):
        response = self.get_export(
            self.employee,
            {'nip': self.other_employee.profil_user.nip},
        )

        content = self.get_content(response)
        self.assertEqual(response.status_code, 200)
        self.assertIn('Data Pendidikan Milik Sendiri', content)
        self.assertNotIn('Data Pendidikan Milik Pegawai Lain', content)

    def test_document_admin_can_export_selected_employee(self):
        response = self.get_export(
            self.admin,
            {'nip': self.other_employee.profil_user.nip},
        )

        content = self.get_content(response)
        self.assertIn('Data Pendidikan Milik Pegawai Lain', content)
        self.assertNotIn('Data Pendidikan Milik Sendiri', content)

    def test_unknown_document_type_returns_404(self):
        response = self.get_export(self.employee, document_type='rahasia')
        self.assertEqual(response.status_code, 404)

    def test_all_registered_document_types_have_working_endpoint(self):
        for document_type in DOCUMENT_TYPES:
            with self.subTest(document_type=document_type):
                response = self.get_export(self.admin, document_type=document_type)
                self.assertEqual(response.status_code, 200)
                self.assertTrue(self.get_content(response).startswith('Nama Pegawai,NIP,'))

    def test_csv_safe_prevents_spreadsheet_formula(self):
        self.assertEqual(csv_safe('=HYPERLINK("https://example.com")'), "'=HYPERLINK(\"https://example.com\")")


class ProfessionSIPExportExcelTests(TestCase):
    password = 'Password-Excel-123!'

    @classmethod
    def setUpTestData(cls):
        cls.employee = cls.create_employee(
            'pegawai-str@example.com',
            'Pegawai STR',
            '198801',
        )
        cls.other_employee = cls.create_employee(
            'pegawai-sip-lain@example.com',
            'Pegawai SIP Lain',
            '198802',
        )
        cls.admin = cls.create_employee(
            'admin-profesi@example.com',
            'Admin Profesi',
            '198803',
        )
        group, _ = Group.objects.get_or_create(name=ADMIN_DOKUMEN)
        cls.admin.groups.add(group)

        old_jabatan = JenisSDM.objects.create(jenis_sdm='Jabatan Lama')
        current_jabatan = JenisSDM.objects.create(jenis_sdm='Jabatan Sekarang')
        future_jabatan = JenisSDM.objects.create(jenis_sdm='Jabatan Mendatang')
        RiwayatJabatan.objects.create(
            pegawai=cls.employee,
            nama_jabatan=old_jabatan,
            tmt_jabatan=date(2023, 1, 1),
        )
        RiwayatJabatan.objects.create(
            pegawai=cls.employee,
            nama_jabatan=current_jabatan,
            detail_nama_jabatan='Detail Jabatan',
            tmt_jabatan=date(2025, 1, 1),
        )
        RiwayatJabatan.objects.create(
            pegawai=cls.employee,
            nama_jabatan=future_jabatan,
            tmt_jabatan=date.today() + timedelta(days=365),
        )
        RiwayatJabatan.objects.create(
            pegawai=cls.other_employee,
            nama_jabatan=old_jabatan,
            tmt_jabatan=date(2024, 1, 1),
        )

        cls.profession = RiwayatProfesi.objects.create(
            pegawai=cls.employee,
            no_str='STR-PEGAWAI-001',
            tgl_str=date(2024, 2, 1),
        )
        RiwayatSIPProfesi.objects.create(
            riwayat_profesi=cls.profession,
            no_sip='SIP-PEGAWAI-001',
            tgl_sip=date(2024, 3, 1),
            berlaku_sd=date(2029, 3, 1),
        )
        RiwayatSIPProfesi.objects.create(
            riwayat_profesi=cls.profession,
            no_sip='SIP-PEGAWAI-002',
            tgl_sip=date(2025, 3, 1),
            berlaku_sd=date(2030, 3, 1),
        )
        other_profession = RiwayatProfesi.objects.create(
            pegawai=cls.other_employee,
            no_str='STR-LAIN-001',
            tgl_str=date(2023, 2, 1),
        )
        RiwayatSIPProfesi.objects.create(
            riwayat_profesi=other_profession,
            no_sip='SIP-LAIN-001',
            tgl_sip=date(2023, 3, 1),
            berlaku_sd=date(2028, 3, 1),
        )

    @classmethod
    def create_employee(cls, email, first_name, nip):
        user = Users.objects.create_user(
            email=email,
            first_name=first_name,
            last_name='',
            password=cls.password,
        )
        ProfilSDM.objects.create(
            user=user,
            nip=nip,
            no_hp='08123456789',
            email_pribadi=email,
        )
        return user

    def get_export_rows(self, user, params=None):
        self.client.force_login(user)
        response = self.client.get(
            reverse('riwayat_urls:profession_sip_export_excel'),
            params or {},
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        worksheet = workbook['Riwayat Profesi dan SIP']
        rows = list(worksheet.iter_rows(min_row=4, values_only=True))
        return response, worksheet, rows

    def test_pegawai_hanya_mengekspor_data_profesi_dan_sip_sendiri(self):
        response, worksheet, rows = self.get_export_rows(
            self.employee,
            {'nip': self.other_employee.profil_user.nip},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            tuple(cell.value for cell in worksheet[3]),
            (
                'Nama Pegawai',
                'Jabatan',
                'Nomor STR',
                'Tanggal STR',
                'Tanggal Terbit SIP',
                'Tanggal Berakhir SIP',
                'Nomor SIP',
            ),
        )
        self.assertEqual(len(rows), 2)
        self.assertTrue(all(row[0] == self.employee.full_name for row in rows))
        self.assertTrue(
            all(row[1] == 'Jabatan Sekarang (Detail Jabatan)' for row in rows)
        )
        self.assertEqual({row[6] for row in rows}, {
            'SIP-PEGAWAI-001',
            'SIP-PEGAWAI-002',
        })

    def test_admin_dokumen_dapat_mengekspor_pegawai_yang_dipilih(self):
        response, _worksheet, rows = self.get_export_rows(
            self.admin,
            {'nip': self.other_employee.profil_user.nip},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], self.other_employee.full_name)
        self.assertEqual(rows[0][2], 'STR-LAIN-001')
        self.assertEqual(rows[0][6], 'SIP-LAIN-001')
