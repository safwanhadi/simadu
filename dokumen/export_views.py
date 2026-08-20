import csv
from datetime import date, datetime

from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import FieldDoesNotExist
from django.db import models
from django.http import Http404, HttpResponse, StreamingHttpResponse
from django.utils import timezone
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from layanan.access.sip import filter_profession_history_queryset

from .access import get_selected_nip
from .document_registry import DOCUMENT_TYPES
from .models import RiwayatJabatan, RiwayatProfesi, RiwayatSIPProfesi

EXCLUDED_FIELDS = {'id', 'pegawai', 'dokumen'}


class Echo:
    """File-like object agar csv.writer dapat menghasilkan baris streaming."""

    @staticmethod
    def write(value):
        return value


def csv_safe(value):
    """Ubah nilai menjadi teks dan cegah formula injection saat dibuka di Excel."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Ya' if value else 'Tidak'
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        value = value.strftime('%Y-%m-%d %H:%M:%S')
    elif isinstance(value, date):
        value = value.isoformat()

    text = str(value)
    if text.startswith(('=', '+', '-', '@', '\t', '\r')):
        return f"'{text}"
    return text


class DocumentExportCSVView(LoginRequiredMixin, View):
    """Ekspor seluruh queryset yang diizinkan, tanpa pagination halaman HTML."""

    http_method_names = ['get']

    def get_configuration(self):
        try:
            return DOCUMENT_TYPES[self.kwargs['document_type']]
        except KeyError as exc:
            raise Http404('Jenis dokumen tidak mendukung ekspor.') from exc

    @staticmethod
    def get_employee_field(model):
        try:
            return model._meta.get_field('pegawai')
        except FieldDoesNotExist as exc:
            raise Http404('Dokumen tidak memiliki relasi pegawai.') from exc

    def get_queryset(self, model):
        employee_field = self.get_employee_field(model)
        queryset = model.objects.all()
        if employee_field.many_to_many:
            queryset = queryset.prefetch_related('pegawai__profil_user')
        else:
            queryset = queryset.select_related('pegawai', 'pegawai__profil_user')

        selected_nip = get_selected_nip(self.request)
        if self.request.user.is_dokumen_admin:
            if selected_nip:
                queryset = queryset.filter(pegawai__profil_user__nip=selected_nip)
        else:
            queryset = queryset.filter(pegawai=self.request.user)

        if (
            model is RiwayatJabatan
            and self.request.user.is_dokumen_admin
            and not selected_nip
        ):
            jabatan = (self.request.GET.get('jabatan') or '').strip()
            if jabatan:
                queryset = queryset.filter(jns_jabatan__icontains=jabatan)

        field_names = {field.name for field in model._meta.fields}
        order_by = ['no_urut_dokumen', 'pk'] if 'no_urut_dokumen' in field_names else ['pk']
        return queryset.order_by(*order_by).distinct()

    @staticmethod
    def get_export_fields(model):
        return [
            field for field in model._meta.fields
            if field.name not in EXCLUDED_FIELDS
        ]

    def get_employees(self, instance, employee_field):
        if employee_field.many_to_many:
            if not self.request.user.is_dokumen_admin:
                return [self.request.user]
            return list(instance.pegawai.all())
        employee = instance.pegawai
        return [employee] if employee else []

    @staticmethod
    def get_nip(employee):
        profile = getattr(employee, 'profil_user', None)
        return getattr(profile, 'nip', '') if profile else ''

    def serialize_field(self, instance, field):
        display_method = getattr(instance, f'get_{field.name}_display', None)
        if callable(display_method):
            return csv_safe(display_method())

        value = getattr(instance, field.name)
        if isinstance(field, models.FileField):
            if not value:
                return ''
            try:
                return csv_safe(self.request.build_absolute_uri(value.url))
            except (ValueError, NotImplementedError):
                return csv_safe(value.name)
        return csv_safe(value)

    def stream_rows(self, queryset, export_fields, employee_field):
        writer = csv.writer(Echo())
        yield '\ufeff'
        yield writer.writerow([
            'Nama Pegawai',
            'NIP',
            *[str(field.verbose_name).title() for field in export_fields],
        ])
        for instance in queryset.iterator(chunk_size=1000):
            employees = self.get_employees(instance, employee_field)
            yield writer.writerow([
                csv_safe('; '.join(employee.full_name for employee in employees)),
                csv_safe('; '.join(self.get_nip(employee) for employee in employees)),
                *[
                    self.serialize_field(instance, field)
                    for field in export_fields
                ],
            ])

    def get(self, request, *args, **kwargs):
        title, model = self.get_configuration()
        queryset = self.get_queryset(model)
        export_fields = self.get_export_fields(model)
        employee_field = self.get_employee_field(model)
        filename = f'{self.kwargs["document_type"]}-{date.today():%Y%m%d}.csv'
        response = StreamingHttpResponse(
            self.stream_rows(queryset, export_fields, employee_field),
            content_type='text/csv; charset=utf-8',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['X-Content-Type-Options'] = 'nosniff'
        response['X-Document-Export'] = title
        return response


class ProfessionSIPExportExcelView(LoginRequiredMixin, View):
    """Ekspor STR dan seluruh riwayat SIP sesuai cakupan akses pengguna."""

    http_method_names = ['get']
    headers = (
        'Nama Pegawai',
        'Jabatan',
        'Nomor STR',
        'Tanggal STR',
        'Tanggal Terbit SIP',
        'Tanggal Berakhir SIP',
        'Nomor SIP',
    )

    def get_profession_queryset(self):
        sip_queryset = RiwayatSIPProfesi.objects.order_by(
            '-tgl_sip',
            '-pk',
        )
        queryset = (
            RiwayatProfesi.objects
            .select_related('pegawai', 'pegawai__profil_user')
            .prefetch_related(
                models.Prefetch(
                    'riwayatsipprofesi_set',
                    queryset=sip_queryset,
                    to_attr='sip_history',
                )
            )
        )
        queryset = filter_profession_history_queryset(
            queryset, self.request.user
        )
        selected_nip = get_selected_nip(self.request)
        if selected_nip:
            queryset = queryset.filter(
                pegawai__profil_user__nip=selected_nip,
            )
        return queryset.order_by(
            'pegawai__first_name',
            'pegawai__last_name',
            'pegawai_id',
            'no_urut_dokumen',
            'pk',
        )

    @staticmethod
    def get_latest_jabatan_map(profession_rows):
        employee_ids = {
            profession.pegawai_id
            for profession in profession_rows
            if profession.pegawai_id
        }
        if not employee_ids:
            return {}

        latest_jabatan_pk = (
            RiwayatJabatan.objects
            .filter(
                models.Q(tmt_jabatan__lte=date.today())
                | models.Q(tmt_jabatan__isnull=True),
                pegawai_id=models.OuterRef('pegawai_id'),
            )
            .order_by(
                models.F('tmt_jabatan').desc(nulls_last=True),
                '-updated_at',
                '-pk',
            )
            .values('pk')[:1]
        )
        latest_rows = (
            RiwayatJabatan.objects
            .filter(
                pegawai_id__in=employee_ids,
                pk=models.Subquery(latest_jabatan_pk),
            )
            .select_related('nama_jabatan')
        )
        return {row.pegawai_id: row for row in latest_rows}

    @staticmethod
    def get_jabatan_display(jabatan):
        if jabatan is None:
            return '-'
        nama_jabatan = (
            jabatan.nama_jabatan.jenis_sdm
            if jabatan.nama_jabatan else ''
        )
        detail = (jabatan.detail_nama_jabatan or '').strip()
        if nama_jabatan and detail:
            return f'{nama_jabatan} ({detail})'
        return nama_jabatan or detail or '-'

    @staticmethod
    def safe_text(value):
        if value is None:
            return ''
        text = str(value)
        if text.startswith(('=', '+', '-', '@', '\t', '\r')):
            return f"'{text}"
        return text

    def append_data_rows(self, worksheet, profession_rows, jabatan_map):
        for profession in profession_rows:
            sip_rows = profession.sip_history or [None]
            jabatan = self.get_jabatan_display(
                jabatan_map.get(profession.pegawai_id)
            )
            for sip in sip_rows:
                worksheet.append((
                    self.safe_text(profession.pegawai.full_name),
                    self.safe_text(jabatan),
                    self.safe_text(profession.no_str),
                    profession.tgl_str,
                    sip.tgl_sip if sip else None,
                    sip.berlaku_sd if sip else None,
                    self.safe_text(sip.no_sip if sip else ''),
                ))

    def style_worksheet(self, worksheet, last_row):
        header_row = 3
        header_fill = PatternFill('solid', fgColor='1F4E78')
        thin_border = Border(
            left=Side(style='thin', color='D9E2F3'),
            right=Side(style='thin', color='D9E2F3'),
            top=Side(style='thin', color='D9E2F3'),
            bottom=Side(style='thin', color='D9E2F3'),
        )
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        if last_row > header_row:
            for row in worksheet.iter_rows(
                min_row=header_row + 1,
                max_row=last_row,
                min_col=1,
                max_col=len(self.headers),
            ):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top')
                for date_column in (4, 5, 6):
                    row[date_column - 1].number_format = 'dd-mm-yyyy'

        for column, width in {
            'A': 30,
            'B': 32,
            'C': 22,
            'D': 18,
            'E': 20,
            'F': 22,
            'G': 22,
        }.items():
            worksheet.column_dimensions[column].width = width
        worksheet.freeze_panes = 'A4'
        if last_row >= header_row:
            worksheet.auto_filter.ref = f'A{header_row}:G{last_row}'

    def get(self, request, *args, **kwargs):
        profession_rows = list(self.get_profession_queryset())
        jabatan_map = self.get_latest_jabatan_map(profession_rows)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Riwayat Profesi dan SIP'
        worksheet.merge_cells('A1:G1')
        worksheet['A1'] = 'RIWAYAT PROFESI, STR, DAN SIP PEGAWAI'
        worksheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        worksheet['A1'].fill = PatternFill('solid', fgColor='1F4E78')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet.append(())
        worksheet.append(self.headers)
        self.append_data_rows(worksheet, profession_rows, jabatan_map)
        self.style_worksheet(worksheet, worksheet.max_row)

        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        )
        response['Content-Disposition'] = (
            f'attachment; filename="riwayat_profesi_sip_{date.today():%Y%m%d}.xlsx"'
        )
        response['X-Content-Type-Options'] = 'nosniff'
        workbook.save(response)
        return response
