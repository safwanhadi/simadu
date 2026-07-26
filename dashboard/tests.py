from datetime import date, datetime, time
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from disiplinsdm.models import AbsensiHarian, LogAktivitasAbsen
from myaccount.models import ProfilAdmin, Users
from strukturorg.models import (
    Bidang,
    InstansiDaerah,
    SatuanKerjaInduk,
    SubBidang,
    UnitInstalasi,
    UnitOrganisasi,
)

from .views import get_accessible_takah


class DashboardAbsensiHariIniTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = Users.objects.create_superuser(
            email='admin-dashboard-absensi@example.com',
            first_name='Admin',
            last_name='Absensi',
            password='Password-123!',
        )
        cls.pegawai = Users.objects.create_user(
            email='pegawai-sudah-presensi@example.com',
            first_name='Sudah',
            last_name='Presensi',
            password='Password-123!',
        )
        cls.pegawai_apel = Users.objects.create_user(
            email='pegawai-hanya-apel@example.com',
            first_name='Hanya',
            last_name='Apel',
            password='Password-123!',
        )
        instansi = InstansiDaerah.objects.create(instansi='Instansi Tes')
        satker = SatuanKerjaInduk.objects.create(
            instansi_daerah=instansi,
            satuan_kerja='Satker Tes',
        )
        unor = UnitOrganisasi.objects.create(
            satker_induk=satker,
            unor='Unor Tes',
        )
        bidang = Bidang.objects.create(unor=unor, bidang='Bidang Tes')
        sub_bidang = SubBidang.objects.create(
            bidang=bidang,
            sub_bidang='Subbidang Tes',
        )
        instalasi = UnitInstalasi.objects.create(
            sub_bidang=sub_bidang,
            instalasi='Instalasi Tes',
        )
        absensi = AbsensiHarian.objects.create(
            pegawai=cls.pegawai,
            tanggal=date.today(),
            unor=unor,
            instalasi=instalasi,
            status_final='HADIR',
        )
        LogAktivitasAbsen.objects.create(
            absensi_harian=absensi,
            tipe='DATANG',
            waktu=datetime.combine(date.today(), time(7, 30)),
            status_ketepatan='Tepat Waktu',
        )
        LogAktivitasAbsen.objects.create(
            absensi_harian=absensi,
            tipe='PULANG',
            waktu=datetime.combine(date.today(), time(16, 0)),
        )
        absensi_apel = AbsensiHarian.objects.create(
            pegawai=cls.pegawai_apel,
            tanggal=date.today(),
            unor=unor,
            instalasi=instalasi,
            status_final='',
        )
        LogAktivitasAbsen.objects.create(
            absensi_harian=absensi_apel,
            tipe='APEL',
            waktu=datetime.combine(date.today(), time(7, 0)),
        )

    def test_dashboard_hanya_menampilkan_pegawai_dengan_presensi_datang(self):
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse('dashboard_urls:dashboard_absensi_view')
        )

        self.assertEqual(response.status_code, 200)
        daftar = list(response.context['sudah_presensi_hari_ini'])
        self.assertEqual([item.pegawai for item in daftar], [self.pegawai])
        self.assertContains(response, 'Sudah Presensi Hari Ini')
        self.assertContains(response, '07:30:00')
        self.assertContains(response, '16:00:00')
        self.assertEqual(response.context['total_presensi_hari_ini'], 1)
        self.assertEqual(response.context['tepat_waktu_hari_ini'], 1)
        self.assertEqual(response.context['belum_pulang_hari_ini'], 0)
        self.assertEqual(response.context['tk_kemarin_presensi_hari_ini'], 1)

    def test_daftar_lama_tidak_lagi_tersedia(self):
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse('dashboard_urls:dashboard_absensi_view')
        )

        self.assertNotIn('hari_ini_alpa', response.context)
        self.assertNotIn('hari_ini_tidak_apel', response.context)
        self.assertNotIn('belum_presensi_hari_ini', response.context)
        self.assertNotIn('kemarin_tidak_apel', response.context)
        self.assertNotContains(response, 'Daftar TK/ALPA Hari Ini')
        self.assertNotContains(response, 'Belum Presensi Hari Ini')
        self.assertNotContains(response, 'Tidak Ikut Apel Hari Ini')
        self.assertNotContains(response, 'Tidak Ikut Apel Kemarin')

    def test_excel_memisahkan_tk_kemarin_dan_presensi_hari_ini(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse('dashboard_urls:export_harian_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['TK Kemarin', 'Presensi Hari Ini'])

        tk_values = [
            str(cell.value)
            for row in workbook['TK Kemarin'].iter_rows()
            for cell in row
            if cell.value is not None
        ]
        presensi_values = [
            str(cell.value)
            for row in workbook['Presensi Hari Ini'].iter_rows()
            for cell in row
            if cell.value is not None
        ]
        isi_tk = ' '.join(tk_values)
        isi_presensi = ' '.join(presensi_values)

        self.assertIn('Sudah Presensi', isi_tk)
        self.assertIn('Hanya Apel', isi_tk)
        self.assertNotIn('07:30:00', isi_tk)

        self.assertIn('Sudah Presensi', isi_presensi)
        self.assertNotIn('Hanya Apel', isi_presensi)
        self.assertIn('07:30:00', isi_presensi)
        self.assertIn('16:00:00', isi_presensi)


class DashboardTakahAccessTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.staff = Users.objects.create_user(
            email='staff-tanpa-profil@example.com',
            first_name='Staff',
            last_name='Tanpa Profil',
            password='Password-123!',
            is_staff=True,
        )
        cls.staff_empty_scope = Users.objects.create_user(
            email='staff-scope-kosong@example.com',
            first_name='Staff',
            last_name='Scope Kosong',
            password='Password-123!',
            is_staff=True,
        )
        ProfilAdmin.objects.create(user=cls.staff_empty_scope)
        cls.pegawai = Users.objects.create_user(
            email='pegawai-dashboard@example.com',
            first_name='Pegawai',
            last_name='Dashboard',
            password='Password-123!',
        )

    def test_staff_tanpa_profil_admin_mendapat_queryset_kosong(self):
        self.assertFalse(get_accessible_takah(self.staff).exists())

    def test_staff_tanpa_cakupan_mendapat_queryset_kosong(self):
        self.assertFalse(get_accessible_takah(self.staff_empty_scope).exists())

    def test_dashboard_staff_tanpa_profil_tidak_error(self):
        self.client.force_login(self.staff)

        response = self.client.get(reverse('dashboard_urls:dashboard_view'))

        self.assertEqual(response.status_code, 200)
        self.assertQuerySetEqual(response.context['takah'], [])
