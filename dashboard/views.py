from django.shortcuts import render, redirect
from django.urls import reverse
from django.views import View
from django.views.generic import ListView, TemplateView
from django.db.models import Sum, F, Q, Case, When, Value, Count, Prefetch, Max, Min, IntegerField, ExpressionWrapper, FloatField, OuterRef, Subquery, Exists
from django.db.models.functions import Concat, Coalesce
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime, date, time, timedelta
from rest_framework.views import APIView
from rest_framework.response import Response
import calendar
from django.utils.timezone import make_aware

from disiplinsdm.models import (
    KehadiranKegiatan, 
    AbsensiHarian, 
    LogAktivitasAbsen, 
    ApprovedJadwalDinasSDM,
    JadwalDinasSDM,
    JenisSDMPerinstalasi, 
    HariLibur
)
from jenissdm.models import JenisSDM 
from strukturorg.models import StandarInstalasi, UnitInstalasi
from dokumen.models import RiwayatJabatan, RiwayatPenempatan, Kompetensi, RiwayatProfesi
from disiplinsdm.models import DaftarKegiatanPegawai
from myaccount.models import Users
from dokumen.views import file_kepegawaian
from itertools import zip_longest

from .services import (
    employees_for_jabatan,
    installation_groups,
    installation_standard_data,
    installation_standard_summaries,
    jabatan_cards,
    workforce_summary,
)

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_accessible_takah(user):
    """Daftar pegawai yang boleh dilihat pada kartu takah dashboard."""
    queryset = Users.objects.exclude(is_superuser=True, pegawai__is_active=False)
    if user.is_superuser:
        return queryset
    if not user.is_staff or not hasattr(user, 'profil_admin'):
        return Users.objects.none()

    profil_admin = user.profil_admin
    scope_fields = (
        ('instalasi', 'riwayat_penempatan__penempatan_level4'),
        ('sub_bidang', 'riwayat_penempatan__penempatan_level3'),
        ('bidang', 'riwayat_penempatan__penempatan_level2'),
        ('unor', 'riwayat_penempatan__penempatan_level1'),
    )
    for admin_field, placement_field in scope_fields:
        scope = getattr(profil_admin, admin_field)
        if scope.exists():
            return queryset.filter(
                **{
                    f'{placement_field}__in': scope.all(),
                    'riwayat_penempatan__status': True,
                }
            ).distinct()
    return Users.objects.none()

import logging

logger = logging.getLogger(__name__)

# logger.debug('This is a debug message')
# logger.info('This is an info message')
# logger.warning('This is a warning message')

# Create your views here.


def pegawai_tk_pada_tanggal(target_date):
    """Pegawai aktif yang tidak memiliki kehadiran/izin/libur sah pada tanggal."""
    aman_ids = AbsensiHarian.objects.filter(
        tanggal=target_date,
        status_final__in=['HADIR', 'IZIN', 'DINAS', 'LIBUR'],
    ).values_list('pegawai_id', flat=True).exclude(pegawai__is_active=False)

    libur_jadwal_ids = list(
        JadwalDinasSDM.objects.filter(tanggal=target_date)
        .filter(
            Q(kategori_jadwal__kategori_jadwal__icontains='libur')
            | Q(kategori_jadwal__kategori_dinas__kategori_dinas__icontains='libur')
        )
        .values_list('pegawai__pegawai_id', flat=True)
    )

    if target_date.weekday() == 6 or HariLibur.objects.filter(tanggal=target_date).exists():
        libur_jadwal_ids.extend(
            JadwalDinasSDM.objects.filter(
                tanggal=target_date,
                kategori_jadwal__kategori_dinas__kategori_dinas='Reguler',
            ).values_list('pegawai__pegawai_id', flat=True)
        )

    return (
        Users.objects.filter(is_active=True)
        .exclude(
            Q(is_superuser=True)
            | Q(id__in=aman_ids)
            | Q(id__in=libur_jadwal_ids)
        )
    )


def presensi_datang_pada_tanggal(target_date):
    """Absensi harian yang benar-benar memiliki transaksi presensi DATANG."""
    return (
        AbsensiHarian.objects.filter(tanggal=target_date).exclude(pegawai__is_active=False)
        .annotate(
            memiliki_presensi_datang=Exists(
                LogAktivitasAbsen.objects.filter(
                    absensi_harian_id=OuterRef('pk'),
                    tipe='DATANG',
                )
            )
        )
        .filter(memiliki_presensi_datang=True)
        .select_related('pegawai', 'pegawai__profil_user', 'instalasi', 'unor', 'bidang')
        .annotate(
            jam_datang=Min('logs__waktu', filter=Q(logs__tipe='DATANG')),
            jam_pulang=Max('logs__waktu', filter=Q(logs__tipe='PULANG')),
            status_datang=Max(
                'logs__status_ketepatan',
                filter=Q(logs__tipe='DATANG'),
            ),
        )
    )

def get_date_from_string(tanggal):
    tanggal_sekarang = datetime.now()
    try:
        get_tanggal = datetime.strptime(tanggal, "%Y-%m-%d").date()
        return get_tanggal
    except Exception:
        return tanggal_sekarang.date()
    

class StandarSDMInstalasi(LoginRequiredMixin, View):
    login_url = '/accounts/login/'
    redirect_field_name = 'next'

    def get_instalasi(self, id):
        try:
            data = UnitInstalasi.objects.get(id=id)
            return data
        except UnitInstalasi.DoesNotExist:
            return None
        
    def takahpagination(self, p):
        page_number = self.request.GET.get('page')
        try:
            page_obj = p.get_page(page_number)  # returns the desired page object
        except PageNotAnInteger:
            # if page_number is not an integer then assign the first page
            page_obj = p.page(1)
        except EmptyPage:
            # if page is empty then return last page
            page_obj = p.page(p.num_pages)
        return page_obj

    def check_standar_perinstalasi(self, data, instalasi):
        if instalasi is not None and hasattr(instalasi, 'instalasi'):
            if any(d['status']=='1_kurang' and d['pegawai'] is not None and d['instalasi'] == instalasi.instalasi for d in data):
                #logger.debug({'instalasi': instalasi.instalasi, 'status':'1_kurang'})
                return {'instalasi': instalasi.instalasi, 'status':'1_kurang'}
            elif any(d['status']=='2_bagus' and d['pegawai'] is not None and d['instalasi'] == instalasi.instalasi for d in data):
                return {'instalasi': instalasi.instalasi, 'status':'2_bagus'}
            elif any(d['status']=='3_mantap' and d['pegawai'] is not None and d['instalasi'] == instalasi.instalasi for d in data):
                return {'instalasi': instalasi.instalasi, 'status':'3_mantap'} 
            else:
                return {}  
    
    def check_standar_persdm(self, data):
        standar = StandarInstalasi.objects.all().order_by('instalasi')
        get_data = {}
        for item in standar:
            if data is not None and item.jenis_sdm == data.nama_jabatan and item.instalasi == data.instalasi:
                standar_wajib = item.kompetensi_wajib.values_list('kompetensi', flat=True)
                standar_pendukung = item.kompetensi_pendukung.values_list('kompetensi', flat=True)
                data_instalasi = data.kompetensi.values_list('kompetensi__kompetensi', flat=True)
                # if set(standar_wajib) == set(data_instalasi):
                if all(item in data_instalasi for item in standar_wajib):
                    get_data.update({
                        'id':item.instalasi.id,
                        'pegawai':data.pegawai,
                        'instalasi':str(item.instalasi),
                        'status':'2_bagus'
                        })
                    if any(item in data_instalasi for item in standar_pendukung):
                        get_data.update({
                            'id':item.instalasi.id,
                            'pegawai':data.pegawai,
                            'instalasi':str(item.instalasi),
                            'status':'3_mantap'
                            })
                    return get_data
                else:
                    get_data.update({
                        'id':item.instalasi.id,
                        'pegawai':data.pegawai,
                        'instalasi':str(item.instalasi),
                        'status':'1_kurang'
                        })
                    return get_data
            get_data.update({'id':item.instalasi.id, 'pegawai':None, 'instalasi':str(item.instalasi), 'status':'1_kurang'})
        return get_data
        
    def get(self, request):
        nip = request.GET.get('nip')
        id_instalasi = request.GET.get('instalasi')
        inst = request.GET.get('inst')
        if not inst:
            instalasi = UnitInstalasi.objects.first()
            inst = instalasi.id if instalasi is not None and hasattr(instalasi, 'id') else 0
        data_instalasi = UnitInstalasi.objects.all().order_by('id')
        status=[]
        status_instalasi = []
        takah = get_accessible_takah(request.user)
        tgl = request.GET.get('tanggal')
        get_tanggal = get_date_from_string(tgl)
        tanggal = datetime.now().strftime("%Y-%m-%d")
        if tgl:
            tanggal = datetime(get_tanggal.year, get_tanggal.month, get_tanggal.day).strftime("%Y-%m-%d")
        #menampilkan standarisasi SDM berdasarkan komptensi dalam dashboard
        for item in data_instalasi:
            if item:
                detail_instalasi = item.riwayatjabatan_set.filter(kompetensi__berlaku_sd__gte=tanggal).order_by('instalasi')
                for item2 in detail_instalasi:
                    data_status = self.check_standar_persdm(item2)
                    status.append(data_status)

        for item in data_instalasi:
            if item is not None and len(status) != 0 and len(status[0]) > 0:
                standar= self.check_standar_perinstalasi(status, item)
                status_instalasi.append(standar)
        data = zip(data_instalasi, status_instalasi)
        #menampilkan takah pegawai di dashoard
        file_kepeg = file_kepegawaian(request, nip)
        
        data_fungsional = None
        if id_instalasi:
            get_data_fungsional = RiwayatJabatan.objects.filter(instalasi__id=id_instalasi).order_by('instalasi')
            status = [d for d in status if int(d['id']) == int(id_instalasi)]
            data_fungsional = zip_longest(get_data_fungsional, status)
        #menampilkan data pegawai dalam card dashboard
        # jenissdm = JenisSDM.objects.all().annotate(
        #     jumlah = 
        # )
        jenissdm = JenisSDM.objects.annotate(
            jumlah = Count(F('riwayatjabatan__pegawai'), distinct=True)
        ).distinct()
        #p_takah = Paginator(takah, 10)
        #takah_page = self.takahpagination(p_takah)
        context = {
            'nip':nip,
            'data':data,
            'takah':takah,
            'file_kepeg':file_kepeg,
            'data_fungsional':data_fungsional,
            'id_instalasi':id_instalasi,
            'inst':inst,
            'instalasi':self.get_instalasi(inst),
            'list_instalasi':data_instalasi,
            'jenissdm': jenissdm,
            'tanggal':tanggal,
            'bulan': get_tanggal.month,
            'dash':'active',
            'title_page':'Home',
            'page':'Home'
        }
        if request.user.is_superuser or request.user.is_staff:
            return render(request, 'standar_sdm_instalasi.html', context)
        else:
            return redirect(reverse('riwayat_urls:riwayat_view'))
    

class StandarInstalasiView(LoginRequiredMixin, ListView):
    login_url = '/accounts/login/'
    redirect_field_name = 'next'
    model = UnitInstalasi
    template_name = 'standar_sdm_instalasi.html'
    context_object_name = 'instalasi_list'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return redirect(reverse('riwayat_urls:riwayat_view'))  # Redirect unauthorized users
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        return installation_groups()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        instalasi_list = context['instalasi_list']
        
        installation_slug = self.request.GET.get('instalasi')
        selected_installation = None
        if installation_slug:
            selected_installation = next(
                (
                    installation
                    for installation in instalasi_list
                    if installation['slug'] == installation_slug
                ),
                None,
            )

        if selected_installation:
            standard_data = installation_standard_data(selected_installation)
            context['users'] = standard_data['users']
            context['installation_status'] = standard_data['status']
        else:
            context['data_instalasi'] = installation_standard_summaries(
                instalasi_list
            )
                
        #untuk keperluan menampilkan chart disiplin pegawai
        inst = self.request.GET.get('inst')
        if inst is None:
            inst = 0
        tgl = self.request.GET.get('tanggal')
        get_tanggal = get_date_from_string(tgl)
        tanggal = datetime.now().strftime("%Y-%m-%d")
        if tgl:
            tanggal = datetime(get_tanggal.year, get_tanggal.month, get_tanggal.day).strftime("%Y-%m-%d")
        #menampilkan data pegawai dalam card dashboard
        jenissdm = list(jabatan_cards())
        context['inst'] = inst
        context['slug'] = selected_installation['slug'] if selected_installation else None
        context['selected_installation'] = (
            selected_installation['instalasi'] if selected_installation else None
        )
        context['tanggal'] = tanggal
        context['bulan'] = get_tanggal.month
        context['jenissdm'] = jenissdm
        context['jabatan_total'] = sum(item.jumlah for item in jenissdm)
        context['workforce_summary'] = workforce_summary()
        # context['data_instalasi'] = data_instalasi
        context['dash'] = 'active'
        context['title_page'] = 'Standar Instalasi'
        context['page'] = 'Standar Instalasi'
        return context
    

class ExportWorkforceProfessionExcelView(
    LoginRequiredMixin,
    UserPassesTestMixin,
    View,
):
    """Ekspor ringkasan SDM aktif dan jabatan terakhir pegawai."""

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        return redirect('dashboard_urls:dashboard_view')

    def get(self, request, *args, **kwargs):
        summary = workforce_summary()
        jabatan_list = list(jabatan_cards())

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Ringkasan SDM'
        worksheet.merge_cells('A1:D1')
        worksheet['A1'] = 'RINGKASAN SDM AKTIF PER JABATAN TERAKHIR'
        worksheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        worksheet['A1'].fill = PatternFill('solid', fgColor='1F4E78')
        worksheet['A1'].alignment = Alignment(horizontal='center')

        summary_rows = (
            ('Total SDM aktif', summary['total_active']),
            ('SDM tanpa penempatan aktif', summary['without_active_placement']),
            ('SDM tanpa jabatan', summary['without_jabatan']),
        )
        for row_number, (label, value) in enumerate(summary_rows, start=3):
            worksheet.cell(row=row_number, column=1, value=label).font = Font(bold=True)
            worksheet.cell(row=row_number, column=2, value=value)

        header_row = 7
        headers = ('No', 'Jabatan', 'Kategori SDM', 'Jumlah Pegawai')
        for column, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=column, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4472C4')
            cell.alignment = Alignment(horizontal='center')

        thin_border = Border(
            left=Side(style='thin', color='D9E2F3'),
            right=Side(style='thin', color='D9E2F3'),
            top=Side(style='thin', color='D9E2F3'),
            bottom=Side(style='thin', color='D9E2F3'),
        )
        for index, jabatan in enumerate(jabatan_list, start=1):
            row = header_row + index
            values = (
                index,
                jabatan.jenis_sdm,
                jabatan.get_kategori_sdm_display() or '-',
                jabatan.jumlah,
            )
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row, column=column, value=value)
                cell.border = thin_border
                if column in (1, 4):
                    cell.alignment = Alignment(horizontal='center')

        total_row = header_row + len(jabatan_list) + 1
        worksheet.merge_cells(
            start_row=total_row,
            start_column=1,
            end_row=total_row,
            end_column=3,
        )
        total_label = worksheet.cell(
            row=total_row,
            column=1,
            value='TOTAL TERHITUNG',
        )
        total_label.font = Font(bold=True)
        total_label.fill = PatternFill('solid', fgColor='D9EAD3')
        total_label.alignment = Alignment(horizontal='right')
        total_value = worksheet.cell(
            row=total_row,
            column=4,
            value=sum(jabatan.jumlah for jabatan in jabatan_list),
        )
        total_value.font = Font(bold=True)
        total_value.fill = PatternFill('solid', fgColor='D9EAD3')
        total_value.alignment = Alignment(horizontal='center')

        worksheet.freeze_panes = f'A{header_row + 1}'
        worksheet.auto_filter.ref = f'A{header_row}:D{header_row + len(jabatan_list)}'
        for column, width in {'A': 8, 'B': 32, 'C': 22, 'D': 18}.items():
            worksheet.column_dimensions[column].width = width

        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        )
        response['Content-Disposition'] = (
            f'attachment; filename="ringkasan_sdm_jabatan_{date.today():%Y%m%d}.xlsx"'
        )
        workbook.save(response)
        return response


class DetailNakes(LoginRequiredMixin, View):
    login_url = '/accounts/login/'
    redirect_field_name = 'next'
    
    def get_object(self, profession_id):
        try:
            data = JenisSDM.objects.get(pk=profession_id)
            return data
        except JenisSDM.DoesNotExist:
            return None
        
    def get(self, request, *args, **kwargs):
        jabatan_id = kwargs.get('sdm')
        
        # # Ambil riwayat jabatan terakhir setiap pegawai berdasarkan profesi
        # latest_riwayat = RiwayatJabatan.objects.filter(nama_jabatan__slug=slug_jenis_nakes).values('pegawai').annotate(
        #     latest_tanggal=Max('created_at')
        # )
        # # Filter hanya data dengan tanggal riwayat terbaru untuk setiap pegawai
        # data = RiwayatJabatan.objects.filter(
        #     nama_jabatan__slug=slug_jenis_nakes,
        #     created_at__in=[entry['latest_tanggal'] for entry in latest_riwayat]
        # ).select_related("pegawai", "nama_jabatan")
        data = employees_for_jabatan(jabatan_id)
        
        jenissdm = self.get_object(jabatan_id)
        jenisnakes = jabatan_cards()
        context={
            'data':data,
            'jenissdm':jenissdm,
            'jenisnakes':jenisnakes,
            'dash':'active',
            'page':'Home',
            'sub_page':'Detail SDM',
            'title_page': ''
        }
        return render(request, 'detail_nakes.html', context)
    

class KehadiranGrafikView(APIView):
    def get(self, request):
        
        #untuk menampilkan data kehadiran apel pegawai di dashboard
        tanggal = request.GET.get('tanggal')
        inst = request.GET.get('inst', 0)
        disiplin = None
        data = DaftarKegiatanPegawai.objects.all().exclude(pegawai__is_active=False)
        ##filter data berdasarkan tanggal (default tanggal hari ini)
        get_tanggal = get_date_from_string(tanggal)
        if get_tanggal:
            data = DaftarKegiatanPegawai.objects.filter(bulan=get_tanggal.month, tahun=get_tanggal.year).order_by('id').exclude(pegawai__is_active=False)
            disiplin = data
        instalasi = UnitInstalasi.objects.first()
        disiplin = data.filter(instalasi__id=instalasi.id)
        if inst is not None:
            disiplin = data.filter(instalasi__id=inst)
        data = {
            'label': [f'{item.pegawai.first_name} {item.pegawai.last_name}' for item in disiplin],
            'data': [item.jumlah_tk for item in disiplin]
        }
        return Response(data)
    

class ProsentaseKedisiplinanInstalasi(APIView):
    '''
    Prosentase ini didapatkan dari ==> (Jumlah SDM yang hadir apel pagi dalam 1 bulan / jumlah kali presensi seharusnya seluruh SDM yang hadir dan TK 
    dalam 1 bulan) * 100%  (dengan status kehadiran SAKIT, IJIN, WFH, PIKET, TUBEL DAN TUGAS DINAS tidak dihituang)

    '''
    def get(self, request):
        tanggal =request.GET.get('tanggal')
        inst = request.GET.get('inst')
        get_tanggal = get_date_from_string(tanggal)
        data = DaftarKegiatanPegawai.objects.all()
        if get_tanggal:
            data = DaftarKegiatanPegawai.objects.filter(bulan=get_tanggal.month, tahun=get_tanggal.year).values('bulan', 'instalasi__instalasi').annotate(
                jlh_sdm_presensi = Count(Case(When(Q(bulan=get_tanggal.month) & Q(kehadirankegiatan__alasan__alasan='Tanpa Keterangan') | Q(kehadirankegiatan__hadir=True), then=F('pegawai')))),
                pegawai_hadir = Count(Case(When(Q(kehadirankegiatan__hadir=True) & Q(bulan=get_tanggal.month), then=F('pegawai')))),
                prosentase_disiplin = F('pegawai_hadir')*100/F('jlh_sdm_presensi')
            ).order_by('instalasi__instalasi')
        context={
            'label': [item['instalasi__instalasi'] for item in data],
            'data':[item['prosentase_disiplin'] for item in data]
        }
        return Response(data=context)


class DashboardAbsensiView(TemplateView):
    template_name = 'absensi/dashboard_absensi.html'

    def safe_int(self, val, default):
        """Helper function to safely convert a value to integer."""
        try:
            return int(val) if str(val).isdigit() else default
        except (ValueError, TypeError):
            return default
        
    def _get_preserved_query(self):
        """
        ## REFACTOR:
        - Fungsi ini sekarang secara eksplisit menghapus SEMUA kemungkinan parameter
          halaman (`page_malas` dan `page_disiplin`).
        - Ini memastikan bahwa query string yang "disimpan" hanya berisi parameter
          filter (seperti periode, bulan, tahun), bukan nomor halaman.
        """
        querydict = self.request.GET.copy()
        querydict.pop('page_malas', None)
        querydict.pop('page_disiplin', None)
        return querydict.urlencode()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        kemarin = date.today() - timedelta(days=1)
        current_year = date.today().year

        # --- 1. Pengaturan Filter ---
        periode = self.request.GET.get('periode', 'bulanan')
        bulan = self.safe_int(self.request.GET.get('bulan'), date.today().month)
        tahun = self.safe_int(self.request.GET.get('tahun'), date.today().year)
        tanggal_str = self.request.GET.get('tgl', kemarin.strftime('%Y-%m-%d'))
        
        status_terlambat_all = ['Terlambat Ringan', 'Terlambat Sedang', 'Terlambat Berat', 'Terlambat']

        # --- 2. Query Data Utama ---
        # Apply date range filter based on selected period
        if periode == 'bulanan':
            tgl_awal = datetime(tahun, bulan, 1)
            hari_terakhir = calendar.monthrange(tahun, bulan)[1]
            tgl_akhir = datetime(tahun, bulan, hari_terakhir)
            start_date = datetime.combine(tgl_awal, time.min)
            end_date = datetime.combine(tgl_akhir, time.max)
        else:
            tgl = datetime.strptime(tanggal_str, '%Y-%m-%d').date()
            start_date = datetime.combine(tgl, time.min)
            end_date = datetime.combine(tgl, time.max)

        data = KehadiranKegiatan.objects.filter(
            tanggal__range=(start_date, end_date)
        ).exclude(
            pegawai__pegawai__is_active=False
        ).select_related('pegawai__pegawai', 'pegawai__instalasi')
        # print('data : ', data)
        # --- 3. Statistik Ringkas (Summary Boxes) ---
        ringkasan = data.aggregate(
            hadir_tepat_waktu=Count('id', filter=Q(status_ketepatan='Tepat Waktu')),
            total_terlambat=Count('id', filter=Q(status_ketepatan__in=status_terlambat_all)),
            total_tidak_hadir=Count('id', filter=Q(hadir=False)),
            total_cepat_pulang=Count('id', filter=Q(status_ketepatan='Cepat Pulang'))
        )
        context['statistik_kehadiran'] = [
            {'label': 'Hadir Tepat Waktu', 'jumlah': ringkasan['hadir_tepat_waktu'], 'color': 'info'},
            {'label': 'Terlambat', 'jumlah': ringkasan['total_terlambat'], 'color': 'warning'},
            {'label': 'Tidak Hadir', 'jumlah': ringkasan['total_tidak_hadir'], 'color': 'danger'},
            {'label': 'Cepat Pulang', 'jumlah': ringkasan['total_cepat_pulang'], 'color': 'primary'},
        ]
        
        
        # --- Chart 1: Persentase Kehadiran per Instalasi ---
        instalasi_data = data.filter(pegawai__instalasi__isnull=False).values('pegawai__instalasi__instalasi').annotate(
            total=Count('id'),
            hadir=Count('id', filter=Q(hadir=True))
        ).annotate(
            persen_hadir=Case(
                When(total=0, then=Value(0.0)),
                default=ExpressionWrapper(100.0 * F('hadir') / F('total'), output_field=FloatField())
            )
        ).order_by('-persen_hadir')
        
        context['chart_instalasi_hadir_labels'] = [item['pegawai__instalasi__instalasi'] or 'Lainnya' for item in instalasi_data]
        context['chart_instalasi_hadir_data'] = [round(item['persen_hadir'], 2) for item in instalasi_data]

        # --- Chart 2: TK dan Terlambat per Instalasi ---
        stats_pelanggaran_instalasi = data.filter(pegawai__instalasi__isnull=False).values('pegawai__instalasi__instalasi').annotate(
            tk=Count('id', filter=Q(hadir=False)),
            terlambat=Count('id', filter=Q(status_ketepatan__in=status_terlambat_all))
        ).order_by('-tk', '-terlambat')

        context['chart_instalasi_labels'] = [i['pegawai__instalasi__instalasi'] or 'Lainnya' for i in stats_pelanggaran_instalasi]
        context['chart_instalasi_tk_data'] = [i['tk'] for i in stats_pelanggaran_instalasi]
        context['chart_instalasi_terlambat_data'] = [i['terlambat'] for i in stats_pelanggaran_instalasi]
        
        # --- Queryset Dasar untuk Analisis Per Pegawai dengan Skor Pelanggaran ---
        pegawai_stats_base = data.filter(pegawai__pegawai__isnull=False).annotate(
            skor_pelanggaran=Case(
                When(hadir=False, then=Value(4)),
                When(status_ketepatan='Terlambat Berat', then=Value(3)),
                When(status_ketepatan='Terlambat Sedang', then=Value(2)),
                When(status_ketepatan__in=['Terlambat Ringan', 'Terlambat'], then=Value(1)),
                default=Value(0),
            )
        ).values('pegawai__pegawai__id').annotate(
            nama=Concat('pegawai__pegawai__first_name', Value(' '), 'pegawai__pegawai__last_name'),
            tk=Count('id', filter=Q(hadir=False)),
            terlambat=Count('id', filter=Q(status_ketepatan__in=status_terlambat_all)),
            tepat_waktu=Count('id', filter=Q(status_ketepatan='Tepat Waktu')),
            skor_terberat=Max('skor_pelanggaran')
        )

        # --- Chart 3 & Tabel: Pegawai Pelanggar Terbanyak (dengan sorting berbobot) ---
        top_pelanggar_qs = pegawai_stats_base.filter(
            Q(tk__gt=0) | Q(terlambat__gt=0)
        ).order_by('-skor_terberat', '-tk', '-terlambat', 'nama')

        # Data untuk Chart (Top 10)
        top_10_pelanggar = top_pelanggar_qs[:10]
        context['chart_pegawai_pelanggar_labels'] = [p['nama'] for p in top_10_pelanggar]
        context['chart_pegawai_pelanggar_tk_data'] = [p['tk'] for p in top_10_pelanggar]
        context['chart_pegawai_pelanggar_terlambat_data'] = [p['terlambat'] for p in top_10_pelanggar]
        
        # Data untuk Tabel Paginasi (Semua Pelanggar)
        paginator_malas = Paginator(top_pelanggar_qs, 10)
        context['top_malas'] = paginator_malas.get_page(self.request.GET.get('page_malas'))

        # --- Chart 4 & Tabel: Pegawai Paling Disiplin ---
        top_disiplin_qs = pegawai_stats_base.filter(tepat_waktu__gt=0).order_by('-tepat_waktu', 'nama')

        # Data untuk Chart (Top 10)
        top_10_disiplin = top_disiplin_qs[:10]
        context['chart_pegawai_terdisiplin_labels'] = [p['nama'] for p in top_10_disiplin]
        context['chart_pegawai_terdisiplin_data'] = [p['tepat_waktu'] for p in top_10_disiplin]

        # Data untuk Tabel Paginasi (Semua yang disiplin)
        paginator_disiplin = Paginator(top_disiplin_qs, 10)
        context['top_disiplin'] = paginator_disiplin.get_page(self.request.GET.get('page_disiplin'))

        # --- 5. Menambahkan data filter ke context ---
        context.update({
            'bulan_list': [(i, calendar.month_name[i]) for i in range(1, 13)],
            'tahun_list': list(range(current_year - 5, current_year + 2)),
            'nama_bulan': calendar.month_name[bulan],
            'kemarin_string': kemarin.strftime('%Y-%m-%d'),
            'periode': periode,
            'tanggal': tanggal_str,
            'bulan': bulan,
            'tahun': tahun,
            'dash2': 'active', # Untuk menandai menu aktif di sidebar
            'preserved_query': self._get_preserved_query()
        })

        return context
    

class DashboardAbsensiTemplateView(LoginRequiredMixin, TemplateView):
    template_name = 'absensi/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        yesterday = today - timedelta(days=1)
        
        # =========================================================================
        # INI ALUR LOGIKA UTAMA: INVENTARISASI PARAMETER & PERIODE
        # =========================================================================
        periode = self.request.GET.get('periode', 'bulan')
        pegawai_id = self.request.GET.get('pegawai', None)

        if periode == 'harian':
            start_date = today
        elif periode == 'minggu':
            start_date = today - timedelta(days=today.weekday())
        elif periode == 'tahun':
            start_date = date(today.year, 1, 1)
        else: 
            start_date = date(today.year, today.month, 1)

        end_date = today

        # Hak Akses: Superuser bisa ganti target pantauan, staf biasa hanya mengintip dirinya sendiri
        target_pegawai_id = pegawai_id if self.request.user.is_superuser else self.request.user.id
        target_user = Users.objects.get(pk=target_pegawai_id) if target_pegawai_id else self.request.user


        # =========================================================================
        # LOGIKA 1: STATISTIK INDIVIDU (UNTUK INFO BOX / WIDGET CARDS)
        # =========================================================================
        absensi_periode = AbsensiHarian.objects.filter(
            pegawai=target_user, 
            tanggal__range=(start_date, end_date)
        )

        stats_individual = absensi_periode.aggregate(
            alpa_count=Count('id', filter=Q(status_final='ALPA')),
            izin_count=Count('id', filter=Q(status_final='IZIN')),
            dinas_count=Count('id', filter=Q(status_final='DINAS')),
            libur_count=Count('id', filter=Q(status_final='LIBUR')),
        )

        logs_periode = LogAktivitasAbsen.objects.filter(
            absensi_harian__pegawai=target_user,
            absensi_harian__tanggal__range=(start_date, end_date)
        )

        stats_logs = logs_periode.aggregate(
            terlambat_count=Count('id', filter=Q(tipe='DATANG', status_ketepatan__icontains='Terlambat')),
            cepat_pulang_count=Count('id', filter=Q(tipe='PULANG', status_ketepatan='Cepat Pulang')),
        )

        hadir_ids = absensi_periode.filter(status_final='HADIR').values_list('id', flat=True)
        tidak_apel_count = AbsensiHarian.objects.filter(id__in=hadir_ids).exclude(logs__tipe='APEL').count()


        # =========================================================================
        # LOGIKA 2: MONITORING KEMARIN (FOKUS UTAMA EVALUASI ADMIN - MAX 50 USER)
        # =========================================================================
        
        # ALUR PIKIR MENCARI PEGAWAI MANGKIR/TANPA JADWAL KEMARIN:
        # Pikirkan ini secara terbalik: Kita cari siapa saja orang yang "Aman" kemarin,
        # lalu kita keluarkan mereka dari total daftar Users. Sisanya adalah orang yang "Bermasalah".
        
        kemarin_alpa_qs = pegawai_tk_pada_tanggal(yesterday)
        kemarin_alpa = kemarin_alpa_qs.prefetch_related(
            Prefetch(
                'riwayat_penempatan',
                queryset=RiwayatPenempatan.objects.filter(status=True).select_related('penempatan_level4'),
                to_attr='sk_aktif'
            )
        ).order_by('first_name', 'last_name')[:50]

        # =========================================================================
        # LOGIKA 3: MONITORING HARI INI (REALTIME RUNNING - MAX 50 USER)
        # =========================================================================
        presensi_hari_ini_qs = presensi_datang_pada_tanggal(today)
        sudah_presensi_hari_ini = presensi_hari_ini_qs.order_by(
            'jam_datang', 'pegawai__first_name', 'pegawai__last_name'
        )[:100]
        total_presensi_hari_ini = presensi_hari_ini_qs.count()
        tepat_waktu_hari_ini = presensi_hari_ini_qs.filter(
            status_datang__iexact='Tepat Waktu'
        ).count()
        terlambat_hari_ini = presensi_hari_ini_qs.filter(
            status_datang__icontains='Terlambat'
        ).count()
        belum_pulang_hari_ini = presensi_hari_ini_qs.filter(
            jam_pulang__isnull=True
        ).count()
        tk_kemarin_presensi_hari_ini = presensi_hari_ini_qs.filter(
            pegawai_id__in=kemarin_alpa_qs.values('id')
        ).count()


        # =========================================================================
        # LOGIKA 4: ANALISIS RANKING INSTALASI & DATA SELISIH JAM KURANG
        # =========================================================================
        instalasi_alpa = AbsensiHarian.objects.filter(
            tanggal__range=(start_date, end_date), status_final='ALPA'
        ).values('instalasi__instalasi').annotate(total=Count('id')).order_by('-total')[:5]

        instalasi_hadir = AbsensiHarian.objects.filter(
            tanggal__range=(start_date, end_date), status_final='HADIR'
        ).values('instalasi__instalasi').annotate(total=Count('id')).order_by('-total')[:5]

        jam_kurang_qs = JenisSDMPerinstalasi.objects.filter(
            pegawai=target_user, bulan=today.month, tahun=today.year
        ).first()


        # =========================================================================
        # LOGIKA 5: PACKAGING & PENGIRIMAN DATA KE HTML TEMPLATE
        # =========================================================================
        context.update({
            'periode': periode,
            'today': today,
            'target_user': target_user,
            'list_pegawai': Users.objects.filter(is_active=True).exclude(is_superuser=True).order_by('first_name') if self.request.user.is_superuser else None,
            
            # Counter Boxes
            'alpa_count': stats_individual['alpa_count'],
            'izin_count': stats_individual['izin_count'],
            'dinas_count': stats_individual['dinas_count'],
            'libur_count': stats_individual['libur_count'], 
            'terlambat_count': stats_logs['terlambat_count'],
            'tidak_apel_count': tidak_apel_count,
            'cepat_pulang_count': stats_logs['cepat_pulang_count'],
            'jam_kurang': jam_kurang_qs.selisih_jam_kerja if jam_kurang_qs else 0,
            
            # Lists Data Terbatas (Max 50)
            'kemarin_alpa': kemarin_alpa,
            'sudah_presensi_hari_ini': sudah_presensi_hari_ini,
            'total_presensi_hari_ini': total_presensi_hari_ini,
            'tepat_waktu_hari_ini': tepat_waktu_hari_ini,
            'terlambat_hari_ini': terlambat_hari_ini,
            'belum_pulang_hari_ini': belum_pulang_hari_ini,
            'tk_kemarin_presensi_hari_ini': tk_kemarin_presensi_hari_ini,
            
            # Analytics
            'instalasi_alpa': instalasi_alpa,
            'instalasi_hadir': instalasi_hadir,
        })

        return context


# ==========================================
# VIEW EXPORT EXCEL (SUPERUSER ONLY)
# ==========================================
class ExportAbsensiHarianExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    
    def test_func(self):
        # Hanya Superuser/Admin Manajemen yang bisa mengakses unduhan ini
        return self.request.user.is_superuser

    def get(self, request, *args, **kwargs):
        today = date.today()
        yesterday = today - timedelta(days=1)
        tk_kemarin_qs = (
            pegawai_tk_pada_tanggal(yesterday)
            .prefetch_related(
                Prefetch(
                    'riwayat_penempatan',
                    queryset=RiwayatPenempatan.objects.filter(status=True).select_related(
                        'penempatan_level1',
                        'penempatan_level2',
                        'penempatan_level4',
                    ),
                    to_attr='sk_aktif',
                )
            )
            .order_by('first_name', 'last_name')
        )
        presensi_hari_ini_qs = presensi_datang_pada_tanggal(today).order_by(
            'jam_datang', 'pegawai__first_name', 'pegawai__last_name'
        )
        
        # =========================================================================
        # INASIALISASI & DEKORASI WORKBOOK
        # =========================================================================
        wb = Workbook()
        
        # Styles Template
        font_title = Font(name='Arial', size=14, bold=True)
        font_subtitle = Font(name='Arial', size=10, italic=True)
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        font_data = Font(name='Arial', size=10)
        
        fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Navy Corporate
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center')
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        # ==========================================
        # SHEET 1: SELURUH PEGAWAI TK KEMARIN
        # ==========================================
        ws1 = wb.active
        ws1.title = "TK Kemarin"
        ws1.views.sheetView[0].showGridLines = True
        
        # Header Dokumen Sheet 1
        ws1['A1'] = "DAFTAR PEGAWAI TK / ALPA KEMARIN"
        ws1['A1'].font = font_title
        ws1['A2'] = f"Tanggal evaluasi: {yesterday.strftime('%d-%m-%Y')}"
        ws1['A2'].font = font_subtitle
        
        # Header Struktur Tabel Sheet 1
        headers_ws1 = [
            'No', 'Nama Pegawai', 'NIP', 'Unit Organisasi', 'Bidang',
            'Instalasi', 'Keterangan',
        ]
        ws1.append([]) # Spacer baris 3
        ws1.append(headers_ws1) # Baris 4
        
        for cell in ws1[4]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        ws1.row_dimensions[4].height = 26

        for idx, pegawai in enumerate(tk_kemarin_qs, 1):
            profil_user = getattr(pegawai, 'profil_user', None)
            nip = getattr(profil_user, 'nip', '-') if profil_user else '-'
            sk = pegawai.sk_aktif[0] if pegawai.sk_aktif else None
            
            row_values = [
                idx,
                pegawai.full_name,
                nip,
                sk.penempatan_level1.unor if sk and sk.penempatan_level1 else '-',
                sk.penempatan_level2.bidang if sk and sk.penempatan_level2 else '-',
                sk.penempatan_level4.instalasi if sk and sk.penempatan_level4 else '-',
                'Tidak memiliki kehadiran, izin, dinas, atau libur yang sah',
            ]
            ws1.append(row_values)
            
            current_row = ws1.max_row
            ws1.row_dimensions[current_row].height = 20
            for col_idx, cell in enumerate(ws1[current_row], start=1):
                cell.font = font_data
                cell.border = border_thin
                cell.alignment = align_center if col_idx in [1, 3] else align_left

        # ==========================================
        # SHEET 2: SELURUH PEGAWAI SUDAH PRESENSI HARI INI
        # ==========================================
        ws2 = wb.create_sheet(title="Presensi Hari Ini")
        ws2.views.sheetView[0].showGridLines = True
        ws2['A1'] = "DAFTAR PEGAWAI SUDAH PRESENSI HARI INI"
        ws2['A1'].font = font_title
        ws2['A2'] = f"Tanggal presensi: {today.strftime('%d-%m-%Y')}"
        ws2['A2'].font = font_subtitle

        headers_ws2 = [
            'No', 'Nama Pegawai', 'NIP', 'Unit Organisasi', 'Bidang',
            'Instalasi', 'Jam Datang', 'Status Datang', 'Jam Pulang',
        ]
        ws2.append([])
        ws2.append(headers_ws2)
        for cell in ws2[4]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        ws2.row_dimensions[4].height = 26

        for idx, absen in enumerate(presensi_hari_ini_qs, 1):
            profil_user = getattr(absen.pegawai, 'profil_user', None)
            nip = getattr(profil_user, 'nip', '-') if profil_user else '-'
            ws2.append([
                idx,
                absen.pegawai.full_name,
                nip,
                absen.unor.unor if absen.unor else '-',
                absen.bidang.bidang if absen.bidang else '-',
                absen.instalasi.instalasi if absen.instalasi else '-',
                absen.jam_datang.strftime('%H:%M:%S') if absen.jam_datang else '-',
                absen.status_datang or '-',
                absen.jam_pulang.strftime('%H:%M:%S') if absen.jam_pulang else 'Belum pulang',
            ])
            current_row = ws2.max_row
            ws2.row_dimensions[current_row].height = 20
            for col_idx, cell in enumerate(ws2[current_row], start=1):
                cell.font = font_data
                cell.border = border_thin
                cell.alignment = align_center if col_idx in [1, 3, 7, 8, 9] else align_left

        # ==========================================
        # KOREKSI OTOMATIS: AUTO-FIT COLUMN WIDTH
        # ==========================================
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row < 4: # Biarkan cell judul dokumen dilewati agar kalkulasi kolom presisi
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Response Output
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Monitoring_TK_dan_Presensi_{today.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
